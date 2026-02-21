#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesador de asistencias (HikCentral / iVMS)
- Lee un Excel exportado (HikCentral/iVMS) que contiene una columna "Registro" con horas del día.
- Interpreta los primeros N registros (2 a 6) como:
    1) Entrada
    2) Salida a comer
    3) Regreso de comer
    4) Salida a cenar
    5) Regreso de cenar
    6) Salida
- Calcula Horas trabajadas (restando comida con tope configurable) y Horas extra (umbral 8:00 con redondeo).
- Genera 2 archivos:
    1) _PROCESADO.xlsx  (ID original, ordenado por el orden de grupos)
    2) _IDGRUPO.xlsx    (solo IDGRUPO con prefijo/grupo, ordenado por el orden de grupos)
- Modo Administración (opcional) para mantener el mapeo Empleado -> Grupo -> IDGRUPO en un JSON.
Formato de horas: HH:MM (sin segundos).
"""
from __future__ import annotations
import argparse
import json
import copy
import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta, time, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import pandas as pd


def _read_csv_flexible(path: Path) -> pd.DataFrame:
    """Lee CSV/TXT de forma tolerante a encodings comunes.
    - Intenta utf-8-sig, utf-8, latin1.
    - Usa sep=None (sniff) con engine=python para tolerar delimitadores.
    """
    last_err: Exception | None = None
    for enc in ("utf-8-sig", "utf-8", "latin1"):
        try:
            return pd.read_csv(path, dtype=str, encoding=enc, sep=None, engine="python")
        except Exception as e:
            last_err = e
            continue
    # Si todos fallan, re-lanzar el último error para diagnóstico
    assert last_err is not None
    raise last_err

# Editor/auditoría delegados (migración)
from .corrections import AuditEntry, editar_checadas_interactivo, guardar_auditoria_json
from .io import _drop_export_debug_cols

# ---------------------------
# Compat helpers (robust parsing)
# ---------------------------
def normalize_id(v: object, width: int = 3) -> str:
    """Compatibilidad: normaliza ID preservando ceros."""
    return _coerce_id_str(v, width)
def parse_date(v: object) -> Optional[date]:
    """Parsea fecha desde str/datetime/date. Retorna date o None."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    dt = pd.to_datetime(s, errors="coerce", dayfirst=False)
    if pd.isna(dt):
        return None
    return dt.date()
def parse_time(v: object) -> Optional[time]:
    """Parsea hora HH:MM (o HH:MM:SS) desde str/datetime/time. Retorna time o None."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, time):
        return v.replace(second=0, microsecond=0)
    if isinstance(v, datetime):
        return v.time().replace(second=0, microsecond=0)
    s = str(v).strip()
    if not s:
        return None
    # aceptar 'HH:MM' o 'HH:MM:SS'
    m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", s)

    if not m:
        return None
    hh = int(m.group(1)); mm = int(m.group(2))
    if hh>23 or mm>59:
        return None
    return time(hour=hh, minute=mm)
def _safe_input(prompt: str, default: str = "") -> str:
    """input() seguro: en EOF/CTRL+C regresa default."""
    try:
        return input(prompt)
    except (EOFError, KeyboardInterrupt):
        return default
# ---------------------------
# Utilidades básicas
# ---------------------------
def _norm(s: str) -> str:
    """Normaliza texto: minúsculas, sin acentos, sin espacios extra."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    # quitar acentos
    tr = str.maketrans("áéíóúüñ", "aeiouun")
    s = s.translate(tr)
    # compactar separadores
    s = re.sub(r"[\s\-_]+", " ", s)
    return s.strip()
def _is_digits(s: str) -> bool:
    return bool(re.fullmatch(r"\d+", str(s).strip()))
def make_emp_key(id_val: object, nombre_val: object, cfg_id_width: int = 3) -> Tuple[str, str]:
    """Crea una clave interna estable por empleado y el ID visible.
    Caso especial SIN ID:
      - En el export, a veces ID == Nombre (ej. 'María') y no es numérico.
      - Internamente usamos clave: 'NOMBRE::' para no perder trazabilidad.
      - En reportes visibles, el ID se deja en blanco (para no confundir).
    Retorna: (emp_key, id_display)
    """
    id_raw = _coerce_id_str(id_val, cfg_id_width)
    nombre_raw = "" if nombre_val is None else str(nombre_val).strip()
    if id_raw and nombre_raw and (_norm(id_raw) == _norm(nombre_raw)) and (not _is_digits(id_raw)):
        return f"NOMBRE::{_norm(nombre_raw)}", ""  # ID visible vacío
    # Si el ID no es numérico y parece nombre (contiene letras), úsalo como clave tipo NOMBRE::
    # Esto ayuda a que correcciones/plantilla coincidan si se captura 'María' como ID.
    if id_raw and (not _is_digits(id_raw)) and re.search(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]", id_raw):
        return f"NOMBRE::{_norm(id_raw)}", ""
    return id_raw, id_raw
def apply_id_display(df: Optional[pd.DataFrame], id_col: str = "ID") -> Optional[pd.DataFrame]:
    """Devuelve copia con ID visible en blanco para claves NOMBRE::"""
    if df is None:
        return None
    if id_col not in df.columns:
        return df
    out = df.copy()
    try:
        out[id_col] = out[id_col].astype(str).map(lambda x: "" if str(x).startswith("NOMBRE::") else str(x))
    except Exception:
        pass
    return out
def _guess_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """Devuelve el nombre de la columna del DF que matchea alguno de los candidatos (normalizados)."""
    norm_map = {_norm(c): c for c in df.columns}
    for cand in candidates:
        key = _norm(cand)
        if key in norm_map:

            return norm_map[key]
    return None
def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)
def fmt_hhmm(t: Optional[time]) -> str:
    return "" if t is None else t.strftime("%H:%M")
def hhmm_to_minutes(s: str) -> int:
    """Convierte 'HH:MM' a minutos. Soporta vacío/None."""
    if s is None:
        return 0
    s = str(s).strip()
    if not s or s.lower() in {"nan", "none"}:
        return 0
    m = re.match(r"^(\d+):(\d{2})$", s)
    if not m:
        return 0
    return int(m.group(1)) * 60 + int(m.group(2))
def minutes_to_hhmm(total_min: int) -> str:
    if total_min <= 0:
        return "00:00"
    h = total_min // 60
    m = total_min % 60
    return f"{h:02d}:{m:02d}"
def add_minutes(t: Optional[time], minutes: int) -> Optional[time]:
    if t is None:
        return None
    dt = datetime(2000,1,1,t.hour,t.minute) + timedelta(minutes=minutes)
    return time(dt.hour, dt.minute)
def round_minutes(value_min: int, step: int, mode: str) -> int:
    """
    Redondea minutos a múltiplos de 'step'.
    mode:
      - 'up'      (hacia arriba)
      - 'down'    (hacia abajo)
      - 'nearest' (al más cercano)
    """
    if step <= 1:
        return max(0, value_min)
    v = max(0, value_min)
    if mode == "up":
        return ((v + step - 1) // step) * step
    if mode == "down":
        return (v // step) * step
    # nearest
    lo = (v // step) * step
    hi = lo + step
    return hi if (v - lo) >= (hi - v) else lo
# ---------------------------
# Configuración / Mapeos
# ---------------------------
@dataclass
class AppConfig:
    # Orden de grupos (como los creas/enumeras)
    grupos_orden: List[str] = field(default_factory=lambda: ["000"])
    # Metadatos (prefijo visible en IDGRUPO)
    grupos_meta: Dict[str, Dict[str, str]] = field(default_factory=lambda: {"000": {"prefijo": "000"}})
    # Asignación por empleado (clave: ID original)
    empleado_a_grupo: Dict[str, str] = field(default_factory=dict)
    # IDGRUPO manual por empleado (clave: ID original)
    empleado_a_idgrupo: Dict[str, str] = field(default_factory=dict)

    # Semana personalizada (0=Lun ... 6=Dom). Default: 0 (Lunes, ISO)
    week_start_dow: int = 2
    # Reglas de cálculo
    umbral_extra_min: int = 480  # 8:00
    redondeo_extra_step_min: int = 1  # 5/10/15
    redondeo_extra_modo: str = "none"  # up/down/nearest
    # Descuento comida: se resta el mínimo entre (duración real) y este tope
    tope_descuento_comida_min: int = 30
    # Texto de auditoría (Notas)
    nota_nolab_solape_cd: str = "NoLaborado solapado con comida/cena: fusionado"
    nota_nolab_fuera_jornada: str = "NoLaborado fuera de jornada ignorado"
    # Formato ID: si el ID es numérico y viene como "3", lo rellena a este ancho (003)
    id_min_width: int = 3
    def prefijo_de_grupo(self, grupo: str) -> str:
        return (self.grupos_meta.get(grupo, {}) or {}).get("prefijo", grupo)
def _config_path(script_dir: Path) -> Path:
    return script_dir / "mapa_grupos.json"
def cargar_config(script_dir: Path) -> AppConfig:
    path = _config_path(script_dir)
    if not path.exists():
        cfg = AppConfig()
        guardar_config(script_dir, cfg)
        return cfg
    data = json.loads(path.read_text(encoding="utf-8"))
    cfg = AppConfig()
    cfg.grupos_orden = data.get("grupos_orden", cfg.grupos_orden)
    cfg.grupos_meta = data.get("grupos_meta", cfg.grupos_meta)
    cfg.empleado_a_grupo = data.get("empleado_a_grupo", {})
    cfg.empleado_a_idgrupo = data.get("empleado_a_idgrupo", {})
    reglas = data.get("reglas", {}) or {}
    cfg.umbral_extra_min = int(reglas.get("umbral_extra_min", cfg.umbral_extra_min))
    cfg.redondeo_extra_step_min = int(reglas.get("redondeo_extra_step_min", cfg.redondeo_extra_step_min))
    cfg.redondeo_extra_modo = str(reglas.get("redondeo_extra_modo", cfg.redondeo_extra_modo))
    cfg.tope_descuento_comida_min = int(reglas.get("tope_descuento_comida_min", cfg.tope_descuento_comida_min))
    cfg.id_min_width = int(reglas.get("id_min_width", cfg.id_min_width))
    cfg.week_start_dow = int(reglas.get("week_start_dow", getattr(cfg, "week_start_dow", 0)) or 0)
    return cfg
def guardar_config(script_dir: Path, cfg: AppConfig) -> None:
    path = _config_path(script_dir)
    data = {
        "grupos_orden": cfg.grupos_orden,
        "grupos_meta": cfg.grupos_meta,
        "empleado_a_grupo": cfg.empleado_a_grupo,
        "empleado_a_idgrupo": cfg.empleado_a_idgrupo,
        "reglas": {
            "umbral_extra_min": cfg.umbral_extra_min,
            "redondeo_extra_step_min": cfg.redondeo_extra_step_min,
            "redondeo_extra_modo": cfg.redondeo_extra_modo,
            "tope_descuento_comida_min": cfg.tope_descuento_comida_min,
            "id_min_width": cfg.id_min_width,
            "week_start_dow": getattr(cfg, "week_start_dow", 0),

        },
    }
    try:
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except PermissionError:
        print(f"[ADVERTENCIA] No se pudo guardar configuración en {path} (permiso denegado). Se continuará sin guardar.")
    except Exception as e:
        print(f"[ADVERTENCIA] No se pudo guardar configuración en {path}: {e}. Se continuará sin guardar.")
# ---------------------------
# Auditoría de cambios (checadas)
# ---------------------------

# ---------------------------
# Auditoría/Editor: movidos a procesador.corrections (migración v19)
# ---------------------------

def parse_registro(reg: object) -> List[time]:
    """
    Acepta "10:44; 11:28; 11:59" o similares.
    Importante (turnos que cruzan medianoche):
    - NO se ordena por reloj (HH:MM), porque eso invierte turnos nocturnos (ej. 20:48 ... 03:05).
    - Se conserva el orden tal como viene en el 'Registro' exportado.
    - Se eliminan duplicados exactos (HH:MM) respetando la primera aparición.
    Devuelve lista en el orden original (sin segundos).
    """
    if reg is None:
        return []
    s = str(reg).strip()
    if not s:
        return []
    # extraer HH:MM (ignora segundos si vienen)
    parts = re.findall(r"\b(\d{1,2}):(\d{2})(?::(\d{2}))?\b", s)
    out: List[time] = []
    seen = set()
    for hh, mm, _ss in parts:
        h = int(hh)
        m = int(mm)
        if 0 <= h <= 23 and 0 <= m <= 59:
            t = time(hour=h, minute=m)
            key = t.strftime("%H:%M")
            if key in seen:
                continue
            seen.add(key)
            out.append(t)
    return out
def normalize_registro_times(times: List[time]) -> Tuple[List[time], bool]:
    """Normaliza y ordena checadas para cálculos consistentes.
    - Mantiene turnos nocturnos: si la entrada es tarde y hay horas menores, se asumen del día siguiente.
    - Reordena horas fuera de orden (por captura/pegado) para evitar duraciones 00:00 o negativas.
    Devuelve (lista_normalizada, reordenado_bool).
    """
    if not times or len(times) < 2:
        return times, False
    mins = [t.hour * 60 + t.minute for t in times]
    entry_min = mins[0]
    has_smaller = any(m < entry_min for m in mins[1:])
    span = max(mins) - min(mins)
    # Heurística de cruce de medianoche:
    # - Entrada tarde (>=18:00) y hay horas menores -> probable cruce
    # - O hay horas menores y el rango del día es muy amplio -> probable cruce
    wrap_likely = (entry_min >= 18 * 60 and has_smaller) or (has_smaller and span > 12 * 60)
    adjusted = []
    for idx, (t, m) in enumerate(zip(times, mins)):
        adj = m + (1440 if (wrap_likely and m < entry_min) else 0)
        adjusted.append((adj, idx, t))
    ordered = sorted(adjusted, key=lambda x: (x[0], x[1]))
    out = [t for _adj, _idx, t in ordered]

    return out, (out != times)
# ---------------------------
# Correcciones manuales (opcional)
# ---------------------------
CORR_EVENTOS = ["Entrada", "Salida a comer", "Regreso de comer", "Salida a cenar", "Regreso de cenar", "Salida"]
def _normalizar_evento(ev: object) -> str:
    s = str(ev).strip()
    # normaliza mayúsculas/minúsculas y dobles espacios
    s = re.sub(r"\s+", " ", s).strip()
    # formas cortas/variantes comunes
    alias = {
        "ENTRADA": "Entrada",
        "SALIDA": "Salida",
        "SALIDA A COMER": "Salida a comer",
        "REGRESO DE COMER": "Regreso de comer",
        "SALIDA A CENAR": "Salida a cenar",
        "REGRESO DE CENAR": "Regreso de cenar",
    }
    key = s.upper()
    return alias.get(key, s)
def cargar_correcciones(path_corr: Optional[Path]) -> Tuple[
    Dict[Tuple[str, date], Dict[str, Tuple[Optional[time], str]]],
    Dict[Tuple[str, date], List[Tuple[Optional[time], Optional[time], str]]],
]:
    """
    Lee un archivo de correcciones manuales (control de RRHH).
    Soporta 2 bloques (en el mismo .xlsx):
      1) Hoja 'Ajustes' (o la primera hoja si no existe):
         - ID
         - Fecha (YYYY-MM-DD o fecha Excel)
         - Evento (uno de CORR_EVENTOS)
         - Hora (HH:MM)  -> si pones '-' o 'BORRAR' se borra ese evento
         - Nota (opcional)
         Devuelve: (id, fecha) -> {evento: (hora|None, nota)}
      2) Hoja 'NoLaborado' (opcional):
         - ID
         - Fecha
         - Inicio (HH:MM)
         - Fin (HH:MM)  -> opcional; si falta, el cálculo asume Fin = Salida final
         - Nota (opcional)
         Devuelve: (id, fecha) -> [(inicio, fin, nota), ...]
    Si el archivo no existe, devuelve estructuras vacías.
    """
    corr_eventos: Dict[Tuple[str, date], Dict[str, Tuple[Optional[time], str]]] = {}
    corr_nolabor: Dict[Tuple[str, date], List[Tuple[Optional[time], Optional[time], str]]] = {}
    if not path_corr or not path_corr.exists():
        return corr_eventos, corr_nolabor
    def _read_any(path: Path) -> pd.ExcelFile:
        return pd.ExcelFile(path)
    def _norm_cols(df: pd.DataFrame) -> Dict[str, str]:
        return {str(c).replace('\ufeff','').strip().lower(): c for c in df.columns}

    def _pick(cols_map: Dict[str, str], *names: str) -> Optional[str]:
        for n in names:
            key = n.strip().lower()
            if key in cols_map:
                return cols_map[key]
        return None
    def _parse_date(v) -> Optional[date]:
        d = parse_date(v)
        return d
    def _parse_time(v) -> Optional[time]:
        return parse_time(v)
    if path_corr.suffix.lower() in [".xlsx", ".xls"]:
        xls = _read_any(path_corr)
        sheet_ajustes = "Ajustes" if "Ajustes" in xls.sheet_names else xls.sheet_names[0]
        dfA = xls.parse(sheet_ajustes)
        # --- Ajustes
        cols = _norm_cols(dfA)
        c_id = _pick(cols, "id", "id empleado", "empleado id")
        c_fecha = _pick(cols, "fecha", "date")
        c_evento = _pick(cols, "evento", "event", "tipo")
        c_hora = _pick(cols, "hora", "time")
        c_nota = _pick(cols, "nota", "notas", "comentario")
        if c_id and c_fecha and c_evento and c_hora:
            for _, r in dfA.iterrows():
                emp = make_emp_key(r.get(c_id), r.get(c_id), 3)[0]
                d = _parse_date(r.get(c_fecha))
                ev = str(r.get(c_evento) or "").strip()
                hh = r.get(c_hora)
                nota = str(r.get(c_nota) or "").strip() if c_nota else ""
                if not emp or not d or not ev:
                    continue
                if ev not in CORR_EVENTOS:
                    continue
                hv = str(hh).strip() if not pd.isna(hh) else ""
                if hv.lower() in ["-", "borrar", "delete", ""]:
                    t = None
                else:
                    t = _parse_time(hh)
                corr_eventos.setdefault((emp, d), {})[ev] = (t, nota)
        # --- NoLaborado (opcional)
        if "NoLaborado" in xls.sheet_names:
            dfN = xls.parse("NoLaborado")
            colsN = _norm_cols(dfN)
            c_id = _pick(colsN, "id", "id empleado", "empleado id")
            c_fecha = _pick(colsN, "fecha", "date")
            c_ini = _pick(colsN, "inicio", "start", "desde")
            c_fin = _pick(colsN, "fin", "end", "hasta")
            c_nota = _pick(colsN, "nota", "notas", "comentario")
            if c_id and c_fecha and c_ini:
                for _, r in dfN.iterrows():
                    emp = make_emp_key(r.get(c_id), r.get(c_id), 3)[0]
                    d = _parse_date(r.get(c_fecha))
                    ini = _parse_time(r.get(c_ini))
                    fin_raw = r.get(c_fin) if c_fin else None
                    fin = _parse_time(fin_raw) if (c_fin and not (pd.isna(fin_raw) or str(fin_raw).strip() in ["", "-"])) else None
                    nota = str(r.get(c_nota) or "").strip() if c_nota else ""
                    if not emp or not d or not ini:
                        continue
                    corr_nolabor.setdefault((emp, d), []).append((ini, fin, nota))
    else:
        # CSV: solo Ajustes
        dfA = pd.read_csv(path_corr)
        dfA.columns = [str(c).replace('\ufeff','').strip() for c in dfA.columns]
        cols = _norm_cols(dfA)
        c_id = _pick(cols, "id")
        c_fecha = _pick(cols, "fecha")
        c_evento = _pick(cols, "evento")
        c_hora = _pick(cols, "hora")
        c_nota = _pick(cols, "nota", "notas")

        if c_id and c_fecha and c_evento and c_hora:
            for _, r in dfA.iterrows():
                emp = make_emp_key(r.get(c_id), r.get(c_id), 3)[0]
                d = _parse_date(r.get(c_fecha))
                ev = str(r.get(c_evento) or "").strip()
                hh = r.get(c_hora)
                nota = str(r.get(c_nota) or "").strip() if c_nota else ""
                if not emp or not d or not ev:
                    continue
                if ev not in CORR_EVENTOS:
                    continue
                hv = str(hh).strip() if not pd.isna(hh) else ""
                t = None if hv.lower() in ["-", "borrar", "delete", ""] else _parse_time(hh)
                corr_eventos.setdefault((emp, d), {})[ev] = (t, nota)
    return corr_eventos, corr_nolabor
def aplicar_correcciones(row: Dict[str, object],
                         correcciones: Dict[Tuple[str, date], Dict[str, Tuple[Optional[time], str]]],
                         id_col: str,
                         fecha_col: str) -> Tuple[Dict[str, object], bool, str]:
    """
    Aplica correcciones sobre un dict de salida (row). Devuelve:
      - row corregido
      - flag de si se aplicó algo
      - nota concatenada
    """
    rid = str(row.get(id_col, "")).strip()
    f = row.get(fecha_col)
    if not rid or not isinstance(f, date):
        return row, False, ""
    key = (rid, f)
    if key not in correcciones:
        return row, False, ""
    aplicado = False
    notas = []
    for ev, (hh, nota) in correcciones[key].items():
        if ev not in row:
            # si la columna no existe, la crea (por seguridad)
            row[ev] = ""
        old = row.get(ev, "")
        if hh is None:
            # borrar
            if old != "":
                row[ev] = ""
                aplicado = True
        else:
            new = fmt_hhmm(hh)
            if str(old).strip() != new:
                row[ev] = new
                aplicado = True
        if nota:
            notas.append(f"{ev}: {nota}")
    return row, aplicado, " | ".join(notas)
def map_eventos(times: List[time]) -> Dict[str, Optional[time]]:
    """
    Mapea checadas a eventos con la regla operativa solicitada:
      - 1 checada: Entrada

      - 2..N checadas: Primera = Entrada, Última = Salida
      - Intermedias (en orden):
          2ª = Salida a comer
          3ª = Regreso de comer
          4ª = Salida a cenar
          5ª = Regreso de cenar
    Si hay más de 6 checadas, se usan únicamente las 4 intermedias más tempranas
    y se conserva la última como Salida.
    """
    labels = ["Entrada", "Salida a comer", "Regreso de comer", "Salida a cenar", "Regreso de cenar", "Salida"]
    out = {k: None for k in labels}
    n = len(times)
    out["_extra_registros"] = max(0, n - 6)
    if n == 0:
        return out
    if n == 1:
        out["Entrada"] = times[0]
        return out
    # Siempre: primera = Entrada, última = Salida
    out["Entrada"] = times[0]
    out["Salida"] = times[-1]
    # Intermedias
    middle = times[1:-1]
    middle_labels = ["Salida a comer", "Regreso de comer", "Salida a cenar", "Regreso de cenar"]
    for lab, t in zip(middle_labels, middle[:4]):
        out[lab] = t
    return out
def minutos_entre(t1: Optional[time], t2: Optional[time]) -> int:
    if t1 is None or t2 is None:
        return 0
    dt1 = datetime(2000, 1, 1, t1.hour, t1.minute)
    dt2 = datetime(2000, 1, 1, t2.hour, t2.minute)
    if dt2 < dt1:
        # cruce de medianoche (muy raro para checadas del mismo día, pero por seguridad)
        dt2 += timedelta(days=1)
    return int((dt2 - dt1).total_seconds() // 60)

def calcular_descuento_comida(eventos: Dict[str, Optional[time]], cfg: AppConfig) -> Tuple[int, int, str]:
    """Calcula duración real y descuento aplicable para comida.

    Regla:
      - Si duración real <= cfg.umbral_comida_media_hora_min (default 60): descuenta min(duración_real, cfg.tope_descuento_comida_min) (default 30)
        (pero nunca descuenta más que la duración real).
      - Si duración real > umbral: descuenta duración real completa.
    Casos incompletos:
      - Si hay 'Salida a comer' pero no 'Regreso de comer' y existe 'Salida' final, se asume fin = Salida.
    Devuelve: (dur_real_min, desc_min, motivo_str)
    """
    sal = eventos.get("Salida")
    sal_com = eventos.get("Salida a comer")
    reg_com = eventos.get("Regreso de comer")

    if not sal_com:
        return 0, 0, "sin comida"
    fin = reg_com or sal
    if not fin:
        # no hay fin estimable
        return 0, 0, "comida incompleta (sin fin)"
    dur = minutos_entre(sal_com, fin)
    if dur <= 0:
        return 0, 0, "sin comida"
    umbral = int(getattr(cfg, "umbral_comida_media_hora_min", 60) or 60)
    tope = int(getattr(cfg, "tope_descuento_comida_min", 30) or 30)
    if dur <= umbral:
        desc = min(dur, tope)
        if desc == dur and dur < tope:
            return dur, desc, f"<= {umbral} min (real < {tope})"
        return dur, desc, f"<= {umbral} min (tope {tope})"
    # excede umbral: descuento completo
    return dur, dur, f"> {umbral} min (completo)"


def calcular_descuento_cena(eventos: Dict[str, Optional[time]], cfg: AppConfig) -> Tuple[int, int, str]:
    """Calcula duración real y descuento aplicable para cena (siempre real completo).

    Casos incompletos:
      - Si hay 'Salida a cenar' pero no 'Regreso de cenar' y existe 'Salida' final, se asume fin = Salida.
    Devuelve: (dur_real_min, desc_min, motivo_str)
    """
    sal = eventos.get("Salida")
    sal_cen = eventos.get("Salida a cenar")
    reg_cen = eventos.get("Regreso de cenar")

    if not sal_cen:
        return 0, 0, "sin cena"
    fin = reg_cen or sal
    if not fin:
        return 0, 0, "cena incompleta (sin fin)"
    dur = minutos_entre(sal_cen, fin)
    if dur <= 0:
        return 0, 0, "sin cena"
    return dur, dur, "completo"


def calcular_trabajado(eventos: Dict[str, Optional[time]], cfg: AppConfig, no_laborado_extra: Optional[List[Tuple[Optional[time], Optional[time], str]]] = None) -> Tuple[int, int, int, int, int, int, int, int]:
    """
    Devuelve (minutos_trabajados, minutos_extra_redondeados).
    Enfoque (robusto y alineado a RRHH):
      1) Base = minutos totales entre Entrada y Salida (cruza medianoche si aplica).
      2) Restas por tiempo NO laborado:
         - Comida:
             * Si hay Salida a comer y Regreso de comer ->
                 - si duración_real <= umbral_comida_media_hora_min: resta min(duración_real, tope_descuento_comida_min)
                 - si duración_real > umbral: resta duración_real completa
             * Si falta el regreso (pero hay salida final) -> asume fin = Salida y aplica el mismo tope
         - Cena:
             * Si hay Salida a cenar y Regreso de cenar -> resta duración_real
             * Si falta el regreso (pero hay salida final) -> asume fin = Salida y resta duración_real
         - Salidas extraordinarias (por inconveniente):
             * Intervalos extra (inicio, fin). Si falta fin y hay salida final -> asume fin = Salida.
             * Se descuenta duración real (sin tope).
      3) Umbral de extra: a partir de cfg.umbral_extra_min (8:00) y redondeo según cfg.
    Nota: si falta Entrada o Salida, devuelve (0,0).
    """
    ent = eventos.get("Entrada")
    sal = eventos.get("Salida")
    if not ent or not sal:
        return 0, 0, 0, 0, 0, 0, 0, 0
    total = minutos_entre(ent, sal)
    # --- Comida (regla: 30 min solo si <= umbral; si excede, descuento completo)
    sal_com = eventos.get("Salida a comer")
    reg_com = eventos.get("Regreso de comer")
    comida_real, comida_ded, _comida_motivo = calcular_descuento_comida(eventos, cfg)

    # --- Cena (real completo)
    sal_cen = eventos.get("Salida a cenar")
    reg_cen = eventos.get("Regreso de cenar")
    cena_real, cena_ded, _cena_motivo = calcular_descuento_cena(eventos, cfg)

    # --- Salidas extraordinarias (real) con validación de solapes
    extra_ded = 0
    nolab_overlap_cd = 0   # solape NoLaborado con comida/cena (no se duplica)
    nolab_solape_interno = 0  # solape entre intervalos NoLaborado (se fusionan)
    ignored_outside_shift = 0  # minutos de NoLaborado capturados fuera de la jornada
    # Construir ventanas "ya descontadas" para evitar doble descuento
    ventanas_descuento = []  # lista de (ini, fin)
    # Ventana de comida: depende del descuento aplicado (tope vs completo)
    if sal_com and comida_ded > 0:
        # fin real estimable
        fin_real = reg_com if reg_com is not None else sal
        if fin_real:
            dur_real = minutos_entre(sal_com, fin_real)
            if dur_real > 0 and comida_ded >= dur_real:
                # descuento completo = ventana real
                ventanas_descuento.append((sal_com, fin_real))
            else:
                # descuento por tope = ventana [inicio, inicio+descuento]
                fin_com = add_minutes(sal_com, comida_ded)
                if fin_com:
                    # acotar a fin real si por algún motivo se pasa
                    if dur_real > 0 and minutos_entre(sal_com, fin_com) > dur_real:
                        fin_com = fin_real
                    ventanas_descuento.append((sal_com, fin_com))

    # Ventana de cena: siempre real completo
    if sal_cen and cena_ded > 0:
        fin_cen = reg_cen if reg_cen is not None else sal
        if fin_cen and minutos_entre(sal_cen, fin_cen) > 0:
            ventanas_descuento.append((sal_cen, fin_cen))

    # Helpers de línea de tiempo relativa a la jornada (soporta cruce de medianoche)
    crosses_midnight = sal < ent  # si la salida es "menor" que la entrada, cruza al día siguiente
    def _to_shift_dt(t: time) -> datetime:
        """Convierte una hora a datetime relativo a la jornada.
        En turnos nocturnos, horas menores que Entrada se consideran del día siguiente."""
        base = datetime(2000, 1, 1, t.hour, t.minute)
        if crosses_midnight and t < ent:
            base += timedelta(days=1)
        return base
    def _window(ent_t: time, sal_t: time) -> Tuple[datetime, datetime]:
        a = _to_shift_dt(ent_t)
        b = _to_shift_dt(sal_t)
        if b < a:
            b += timedelta(days=1)
        return a, b
    def _overlap_min(a1: time, a2: time, b1: time, b2: time) -> int:
        """Minutos de solape entre intervalos (a1,a2) y (b1,b2) en la misma línea temporal de jornada."""
        if not a1 or not a2 or not b1 or not b2:
            return 0
        A1, A2 = _to_shift_dt(a1), _to_shift_dt(a2)
        B1, B2 = _to_shift_dt(b1), _to_shift_dt(b2)
        if A2 < A1:
            A2 += timedelta(days=1)
        if B2 < B1:
            B2 += timedelta(days=1)
        s = max(A1, B1)

        e = min(A2, B2)
        if e <= s:
            return 0
        return int((e - s).total_seconds() // 60)
    def _clip_to_shift(ini_t: time, fin_t: time, ent_t: time, sal_t: time) -> Optional[Tuple[time, time, int]]:
        """Recorta un intervalo (ini, fin) a la ventana [Entrada, Salida] en la línea temporal de la jornada.
        Devuelve (ini_recortado, fin_recortado, minutos_ignorados_fuera_de_jornada)."""
        s0, s1 = _window(ent_t, sal_t)
        a0 = _to_shift_dt(ini_t)
        a1 = _to_shift_dt(fin_t)
        if a1 < a0:
            a1 += timedelta(days=1)
        i0 = max(a0, s0)
        i1 = min(a1, s1)
        if i1 <= i0:
            # todo fuera
            ignorados = int((a1 - a0).total_seconds() // 60)
            return None
        inside = int((i1 - i0).total_seconds() // 60)
        total_i = int((a1 - a0).total_seconds() // 60)
        ignorados = max(0, total_i - inside)
        return time(i0.hour, i0.minute), time(i1.hour, i1.minute), ignorados
    if no_laborado_extra:
        # Normalizar: convertir a lista de intervalos efectivos y ordenar por inicio
        intervals = []
        ignored_outside_shift = 0
        for ini, fin, _nota in no_laborado_extra:
            if ini is None:
                continue
            fin_eff = fin if fin is not None else sal
            if fin_eff is None:
                continue
            # Recortar a la jornada [Entrada, Salida] para evitar descuentos fuera de turno
            clipped = _clip_to_shift(ini, fin_eff, ent, sal)
            if clipped is None:
                # Intervalo completamente fuera de la jornada (se ignora pero se contabiliza como advertencia)
                try:
                    ignored_outside_shift += minutos_entre(ini, fin_eff)
                except Exception:
                    pass
                continue
            ini_c, fin_c, ign = clipped
            ignored_outside_shift += ign
            intervals.append((ini_c, fin_c))
        # Ordenar por inicio (considerando hora)
        # Ordenar por posición relativa desde Entrada (maneja turnos nocturnos)
        intervals.sort(key=lambda x: minutos_entre(ent, x[0]))
        # Fusionar solapes internos
        merged = []
        for ini, fin in intervals:
            if not merged:
                merged.append([ini, fin])
                continue
            last_ini, last_fin = merged[-1]
            # Si solapan o se enciman
            if _overlap_min(last_ini, last_fin, ini, fin) > 0 or (last_fin == ini):
                # calcular solape interno aproximado
                nolab_solape_interno += _overlap_min(last_ini, last_fin, ini, fin)
                # extender fin si es necesario
                # comparar fin en dt
                def _to_dt(t):
                    d = datetime(2000,1,1,t.hour,t.minute)
                    return d
                lf = _to_dt(last_fin); cf = _to_dt(fin)
                li = _to_dt(last_ini); ci = _to_dt(ini)
                if lf < li: lf += timedelta(days=1)
                if cf < ci: cf += timedelta(days=1)
                # mantener inicio más temprano (last_ini) y fin más tardío
                if cf > lf:
                    merged[-1][1] = fin
            else:

                merged.append([ini, fin])
        # Descontar merged evitando doble descuento con comida/cena
        for ini, fin in merged:
            dur = minutos_entre(ini, fin)
            ov = 0
            for v_ini, v_fin in ventanas_descuento:
                ov += _overlap_min(ini, fin, v_ini, v_fin)
            nolab_overlap_cd += ov
            extra_ded += max(0, dur - ov)
    trabajado = max(0, total - comida_ded - cena_ded - extra_ded)
    extra = max(0, trabajado - cfg.umbral_extra_min)
    # Horas extra SIN redondeo (se dejan exactas al minuto). Si en el futuro se desea,
    # se puede habilitar redondeo poniendo redondeo_extra_step_min > 1 y redondeo_extra_modo distinto de "none".
    if getattr(cfg, "redondeo_extra_step_min", 1) and cfg.redondeo_extra_step_min > 1 and getattr(cfg, "redondeo_extra_modo", "none") != "none":
        extra = round_minutes(extra, cfg.redondeo_extra_step_min, cfg.redondeo_extra_modo)
    return trabajado, extra, comida_ded, cena_ded, extra_ded, nolab_overlap_cd, nolab_solape_interno, ignored_outside_shift
# ---------------------------
# Lectura / procesamiento Excel
# ---------------------------
def _coerce_id_str(x: object, min_width: int) -> str:
    s = "" if x is None else str(x).strip()
    if not s:
        return ""
    # Excel a veces convierte 003 -> 3. Si es numérico puro, aplica zfill
    if re.fullmatch(r"\d+", s) and len(s) < min_width:
        return s.zfill(min_width)
    return s
def leer_input(path: Path, cfg: AppConfig) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """
    Lee el archivo y retorna:
      - DataFrame con columnas originales
      - dict canon->col_real: {'id','fecha','semana','nombre','pases','registro'}
    """
    if path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]:
        df = pd.read_excel(path, dtype=str)  # dtype=str para no perder ceros
    else:
        df = _read_csv_flexible(path)
    # Normalizar encabezados (quita BOM/espacios)
    df.columns = [str(c).replace('\ufeff','').strip() for c in df.columns]
    # Detectar columnas con tolerancia
    col_id = _guess_column(df, ["ID", "Id", "Empleado", "No. empleado", "No empleado"])
    col_fecha = _guess_column(df, ["Fecha", "Fecha de registro", "Fecha registro", "Date"])
    col_semana = _guess_column(df, ["Semana", "Día", "Dia", "Día de la semana", "Day"])
    col_nombre = _guess_column(df, ["Nombre", "Name", "Empleado", "Colaborador"])
    col_pases = _guess_column(df, ["Número de pases de la tarjeta", "Numero de pases de la tarjeta", "Pases", "Numero de pases", "Pases de la tarjeta"])
    col_registro = _guess_column(df, ["Registro", "Registros", "Registro de asistencia", "Registro de tiempo", "Time", "Punches"])
    missing = []
    for key, col in [("ID", col_id), ("Fecha", col_fecha), ("Nombre", col_nombre), ("Pases", col_pases), ("Registro", col_registro)]:
        if col is None:
            missing.append(key)

    if missing:
        raise KeyError(
            f"Faltan columnas requeridas: {', '.join(missing)}.\n"
            f"Columnas encontradas en el archivo: {list(df.columns)}\n"
            f"Tip: Asegúrate de exportar con columnas ID/Fecha/Nombre/Pases/Registro (o equivalentes)."
        )
    # Coerciones mínimas
    df[col_id] = df[col_id].map(lambda v: _coerce_id_str(v, cfg.id_min_width))
    # Caso SIN ID: cuando ID == Nombre (texto), se crea una clave interna NOMBRE:: y se deja ID visible en blanco.
    # Guardamos el ID visible en columna auxiliar para usarla al exportar.
    df["_ID_DISPLAY"] = ""
    def _mk(row):
        k, disp = make_emp_key(row.get(col_id, ""), row.get(col_nombre, ""), cfg.id_min_width)
        return pd.Series({"_EMP_KEY": k, "_ID_DISPLAY": disp})
    tmp = df.apply(_mk, axis=1)
    df["_EMP_KEY"] = tmp["_EMP_KEY"]
    df["_ID_DISPLAY"] = tmp["_ID_DISPLAY"]
    df[col_id] = df["_EMP_KEY"]
    if col_pases:
        # Dejar como entero si es posible (pero conservamos dtype=str)
        df[col_pases] = df[col_pases].fillna("").map(lambda x: re.sub(r"\.0$", "", str(x).strip()))
    return df, {"id": col_id, "fecha": col_fecha, "semana": col_semana, "nombre": col_nombre, "pases": col_pases, "registro": col_registro}
def _grupo_sort_key(emp_id: str, cfg: AppConfig) -> Tuple[int, str]:
    g = cfg.empleado_a_grupo.get(emp_id, "")
    try:
        idx = cfg.grupos_orden.index(g)
    except ValueError:
        idx = 9999
    return idx, emp_id
def aplicar_grupos_y_idgrupo(df: pd.DataFrame, cols: Dict[str, str], cfg: AppConfig, *, permitir_interactivo: bool) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Devuelve (df_procesado, df_idgrupo).
    - df_procesado: ID original, ordenado por grupo
    - df_idgrupo: IDGRUPO, ordenado por grupo
    """
    col_id = cols["id"]
    col_nombre = cols["nombre"]
    # Asegurar asignación: si falta grupo/idgrupo y está permitido, pedirlo
    ids = df[col_id].fillna("").astype(str).tolist()
    for emp in ids:
        if not emp:
            continue
        # Caso SIN ID (clave interna NOMBRE::...):
        # - No forzar captura interactiva
        # - Asignar grupo interno SIN_ID solo para ordenar al final
        # - Mantener IDGRUPO vacío para que NO afecte reportes por IDGRUPO
        if str(emp).startswith("NOMBRE::"):
            if emp not in cfg.empleado_a_grupo:
                cfg.empleado_a_grupo[emp] = "SIN_ID"
            # no asignar IDGRUPO
            if emp not in cfg.empleado_a_idgrupo:
                cfg.empleado_a_idgrupo[emp] = ""
            continue
        if emp not in cfg.empleado_a_grupo and permitir_interactivo:
            print(f"\nEmpleado nuevo detectado: ID={emp}.")
            print("Grupos disponibles (en orden): " + " | ".join(cfg.grupos_orden))
            # Selección de grupo con validación (evita teclear algo accidental)
            while True:
                g_in = _safe_input("Asigna grupo (teclea exactamente como aparece) [Enter = primer grupo]: ", "").strip()
                if not g_in:
                    cfg.empleado_a_grupo[emp] = cfg.grupos_orden[0]
                    break
                if g_in in cfg.grupos_orden:
                    cfg.empleado_a_grupo[emp] = g_in
                    break
                resp = _safe_input(f"El grupo '{g_in}' no existe. ¿Deseas crearlo? (S/N): ", "N").strip().upper()
                if resp == "S":
                    cfg.grupos_orden.append(g_in)

                    if g_in not in cfg.grupos_meta:
                        cfg.grupos_meta[g_in] = {"prefijo": g_in}
                    cfg.empleado_a_grupo[emp] = g_in
                    break
        if emp not in cfg.empleado_a_idgrupo and permitir_interactivo:
            # Si es un empleado SIN ID (clave NOMBRE::), no generamos IDGRUPO automático.
            if str(emp).startswith("NOMBRE::"):
                cfg.empleado_a_idgrupo[emp] = ""
            else:
                g = cfg.empleado_a_grupo.get(emp, cfg.grupos_orden[0])
                pref = cfg.prefijo_de_grupo(g)
                sugerido = f"{pref}-{emp}"
                nuevo = _safe_input(f"IDGRUPO para ID={emp} (Enter para usar '{sugerido}'): ", "").strip()
                cfg.empleado_a_idgrupo[emp] = nuevo if nuevo else sugerido
    # Construir columnas calculadas (sin mostrar grupo/prefijo según requerimiento)
    df2 = df.copy()
    df2["_grp_idx"] = df2[col_id].map(lambda x: _grupo_sort_key(str(x), cfg)[0])
    df2["_grp_idgrupo"] = df2[col_id].map(lambda x: cfg.empleado_a_idgrupo.get(str(x), ""))
    # Orden por grupo y luego por ID (estable)
    df2 = df2.sort_values(by=["_grp_idx", col_id], kind="stable").reset_index(drop=True)
    # df_idgrupo: reemplaza ID por IDGRUPO visible
    df_idgrupo = df2.copy()
    df_idgrupo.insert(0, "IDGRUPO", df_idgrupo["_grp_idgrupo"])
    # En archivo _IDGRUPO solo se conserva la columna IDGRUPO (se elimina el ID original)
    df_idgrupo = df_idgrupo.drop(columns=[col_id], errors="ignore")
    # quitar columnas internas
    df_idgrupo = df_idgrupo.drop(columns=["_grp_idx", "_grp_idgrupo"], errors="ignore")
    # df_procesado: mantiene ID original, pero mismo orden de df2
    df_procesado = df2.drop(columns=["_grp_idx", "_grp_idgrupo"], errors="ignore")
    # Guardar config si hubo cambios
    return df_procesado, df_idgrupo
def construir_salida(df: pd.DataFrame, cols: Dict[str, str], cfg: AppConfig, run_id: str = "", *, correcciones_nolabor: Optional[Dict[Tuple[str, date], List[Tuple[Optional[time], Optional[time], str]]]] = None, correcciones_eventos: Optional[Dict[Tuple[str, date], Dict[str, Tuple[Optional[time], str]]]] = None, correccion_interactiva: bool = False, usuario_editor: str = "RRHH", audit_log: Optional[List[AuditEntry]] = None, modo_seguro: bool = False) -> pd.DataFrame:
    col_id = cols["id"]
    col_fecha = cols["fecha"]
    col_semana = cols.get("semana")
    col_nombre = cols["nombre"]
    col_pases = cols["pases"]
    col_registro = cols["registro"]
    correcciones_nolabor = correcciones_nolabor or {}
    out_rows = []
    for _, r in df.iterrows():
        emp_id = normalize_id(r.get(col_id, ""), cfg.id_min_width)
        fecha_d = parse_date(r.get(col_fecha, ""), cfg)
        fecha = fecha_d if fecha_d else str(r.get(col_fecha, "")).strip()
        semana = str(r.get(col_semana, "")).strip() if col_semana else ""
        nombre = str(r.get(col_nombre, "")).strip()
        pases = str(r.get(col_pases, "")).strip()
        registro = r.get(col_registro, "")
        times = parse_registro(registro)

        # Auditoría/edición interactiva por registro (opcional)
        if audit_log is None:
            audit_log = []
        ajuste_manual = ""
        nota_ajuste = ""

        if correccion_interactiva and isinstance(fecha_d, date) and emp_id and not str(emp_id).startswith("NOMBRE::"):
            resp = _safe_input(
                f"¿Deseas revisar/editar las checadas de este registro (ID: {emp_id}, Fecha: {fecha_d.isoformat()})? (S/N): ",
                "N",
            ).strip().lower()
            if resp in ("s", "si", "sí", "y", "yes"):
                no_labor = correcciones_nolabor.get((emp_id, fecha_d), None) if correcciones_nolabor else None
                times_edit, nota_final, _bulk_plan, no_labor_edit = editar_checadas_interactivo(
                    emp_id=emp_id,
                    run_id=run_id,
                    nombre=nombre,
                    fecha_d=fecha_d,
                    registro_raw=str(registro or ""),
                    cfg=cfg,
                    usuario=usuario_editor,
                    audit_log=audit_log,
                    modo_seguro=modo_seguro,
                    no_labor=no_labor,
                )
                if times_edit is not None:
                    times = times_edit
                    # Persistir NoLaborado también en legacy
                    if isinstance(fecha_d, date):
                        key = (emp_id, fecha_d)
                        if no_labor_edit:
                            correcciones_nolabor[key] = copy.deepcopy(no_labor_edit)
                        else:
                            correcciones_nolabor.pop(key, None)
                    if nota_final:
                        ajuste_manual = "Sí"
                        nota_ajuste = nota_final

        # Normalización (cruce de medianoche) – puede desactivarse con modo_seguro
        if modo_seguro:
            times_norm, _reordenado = times[:], False
        else:
            times_norm, _reordenado = normalize_registro_times(times)

        eventos = map_eventos(times_norm)
        # Aplicar correcciones de eventos (Ajustes) ANTES del cálculo, para que afecten Horas trabajadas/extra

        corr_evs = correcciones_eventos.get((emp_id, fecha), None) if isinstance(fecha, date) else None
        if corr_evs:
            ajuste_manual = "Sí"
            parts = []
            for ev, (t_corr, nota_corr) in corr_evs.items():
                if ev in eventos:
                    eventos[ev] = t_corr
                if nota_corr:
                    parts.append(f"{ev}: {nota_corr}")
                else:
                    if t_corr is None:
                        parts.append(f"{ev}: borrado")
                    else:
                        try:
                            parts.append(f"{ev}: {t_corr.strftime('%H:%M')}")
                        except Exception:
                            parts.append(f"{ev}: ajustado")
            nota_ajuste = " | ".join(parts)
        no_lab = correcciones_nolabor.get((emp_id, fecha), []) if isinstance(fecha, date) else []
        trabajado_min, extra_min, comida_ded, cena_ded, no_lab_ded, nolab_ov, nolab_intov, nolab_ign = calcular_trabajado(eventos, cfg, no_lab)
        # Auditoría de descuentos por tiempo NO laborado (para RRHH)
        notas_partes = []
        extra_regs = int(eventos.get("_extra_registros", 0) or 0)
        if extra_regs > 0:
            notas_partes.append(f"Registros extra ignorados: {extra_regs}")
        if "_reordenado" in locals() and _reordenado:
            notas_partes.append("Registro fuera de orden; reordenado")
        if no_lab_ded and no_lab_ded > 0:
            notas_partes.append(f"Descuento NoLaborado: {minutes_to_hhmm(no_lab_ded)}")
            # Auditoría: indicar solape de NoLaborado con comida/cena (sin doble descuento)
            if nolab_ov and nolab_ov > 0:
                notas_partes.append(getattr(cfg, "nota_nolab_solape_cd", "NoLaborado solapado con comida/cena: fusionado"))
        if nolab_intov and nolab_intov > 0:
            notas_partes.append(f"Solape interno NoLaborado: {minutes_to_hhmm(nolab_intov)} (fusionado)")
        if nolab_ign and nolab_ign > 0:
            notas_partes.append(f"{getattr(cfg, 'nota_nolab_fuera_jornada', 'NoLaborado fuera de jornada ignorado')}: {minutes_to_hhmm(nolab_ign)}")
        # Nota opcional si el sistema aplicó cierre automático por regreso faltante
        # (solo informativo; no afecta el cálculo más allá de lo ya descontado)
        if eventos.get("Salida a comer") and not eventos.get("Regreso de comer") and eventos.get("Salida"):
            notas_partes.append("Comida incompleta: fin asumido=Salida")
        if eventos.get("Salida a cenar") and not eventos.get("Regreso de cenar") and eventos.get("Salida"):
            notas_partes.append("Cena incompleta: fin asumido=Salida")
        notas_str = " | ".join(notas_partes)
        row = {
            "ID": emp_id,
            "Fecha": fecha,
            "Semana": semana,
            "Nombre": nombre,
            "Pases": pases,
            "Entrada": fmt_hhmm(eventos["Entrada"]),
            "Salida a comer": fmt_hhmm(eventos["Salida a comer"]),
            "Regreso de comer": fmt_hhmm(eventos["Regreso de comer"]),
            "Salida a cenar": fmt_hhmm(eventos["Salida a cenar"]),
            "Regreso de cenar": fmt_hhmm(eventos["Regreso de cenar"]),
            "Salida": fmt_hhmm(eventos["Salida"]),
            "Horas trabajadas": minutes_to_hhmm(trabajado_min),
            "Horas extra": minutes_to_hhmm(extra_min),
            "Notas": notas_str,
            "Ajuste manual": ajuste_manual,
            "Nota ajuste": nota_ajuste,
        }
        out_rows.append(row)
    out = pd.DataFrame(out_rows)
    # Compactar Semana si viene vacía (algunas exportaciones no la traen)
    if "Semana" in out.columns and out["Semana"].astype(str).str.strip().eq("").all():
        out = out.drop(columns=["Semana"])
    return out

def rango_semana(fecha: pd.Timestamp, week_start_dow: int) -> str:
    dow = fecha.weekday()
    delta = (dow - week_start_dow) % 7
    inicio = fecha - pd.Timedelta(days=delta)
    fin = inicio + pd.Timedelta(days=6)
    return f"{inicio.strftime('%Y-%m-%d')} a {fin.strftime('%Y-%m-%d')}"
def construir_resumen_semanal_vertical(df_out: pd.DataFrame, cfg, faltas_semanal: pd.DataFrame = None) -> pd.DataFrame:
    """Resumen semanal VERTICAL por ID: Días presentes, Faltas, Totales."""
    df = df_out.copy()
    if "Fecha" not in df.columns:
        return pd.DataFrame()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.dropna(subset=["Fecha"])
    if len(df)==0:
        return pd.DataFrame()
    df["Semana"] = df["Fecha"].apply(lambda x: _week_key(x, cfg))
    df["_min_trab"] = df.get("Horas trabajadas","").apply(hhmm_to_minutes)
    df["_min_extra"] = df.get("Horas extra","").apply(hhmm_to_minutes)
    grp_cols = ["ID","Nombre","Semana"]
    for c in grp_cols:
        if c not in df.columns:
            return pd.DataFrame()
    g = df.groupby(grp_cols, dropna=False, as_index=False).agg(
        Dias_presentes=("Fecha", lambda x: x.dt.date.nunique()),
        Total_trab_min=("_min_trab","sum"),
        Total_extra_min=("_min_extra","sum"),
    )
    g["Total horas trabajadas"] = g["Total_trab_min"].apply(minutes_to_hhmm)
    g["Total horas extra"] = g["Total_extra_min"].apply(minutes_to_hhmm)
    g = g.drop(columns=["Total_trab_min","Total_extra_min"])
    if faltas_semanal is not None and len(faltas_semanal)>0 and set(["ID","Semana","Faltas"]).issubset(faltas_semanal.columns):
        g = g.merge(faltas_semanal[["ID","Semana","Faltas"]], on=["ID","Semana"], how="left")
    if "Faltas" not in g.columns:
        g["Faltas"]=0
    g["Faltas"] = pd.to_numeric(g["Faltas"], errors="coerce").fillna(0).astype(int)
    g["_grp_idx"] = g["ID"].map(lambda x: _grupo_sort_key(str(x), cfg)[0])
    g = g.sort_values(by=["_grp_idx","ID","Semana"], kind="stable").drop(columns=["_grp_idx"]).reset_index(drop=True)
    # Agregar rango real de la semana (según cfg.week_start_dow)
    try:
        # tomar la primera fecha de la semana por ID
        fechas = df.groupby(["ID","Semana"])["Fecha"].min().reset_index()
        fechas["Rango semana"] = fechas["Fecha"].apply(lambda f: rango_semana(pd.to_datetime(f), cfg.week_start_dow))
        g = g.merge(fechas[["ID","Semana","Rango semana"]], on=["ID","Semana"], how="left")
    except Exception:
        pass
    return g
def construir_resumen_semanal(df_out: pd.DataFrame, cfg, faltas_semanal: pd.DataFrame = None) -> pd.DataFrame:
    """Resumen semanal HORIZONTAL:
    - 1 fila por (ID, Semana)
    - 7 columnas de días (según week_start_dow) con HH:MM de horas trabajadas por día o 'F' si falta
    - Totales al final (horas trabajadas, horas extra, faltas)
    """
    df = df_out.copy()
    if "Fecha" not in df.columns:
        return pd.DataFrame()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.dropna(subset=["Fecha"])

    if len(df) == 0:
        return pd.DataFrame()
    # Claves
    df["Semana"] = df["Fecha"].apply(lambda x: _week_key(x, cfg))
    df["_min_trab"] = df.get("Horas trabajadas", "").apply(hhmm_to_minutes)
    df["_min_extra"] = df.get("Horas extra", "").apply(hhmm_to_minutes)
    # Validar cols base
    for c in ["ID", "Nombre", "Semana"]:
        if c not in df.columns:
            return pd.DataFrame()
    # Day names ES (completo) con orden configurable
    dias_es = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    try:
        ws = int(getattr(cfg, "week_start_dow", 0) or 0) % 7
    except Exception:
        ws = 0
    dias_orden = dias_es[ws:] + dias_es[:ws]
    # Asignar etiqueta de día
    df["_dia"] = df["Fecha"].dt.dayofweek.map(lambda i: dias_es[int(i)] if pd.notna(i) else "")
    # Minutos por día (si hay duplicados por el mismo día, suma)
    daily = df.groupby(["ID", "Nombre", "Semana", "_dia"], as_index=False).agg(
        Trab_min=("_min_trab", "sum"),
        Extra_min=("_min_extra", "sum"),
        Fecha_min=("Fecha", "min"),
    )
    daily["Horas_dia"] = daily["Trab_min"].apply(minutes_to_hhmm)
    # Pivot a horizontal fijo 7 columnas
    pivot = daily.pivot_table(index=["ID", "Nombre", "Semana"], columns="_dia", values="Horas_dia", aggfunc="first").reset_index()
    # Asegurar columnas de días en el orden deseado, tipo texto
    for dname in dias_orden:
        if dname not in pivot.columns:
            pivot[dname] = ""
    pivot[dias_orden] = pivot[dias_orden].astype(object)
    # Totales por semana
    totals = df.groupby(["ID", "Nombre", "Semana"], as_index=False).agg(
        Dias_presentes=("Fecha", lambda x: x.dt.date.nunique()),
        Total_trab_min=("_min_trab", "sum"),
        Total_extra_min=("_min_extra", "sum"),
    )
    totals["Total horas trabajadas"] = totals["Total_trab_min"].apply(minutes_to_hhmm)
    totals["Total horas extra"] = totals["Total_extra_min"].apply(minutes_to_hhmm)
    totals = totals.drop(columns=["Total_trab_min", "Total_extra_min"])
    # Faltas
    if faltas_semanal is not None and len(faltas_semanal) > 0 and set(["ID", "Semana", "Faltas"]).issubset(faltas_semanal.columns):
        totals = totals.merge(faltas_semanal[["ID", "Semana", "Faltas"]], on=["ID", "Semana"], how="left")
    if "Faltas" not in totals.columns:
        totals["Faltas"] = 0
    totals["Faltas"] = pd.to_numeric(totals["Faltas"], errors="coerce").fillna(0).astype(int)
    # Rango semana
    try:
        fechas = df.groupby(["ID", "Semana"])["Fecha"].min().reset_index()
        fechas["Rango semana"] = fechas["Fecha"].apply(lambda f: rango_semana(pd.to_datetime(f), ws))
        totals = totals.merge(fechas[["ID", "Semana", "Rango semana"]], on=["ID", "Semana"], how="left")
    except Exception:
        totals["Rango semana"] = ""
    # Merge todo
    out = pivot.merge(totals, on=["ID", "Nombre", "Semana"], how="left")
    # Marcar faltas en celdas del día si falta (solo si tenemos faltas + plantilla permite conocer esperados)
    # Nota: La lógica de faltas ya calcula por semana, aquí solo ponemos 'F' si la persona NO aparece ese día.
    # Construimos conjunto de fechas presentes por (ID, Semana, Dia)
    present_set = set(zip(daily["ID"].astype(str), daily["Semana"].astype(str), daily["_dia"].astype(str)))

def mark_f(row):
    """Marca 'F' en días vacíos cuando faltas_semanal está disponible."""
    if faltas_semanal is None or len(faltas_semanal) == 0:
        return row
    _id = str(row.get("ID", "")).strip()
    _wk = str(row.get("Semana", "")).strip()
    for dname in dias_orden:
        val = row.get(dname, "")
        if val is None or (isinstance(val, float) and pd.isna(val)):
            val = ""
        if str(val).strip() == "" and (_id, _wk, dname) not in present_set:
            row[dname] = "F"
    return row

    out = out.apply(mark_f, axis=1)
    # Orden estable por grupos
    out["_grp_idx"] = out["ID"].map(lambda x: _grupo_sort_key(str(x), cfg)[0])
    out = out.sort_values(by=["_grp_idx", "ID", "Semana"], kind="stable").drop(columns=["_grp_idx"]).reset_index(drop=True)
    # Reordenar columnas: ID, Nombre, Semana, Rango, días..., totales
    cols = ["ID", "Nombre", "Semana"]
    if "Rango semana" in out.columns:
        cols.append("Rango semana")
    cols += dias_orden
    for c in ["Dias_presentes", "Faltas", "Total horas trabajadas", "Total horas extra"]:
        if c in out.columns:
            cols.append(c)
    # preservar otras cols al final
    restantes = [c for c in out.columns if c not in cols]
    out = out[cols + restantes]
    return out
def construir_resumen_mensual(df_out: pd.DataFrame, cfg, faltas_mensual: pd.DataFrame = None) -> pd.DataFrame:
    """Resumen mensual por ID: Días presentes, Faltas, Totales."""
    df = df_out.copy()
    if "Fecha" not in df.columns:
        return pd.DataFrame()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.dropna(subset=["Fecha"])
    if len(df)==0:
        return pd.DataFrame()
    df["Mes"] = df["Fecha"].apply(_month_key)
    df["_min_trab"] = df.get("Horas trabajadas","").apply(hhmm_to_minutes)
    df["_min_extra"] = df.get("Horas extra","").apply(hhmm_to_minutes)
    grp_cols = ["ID","Nombre","Mes"]
    for c in grp_cols:
        if c not in df.columns:
            return pd.DataFrame()

    g = df.groupby(grp_cols, dropna=False, as_index=False).agg(
        Dias_presentes=("Fecha", lambda x: x.dt.date.nunique()),
        Total_trab_min=("_min_trab","sum"),
        Total_extra_min=("_min_extra","sum"),
    )
    g["Total horas trabajadas"] = g["Total_trab_min"].apply(minutes_to_hhmm)
    g["Total horas extra"] = g["Total_extra_min"].apply(minutes_to_hhmm)
    g = g.drop(columns=["Total_trab_min","Total_extra_min"])
    if faltas_mensual is not None and len(faltas_mensual)>0 and set(["ID","Mes","Faltas"]).issubset(faltas_mensual.columns):
        g = g.merge(faltas_mensual[["ID","Mes","Faltas"]], on=["ID","Mes"], how="left")
    if "Faltas" not in g.columns:
        g["Faltas"]=0
    g["Faltas"] = pd.to_numeric(g["Faltas"], errors="coerce").fillna(0).astype(int)
    g["_grp_idx"] = g["ID"].map(lambda x: _grupo_sort_key(str(x), cfg)[0])
    g = g.sort_values(by=["_grp_idx","ID","Mes"], kind="stable").drop(columns=["_grp_idx"]).reset_index(drop=True)
    return g
def _iso_week_key(d: pd.Timestamp) -> str:
    iso = d.isocalendar()
    return f"{iso.year}-W{int(iso.week):02d}"
def _week_key(d: pd.Timestamp, cfg: AppConfig) -> str:
    """Clave de semana según el inicio configurado (week_start_dow).
    - 0=Lunes (ISO), 2=Miércoles, etc.
    Devuelve una clave estable tipo: YYYY-WK-YYYYMMDD (fecha de inicio de semana).
    """
    try:
        dow = int(getattr(cfg, "week_start_dow", 0) or 0)
    except Exception:
        dow = 0
    dow = dow % 7
    # pandas Timestamp -> datetime.date
    dd = pd.to_datetime(d).date()
    delta = (dd.weekday() - dow) % 7
    ws = dd - timedelta(days=delta)
    return f"{ws.year:04d}-WK-{ws.strftime('%Y%m%d')}"
def _month_key(d: pd.Timestamp) -> str:
    return f"{d.year:04d}-{d.month:02d}"
def calcular_faltas(df_out: pd.DataFrame, plantilla: pd.DataFrame, cfg: AppConfig) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Calcula faltas por semana y mes + detalle, usando plantilla (empleados activos).
    Fechas esperadas: rango completo Fecha_min..Fecha_max del archivo procesado.
    """
    if df_out is None or len(df_out)==0 or plantilla is None or len(plantilla)==0:
        return (pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    if "Fecha" not in df_out.columns:
        return (pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    df = df_out.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.dropna(subset=["Fecha"])
    if len(df)==0:
        return (pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    min_d = df["Fecha"].min().date()
    max_d = df["Fecha"].max().date()
    # Expandir rango esperado a semanas completas según inicio configurado.
    # Esto permite detectar FALTAS aunque proceses un archivo de 1 solo día.
    def _week_start(d: date) -> date:
        dow = d.weekday()
        delta = (dow - int(getattr(cfg, "week_start_dow", 0))) % 7
        return d - timedelta(days=delta)
    def _week_end(d: date) -> date:
        return _week_start(d) + timedelta(days=6)
    exp_ini = _week_start(min_d)
    exp_fin = _week_end(max_d)
    fechas = pd.date_range(exp_ini, exp_fin, freq="D").date
    pad = lambda v: _coerce_id_str(v, 3)
    presentes = set((pad(r.get("ID","")), r["Fecha"].date()) for _, r in df.iterrows() if str(r.get("ID","")).strip()!="")
    nombre_por_id = {}
    if "Nombre" in df.columns:

        for _, r in df[["ID","Nombre"]].dropna().drop_duplicates().iterrows():
            i = pad(r.get("ID",""))
            if i and i not in nombre_por_id:
                nombre_por_id[i] = str(r.get("Nombre","") or "")
    for _, r in plantilla.iterrows():
        i = pad(r.get("ID",""))
        n = str(r.get("Nombre","") or "")
        if i and n:
            nombre_por_id[i]=n
    detalles=[]
    for _, emp in plantilla.iterrows():
        emp_id = pad(emp.get("ID",""))
        if not emp_id or not bool(emp.get("_activo", True)):
            continue
        alta = emp.get("FechaAlta"); baja = emp.get("FechaBaja")
        for d in fechas:
            if pd.notna(alta) and alta and d < alta: 
                continue
            if pd.notna(baja) and baja and d > baja:
                continue
            if (emp_id, d) not in presentes:
                ts = pd.Timestamp(d)
                detalles.append({"ID":emp_id,"Nombre":nombre_por_id.get(emp_id,""),"Fecha":ts.date(),
                                 "Semana":_week_key(ts, cfg),"Mes":_month_key(ts)})
    df_det = pd.DataFrame(detalles)
    if len(df_det)==0:
        return (pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    df_sem = df_det.groupby(["ID","Nombre","Semana"], as_index=False).agg(Faltas=("Fecha","count"))
    df_mes = df_det.groupby(["ID","Nombre","Mes"], as_index=False).agg(Faltas=("Fecha","count"))
    return (df_sem, df_mes, df_det)
def _dia_abrev_es(weekday: int) -> str:
    # 0=Lun ... 6=Dom
    mapa = {0: "Lun", 1: "Mar", 2: "Mié", 3: "Jue", 4: "Vie", 5: "Sáb", 6: "Dom"}
    return mapa.get(int(weekday) % 7, "Dia")
def crear_resumen_semanal_checadas(df_rep: pd.DataFrame, cfg: AppConfig, modo: str = "PROCESADO") -> pd.DataFrame:
    """
    Hoja de control interno: RESUMEN SEMANAL DE CHECADAS (por empleado).
    - Una fila por empleado por semana (según cfg.week_start_dow).
    - Columnas por día y por evento:
        Entrada, Salida a comer, Regreso de comer, Salida a cenar, Regreso de cenar, Salida
    - Para modo IDGRUPO, se omite columna ID (archivo orientado a IDGRUPO).
    Nota: No depende de que la columna "Semana" venga en formato clave; calcula semana desde "Fecha".
    """
    if df_rep is None or len(df_rep) == 0:
        return pd.DataFrame()
    eventos = ["Entrada", "Salida a comer", "Regreso de comer", "Salida a cenar", "Regreso de cenar", "Salida"]
    if "Fecha" not in df_rep.columns or "Nombre" not in df_rep.columns:
        return pd.DataFrame()
    if not set(eventos).issubset(set(df_rep.columns)):
        return pd.DataFrame()
    df = df_rep.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce").dt.date
    df = df.dropna(subset=["Fecha"])
    if len(df) == 0:
        return pd.DataFrame()
    # Semana según week_start_dow (por defecto 2=Miércoles para tu operación)

    try:
        dow = int(getattr(cfg, "week_start_dow", 2) or 2)
    except Exception:
        dow = 2
    dow = dow % 7
    def _week_start(dd):
        delta = (dd.weekday() - dow) % 7
        return dd - timedelta(days=delta)
    df["WSTART"] = df["Fecha"].apply(_week_start)
    df["Semana"] = df["WSTART"].apply(lambda ws: f"{ws.strftime('%Y-%m-%d')} a {(ws + timedelta(days=6)).strftime('%Y-%m-%d')}")
    df["DIA_OFF"] = df.apply(lambda r: (r["Fecha"] - r["WSTART"]).days, axis=1)
    df = df[(df["DIA_OFF"] >= 0) & (df["DIA_OFF"] <= 6)]
    if len(df) == 0:
        return pd.DataFrame()
    df["DIA_LABEL"] = df.apply(lambda r: f"{_dia_abrev_es((r['WSTART'] + timedelta(days=int(r['DIA_OFF']))).weekday())} {(r['WSTART'] + timedelta(days=int(r['DIA_OFF']))).strftime('%d/%m')}", axis=1)
    modo = (modo or "PROCESADO").upper().strip()
    if modo == "IDGRUPO":
        if "IDGRUPO" not in df.columns:
            return pd.DataFrame()
        keys = ["IDGRUPO", "Nombre", "Semana"]
    else:
        if "ID" not in df.columns:
            return pd.DataFrame()
        keys = ["ID", "Nombre", "Semana"]
    long = df[keys + ["DIA_LABEL"] + eventos].melt(
        id_vars=keys + ["DIA_LABEL"],
        value_vars=eventos,
        var_name="Evento",
        value_name="Hora",
    )
    long["Col"] = long["DIA_LABEL"].astype(str) + " " + long["Evento"].astype(str)
    wide = long.pivot_table(index=keys, columns="Col", values="Hora", aggfunc="first").reset_index()
    # Orden de columnas: 7 días (según WSTART) x eventos
    day_labels = [f"{_dia_abrev_es((date(2000,1,3) + timedelta(days=((dow + i) % 7 - 0))).weekday())}" for i in range(7)]  # dummy, luego se reemplaza
    try:
        ws0 = pd.to_datetime(df["WSTART"].min()).date()
        day_labels = [f"{_dia_abrev_es((ws0 + timedelta(days=i)).weekday())} {(ws0 + timedelta(days=i)).strftime('%d/%m')}" for i in range(7)]
    except Exception:
        day_labels = []
    ordered_cols = []
    for dl in day_labels:
        for ev in eventos:
            cname = f"{dl} {ev}"
            if cname in wide.columns:
                ordered_cols.append(cname)
    base_cols = keys
    remaining = [c for c in wide.columns if c not in base_cols and c not in ordered_cols]
    out = wide[base_cols + ordered_cols + remaining].copy()
    for c in out.columns:
        if c in base_cols:
            continue
        out[c] = out[c].fillna("")
    # Orden filas: dejar blancos al final
    if modo == "IDGRUPO":
        out["__SORT"] = out["IDGRUPO"].astype(str).replace({"": "ZZZ"}).fillna("ZZZ")
        out = out.sort_values(["Semana", "__SORT", "Nombre"]).drop(columns=["__SORT"])
    else:
        out["__SORT"] = out["ID"].astype(str).replace({"": "ZZZ"}).fillna("ZZZ")
        out = out.sort_values(["Semana", "__SORT", "Nombre"]).drop(columns=["__SORT"])
    return out
def exportar_excel(df: pd.DataFrame, out_path: Path, extra_sheets: dict = None) -> None:
    from openpyxl.utils import get_column_letter
    _ensure_dir(out_path.parent)
    extra_sheets = extra_sheets or {}

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        _drop_export_debug_cols(df).to_excel(writer, index=False, sheet_name="Reporte")
        for sheet_name, sdf in extra_sheets.items():
            if sdf is None or len(sdf) == 0:
                continue
            # Limitar nombre de hoja a 31 chars (Excel)
            sname = str(sheet_name)[:31]
            _drop_export_debug_cols(sdf).to_excel(writer, index=False, sheet_name=sname)
        # Formato: asegurar que la columna ID se guarde como TEXTO (para preservar ceros a la izquierda)
        for sname, ws in writer.sheets.items():
            # Detectar columna "ID" por encabezado
            try:
                headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
                if "ID" in headers:
                    id_col = headers.index("ID") + 1
                    for r in range(2, ws.max_row+1):
                        cell = ws.cell(row=r, column=id_col)
                        if cell.value is None:
                            continue
                        cell.value = _coerce_id_str(cell.value, 3)
                        cell.number_format = "@"
                    ws.cell(row=1, column=id_col).number_format = "@"
            except Exception:
                pass
        # Ajuste básico de ancho para todas las hojas
        for sname, ws in writer.sheets.items():
            try:
                df_ref = df if sname == "Reporte" else extra_sheets.get(sname) or extra_sheets.get(sheet_name)  # best-effort
            except Exception:
                df_ref = None
            if df_ref is None:
                continue
            for i, col in enumerate(df_ref.columns, start=1):
                try:
                    max_len = max([len(str(col))] + [len(str(x)) for x in df_ref[col].astype(str).head(200).tolist()])
                except Exception:
                    max_len = len(str(col))
                ws.column_dimensions[get_column_letter(i)].width = min(40, max(10, max_len + 2))
# ---------------------------
# Modo administración (simple)
# ---------------------------
def modo_administracion(script_dir: Path, plantilla_path: Optional[str] = None) -> None:
    """Menú interactivo para administrar grupos y asignaciones.
    Archivos:
      - mapa_grupos.json: guarda orden de grupos, prefijos, empleado->grupo, empleado->idgrupo, y reglas.
      - (opcional) plantilla_empleados*.xlsx: si existe, se usa para mostrar nombres en el listado.
    """
    cfg_base = cargar_config(script_dir)
    cfg = copy.deepcopy(cfg_base)
    dirty = False
    def _cargar_nombres() -> dict[str, str]:
        """Carga ID->Nombre desde plantilla(s) si existen.
        Busca, en orden:
          - plantilla_path (si se indicó)

          - plantilla_empleados_OFICIAL.xlsx / plantilla_empleados.xlsx
        Acepta hoja 'Empleados' o, si no existe, la primera hoja.
        """
        candidatos = []
        if plantilla_path:
            candidatos.append(Path(plantilla_path))
        candidatos += [
            script_dir / "plantilla_empleados_OFICIAL.xlsx",
            script_dir / "plantilla_empleados_OFICIAL_PRUEBA_86.xlsx",
            script_dir / "plantilla_empleados.xlsx",
        ]
        for p in candidatos:
            try:
                if not p.exists():
                    continue
                try:
                    df = pd.read_excel(p, sheet_name="Empleados", dtype=str)
                except Exception:
                    df = pd.read_excel(p, dtype=str)
                df.columns = [str(c).replace('\ufeff','').strip() for c in df.columns]
                cols = {str(c).strip().lower(): c for c in df.columns}
                id_col = cols.get("id")
                nombre_col = cols.get("nombre")
                if not id_col:
                    continue
                out = {}
                for _, r in df.iterrows():
                    emp = str(r.get(id_col, "")).strip()
                    if not emp:
                        continue
                    emp = normalize_id(emp, cfg.id_min_width)
                    nom = ""
                    if nombre_col:
                        nom = str(r.get(nombre_col, "")).strip()
                    if nom:
                        out[emp] = nom
                if out:
                    return out
            except Exception:
                continue
        return {}
    nombres = _cargar_nombres()
    def _autosave():
        nonlocal dirty
        dirty = True
    def _save():
        nonlocal dirty
        guardar_config(script_dir, cfg)
        dirty = False
    def _hr(ch: str = "■", n: int = 64) -> str:
        return ch * n
    def _status() -> str:
        return "SIN GUARDAR" if dirty else "OK"
    def menu():
        print("\n" + _hr())
        print(f"MODO ADMINISTRACIÓN (v18)".ljust(44) + f"Estado: {_status()}")
        print(_hr())
        print(f"Config: {script_dir / 'mapa_grupos.json'}")
        print(f"Grupos: {len(cfg.grupos_orden)} | Empleados asignados: {len(cfg.empleado_a_grupo)}")
        print("Atajos: 90=Guardar | 99=Guardar y salir | 0=Salir")
        print(_hr("-"))
        opciones = [
            ("1",  "Ver grupos (orden)"),
            ("2",  "Crear grupo"),
            ("3",  "Renombrar grupo"),
            ("4",  "Cambiar prefijo de grupo"),
            ("5",  "Reordenar grupos (mover)"),

            ("6",  "Eliminar grupo (con confirmación)"),
            ("7",  "Asignación: empleado -> grupo / IDGRUPO (manual)"),
            ("8",  "Quitar asignación de empleado"),
            ("9",  "Ver asignación (empleados)"),
            ("10", "Cambiar reglas (extra/comida/ID)"),
            ("11", "Buscar empleado por ID"),
            ("12", "Asignación masiva (pegar IDs)"),
            ("13", "Backup de configuración (mapa_grupos.json)"),
            ("90", "Guardar cambios"),
            ("99", "Guardar y salir"),
            ("0",  "Salir sin guardar"),
        ]
        colw = 56
        sep = "  "
        for i in range(0, len(opciones), 2):
            a = opciones[i]
            left = f"{a[0]}) {a[1]}"
            right = ""
            if i + 1 < len(opciones):
                b = opciones[i + 1]
                right = f"{b[0]}) {b[1]}"
            # Si el texto es muy largo, se imprime en 2 líneas para que no se pegue.
            if len(left) > colw - 2:
                print(left)
                if right:
                    print((" " * 4) + right)
                continue
            print(left.ljust(colw) + (sep + right if right else ""))
        print(_hr("-"))
    def _pick_grupo(prompt: str = "Grupo: ", allow_blank: bool = False) -> str:
        """Pide un grupo existente; permite crear si no existe."""
        while True:
            print("Grupos: " + " | ".join(cfg.grupos_orden) if cfg.grupos_orden else "Grupos: (vacío)")
            g = _safe_input(prompt).strip()
            if not g and allow_blank:
                return ""
            if not g:
                print("Grupo vacío.")
                continue
            if g in cfg.grupos_orden:
                return g
            resp = _safe_input(f"El grupo '{g}' no existe. ¿Deseas crearlo? (S/N): ").strip().lower()
            if resp in ("s", "si", "sí", "y", "yes"):
                cfg.grupos_orden.append(g)
                cfg.grupos_meta.setdefault(g, {"prefijo": g})
                _autosave()
                print(f"Creado: {g}")
                return g
            print("Ok. Intenta de nuevo.")
    while True:
        menu()
        op = _safe_input("Opción: ").strip()
        if not op:
            continue
        if op == "0":
            if dirty:
                resp = _safe_input("Hay cambios SIN guardar. ¿Salir sin guardar? (S/N): ").strip().lower()

                if resp not in ("s","si","sí","y","yes"):
                    continue
            print("Saliendo sin guardar.")
            return
        if op == "90":
            _save()
            print("Cambios guardados.")
            continue
        if op == "99":
            _save()
            print("Cambios guardados. Saliendo.")
            return
        if op == "1":
            print("\nOrden de grupos:")
            for i, g in enumerate(cfg.grupos_orden, 1):
                print(f"  {i}. {g} (prefijo: {cfg.prefijo_de_grupo(g)})")
        elif op == "2":
            print("\nCrear grupo")
            nuevo = _safe_input("Nombre del nuevo grupo: ").strip()
            if not nuevo:
                print("Nombre inválido.")
                continue
            if nuevo in cfg.grupos_orden:
                print("Ese grupo ya existe.")
                continue
            cfg.grupos_orden.append(nuevo)
            cfg.grupos_meta.setdefault(nuevo, {"prefijo": nuevo})
            _autosave()
            print(f"Creado: {nuevo}")
        elif op == "3":
            print("\nRenombrar grupo")
            actual = _pick_grupo("Grupo actual: ")
            nuevo = _safe_input("Nuevo nombre: ").strip()
            if not nuevo:
                print("Nombre nuevo inválido.")
                continue
            if nuevo in cfg.grupos_orden and nuevo != actual:
                print("Ese grupo ya existe.")
                continue
            idx = cfg.grupos_orden.index(actual)
            cfg.grupos_orden[idx] = nuevo
            cfg.grupos_meta[nuevo] = cfg.grupos_meta.pop(actual, {"prefijo": nuevo})
            cfg.empleado_a_grupo = {k: (nuevo if v == actual else v) for k, v in cfg.empleado_a_grupo.items()}
            _autosave()
            print(f"Renombrado: {actual} -> {nuevo}")
        elif op == "4":
            print("\nCambiar prefijo de grupo")
            g = _pick_grupo()
            pref = _safe_input(f"Nuevo prefijo para '{g}': ").strip()
            if not pref:
                print("Prefijo inválido.")
                continue
            cfg.grupos_meta.setdefault(g, {})["prefijo"] = pref
            _autosave()
            print("Prefijo actualizado.")
        elif op == "5":
            print("\nReordenar grupos (mover)")
            if len(cfg.grupos_orden) < 2:
                print("No hay suficientes grupos para reordenar.")
                continue
            for i, g in enumerate(cfg.grupos_orden, 1):
                print(f"  {i}. {g}")
            try:
                idx = int(_safe_input("Número de grupo a mover: ").strip())
                if not (1 <= idx <= len(cfg.grupos_orden)):
                    raise ValueError
                dir_ = _safe_input("Mover (U)p / (D)own: ").strip().lower()
                if dir_ not in ("u", "d"):
                    print("Dirección inválida.")

                    continue
                new_idx = idx - 2 if dir_ == "u" else idx
                if not (0 <= new_idx < len(cfg.grupos_orden)):
                    print("No se puede mover más.")
                    continue
                g = cfg.grupos_orden.pop(idx - 1)
                cfg.grupos_orden.insert(new_idx, g)
                _autosave()
                print("Orden actualizado.")
            except ValueError:
                print("Entrada inválida.")
        elif op == "6":
            print("\nEliminar grupo")
            g = _pick_grupo()
            # confirmar
            resp = _safe_input(f"¿Seguro que deseas eliminar '{g}'? Esto NO borra empleados, solo su grupo. (ESCRIBE EL NOMBRE): ").strip()
            if resp != g:
                print("Confirmación no coincide. Cancelado.")
                continue
            cfg.grupos_orden.remove(g)
            cfg.grupos_meta.pop(g, None)
            # empleados asignados a ese grupo quedan sin grupo
            cfg.empleado_a_grupo = {k: v for k, v in cfg.empleado_a_grupo.items() if v != g}
            _autosave()
            print(f"Grupo eliminado: {g}")
        elif op == "7":
            print("\nAsignar empleado -> grupo / IDGRUPO (manual)")
            emp_raw = _safe_input("ID de empleado (original): ").strip()
            if not emp_raw:
                continue
            emp = normalize_id(emp_raw, cfg.id_min_width)
            g = _pick_grupo("Nuevo grupo: ")
            cfg.empleado_a_grupo[emp] = g
            nuevo_idg = _safe_input("Nuevo IDGRUPO (vacío = sugerir): ").strip()
            if nuevo_idg:
                cfg.empleado_a_idgrupo[emp] = nuevo_idg
            else:
                pref = cfg.prefijo_de_grupo(g)
                cfg.empleado_a_idgrupo[emp] = f"{pref}-{emp}"
            _autosave()
            print("Asignación actualizada.")
        elif op == "8":
            print("\nQuitar asignación de empleado")
            emp_raw = _safe_input("ID de empleado: ").strip()
            if not emp_raw:
                continue
            emp = normalize_id(emp_raw, cfg.id_min_width)
            existed = False
            if emp in cfg.empleado_a_grupo:
                cfg.empleado_a_grupo.pop(emp, None)
                existed = True
            if emp in cfg.empleado_a_idgrupo:
                cfg.empleado_a_idgrupo.pop(emp, None)
                existed = True
            if existed:
                _autosave()
                print("Asignación eliminada.")
            else:
                print("Ese empleado no tenía asignación.")
        elif op == "9":

            print("\nAsignación (empleados):")
            if not nombres:
                try:
                    nombres.update(_cargar_nombres() or {})
                except Exception:
                    pass
            items = sorted(cfg.empleado_a_grupo.items(), key=lambda kv: _grupo_sort_key(kv[0], cfg))
            if not items:
                print("  (vacío)")
                continue
            for emp, g in items[:1000]:
                idg = cfg.empleado_a_idgrupo.get(emp, "")
                nom = nombres.get(emp, "")
                nom_txt = f"  Nombre={nom}" if nom else ""
                print(f"  ID={emp}  Grupo={g}  IDGRUPO={idg}{nom_txt}")
        elif op == "10":
            print("\nCambiar reglas")
            # IMPORTANTE: esta opción debe ser transaccional. Si el usuario mete
            # un valor inválido en cualquier campo, NO se deben guardar cambios parciales.
            old_tope = cfg.tope_descuento_comida_min
            old_w = cfg.id_min_width
            old_um = cfg.umbral_extra_min
            old_ws = getattr(cfg, "week_start_dow", 0)
            try:
                # Capturar valores (sin aplicarlos hasta validar todos)
                tope_txt = _safe_input(f"Tope descuento comida (min) [actual {cfg.tope_descuento_comida_min}]: ").strip()
                w_txt = _safe_input(f"Ancho mínimo ID numérico (ej. 3 para 003) [actual {cfg.id_min_width}]: ").strip()
                um_txt = _safe_input(f"Umbral extra (min) [actual {cfg.umbral_extra_min}]: ").strip()
                ws_txt = _safe_input(f"Inicio de semana (0=Lun..6=Dom) [actual {getattr(cfg,'week_start_dow',0)}] (Mié=2): ").strip()
                # Validar/parsear
                tope_new = old_tope if not tope_txt else int(tope_txt)
                w_new = old_w if not w_txt else int(w_txt)
                um_new = old_um if not um_txt else int(um_txt)
                ws_new = old_ws if not ws_txt else int(ws_txt)
                if ws_new < 0 or ws_new > 6:
                    raise ValueError("week_start_dow fuera de rango (0..6)")
                if w_new < 1 or w_new > 10:
                    raise ValueError("id_min_width fuera de rango razonable (1..10)")
                if tope_new < 0 or tope_new > 240:
                    raise ValueError("tope_descuento_comida_min fuera de rango razonable (0..240)")
                if um_new < 0 or um_new > 24*60:
                    raise ValueError("umbral_extra_min fuera de rango razonable (0..1440)")
                # Commit
                cfg.tope_descuento_comida_min = tope_new
                cfg.id_min_width = w_new
                cfg.umbral_extra_min = um_new
                cfg.week_start_dow = ws_new
                print("Reglas actualizadas.")
                _autosave()
            except ValueError:
                # rollback seguro
                cfg.tope_descuento_comida_min = old_tope
                cfg.id_min_width = old_w
                cfg.umbral_extra_min = old_um
                cfg.week_start_dow = old_ws
                print("Valor inválido. No se aplicaron cambios.")
        elif op == "11":
            print("\nBuscar empleado por ID")
            emp_in = _safe_input("ID del empleado (ej. 003): ").strip()
            if not emp_in:
                print("ID vacío. Cancelado.")
                continue
            emp_id = normalize_id(emp_in, cfg.id_min_width)
            if not nombres:
                nombres.update(_cargar_nombres() or {})
            nombre = nombres.get(emp_id, "")
            grp = cfg.empleado_a_grupo.get(emp_id, "")
            idg = cfg.empleado_a_idgrupo.get(emp_id, "")
            if not nombre and not grp and not idg:
                print("No se encontró el empleado en plantilla ni en asignaciones.")

                continue
            print(f"ID: {emp_id}")
            if nombre:
                print(f"Nombre: {nombre}")
            print(f"Grupo: {grp or '-'}")
            print(f"IDGRUPO: {idg or '-'}")
        elif op == "12":
            print("\nAsignación masiva (pegar IDs)")
            ids_raw = _safe_input("Pega IDs (coma/espacio/salto): ").strip()
            if not ids_raw:
                print("Sin IDs. Cancelado.")
                continue
            ids = re.split(r"[,\s]+", ids_raw)
            ids = [x.strip() for x in ids if x and x.strip()]
            # Normalizar IDs numéricos a ancho cfg.id_min_width
            norm_ids = []
            for i in ids:
                if str(i).isdigit():
                    norm_ids.append(str(i).zfill(getattr(cfg, "id_min_width", 3)))
                else:
                    norm_ids.append(str(i))
            # Uniq manteniendo orden
            seen = set()
            norm_ids2 = []
            for i in norm_ids:
                if i in seen:
                    continue
                seen.add(i)
                norm_ids2.append(i)
            norm_ids = norm_ids2
            print(f"IDs detectados: {len(norm_ids)}")
            g = _pick_grupo("Asignar TODOS a grupo: ")
            for emp_id in norm_ids:
                cfg.empleado_a_grupo[str(emp_id)] = g
            resp = _safe_input("¿Asignar también IDGRUPO manual (uno por uno)? (S/N): ").strip().lower()
            if resp == "s":
                for emp_id in norm_ids:
                    nom = nombres.get(str(emp_id), "")
                    actual = cfg.empleado_a_idgrupo.get(str(emp_id), "")
                    prompt = f"IDGRUPO para {emp_id} {('- ' + nom) if nom else ''} (Enter=dejar {actual or 'vacío'}): "
                    val = _safe_input(prompt).strip()
                    if val:
                        cfg.empleado_a_idgrupo[str(emp_id)] = val
            _autosave()
            print("Asignación masiva guardada.")
            continue
        elif op == "13":
            print("\nBackup de configuración")
            backups_dir = script_dir / "backups"
            backups_dir.mkdir(exist_ok=True)
            try:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            except Exception:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            src_cfg = script_dir / "mapa_grupos.json"
            dst_cfg = backups_dir / f"mapa_grupos_backup_{ts}.json"
            try:
                if src_cfg.exists():
                    dst_cfg.write_text(src_cfg.read_text(encoding="utf-8"), encoding="utf-8")
                else:
                    guardar_config(script_dir, cfg)
                    dst_cfg.write_text((script_dir / "mapa_grupos.json").read_text(encoding="utf-8"), encoding="utf-8")
                print(f"Backup creado: {dst_cfg}")

            except Exception as e:
                print(f"No se pudo crear backup: {e}")
            continue
        else:
            print("Opción inválida.")
def procesar_archivo(in_path: Path, *, correccion_interactiva: bool, correcciones_eventos: Optional[Dict[Tuple[str, date], Dict[str, Tuple[Optional[time], str]]]] = None, correcciones_nolabor: Optional[Dict[Tuple[str, date], List[Tuple[Optional[time], Optional[time], str]]]] = None, plantilla_path: str = "", edicion_interactiva: bool = False, usuario_editor: str = "RRHH", modo_seguro: bool = False) -> Tuple[Path, Path]:
    script_dir = Path(__file__).resolve().parent
    cfg = cargar_config(script_dir)
    plantilla = cargar_plantilla_empleados(script_dir, plantilla_path)
    correcciones_eventos = correcciones_eventos or {}
    correcciones_nolabor = correcciones_nolabor or {}
    df_in, cols = leer_input(in_path, cfg)
    # Construir salida “base”
    audit_log: List[AuditEntry] = []
    df_out = construir_salida(
        df_in,
        cols,
        cfg,
        correcciones_nolabor=correcciones_nolabor,
        correcciones_eventos=correcciones_eventos,
        correccion_interactiva=edicion_interactiva,
        usuario_editor=usuario_editor,
        audit_log=audit_log,
        modo_seguro=modo_seguro,
    )
    # Correcciones manuales (si existe correcciones_asistencia.xlsx / --correcciones)
    if False and correcciones_eventos:
        ajustado_flags: List[str] = []
        notas_flags: List[str] = []
        for idx, r in df_out.iterrows():
            row = r.to_dict()
            row2, aplicado, nota = aplicar_correcciones(row, correcciones_eventos, id_col="ID", fecha_col="Fecha")
            ajustado_flags.append("Sí" if aplicado else "")
            notas_flags.append(nota if aplicado else "")
            # Actualiza columnas de eventos (si existen en el DF)
            for ev in CORR_EVENTOS:
                if ev in df_out.columns:
                    df_out.at[idx, ev] = row2.get(ev, "")
        # Para auditoría. Si no quieres estas columnas, se pueden desactivar.
        df_out["Ajuste manual"] = ajustado_flags
        df_out["Nota ajuste"] = notas_flags
# Aplicar orden por grupos usando el ID original (columna "ID" en df_out)
    # Creamos DF temporal para ordenar con cfg.empleado_a_grupo.
    df_out["_grp_idx"] = df_out["ID"].map(lambda x: _grupo_sort_key(str(x), cfg)[0])
    df_out = df_out.sort_values(by=["_grp_idx", "ID"], kind="stable").drop(columns=["_grp_idx"]).reset_index(drop=True)
    # IDGRUPO: agregar columna al inicio según mapping manual
    
    df_idgrupo = df_out.copy()
    # IDGRUPO: por prioridad se toma de la plantilla (si existe y está capturado); si no, usa el mapeo del JSON
    plantilla_idgrupo: Dict[str, str] = {}
    try:
        if plantilla is not None and len(plantilla) > 0 and "IDGRUPO" in plantilla.columns:
            plantilla_idgrupo = dict(
                zip(
                    plantilla["ID"].astype(str).str.strip(),
                    plantilla["IDGRUPO"].astype(str).fillna("").astype(str).str.strip(),
                )
            )
    except Exception:
        plantilla_idgrupo = {}

    def _idgrupo_of(emp_id: object) -> str:
        k = str(emp_id).strip()
        if k.startswith("NOMBRE::"):
            return ""
        v = plantilla_idgrupo.get(k, "")
        if v and v.lower() != "nan":
            return v
        # fallback: JSON mapping
        try:
            grp = cfg.empleado_grupo.get(k, "")
            if grp:
                return cfg.grupo_idgrupo.get(grp, "")
        except Exception:
            return ""
        return ""

    df_idgrupo.insert(0, "IDGRUPO", df_idgrupo["ID"].map(_idgrupo_of))
    # ordenar igual que df_out (ya viene ordenado)
    # Si corrección interactiva está activa: permitir completar mapeos faltantes
    if correccion_interactiva:
        # re-aplicar pero con prompts si faltan
        df_tmp_in = df_in.copy()
        df_tmp_in, df_tmp_idg = aplicar_grupos_y_idgrupo(df_tmp_in, cols, cfg, permitir_interactivo=True)
        # Guardar config y regenerar df_out/df_idgrupo con nuevos mapeos
        guardar_config(script_dir, cfg)

        df_out = construir_salida(df_in, cols, cfg, correcciones_nolabor=correcciones_nolabor, correcciones_eventos=correcciones_eventos)
        df_out["_grp_idx"] = df_out["ID"].map(lambda x: _grupo_sort_key(str(x), cfg)[0])
        df_out = df_out.sort_values(by=["_grp_idx", "ID"], kind="stable").drop(columns=["_grp_idx"]).reset_index(drop=True)
        df_idgrupo = df_out.copy()
        # IDGRUPO: por prioridad se toma de la plantilla (si existe y está capturado); si no, usa el mapeo del JSON
    plantilla_idgrupo = {}
    try:
        if plantilla is not None and len(plantilla) > 0 and "IDGRUPO" in plantilla.columns:
            plantilla_idgrupo = dict(zip(plantilla["ID"].astype(str).str.strip(), plantilla["IDGRUPO"].astype(str).fillna("").str.strip()))
    except Exception:
        plantilla_idgrupo = {}
    def _idgrupo_of(emp_id: object) -> str:
        k = str(emp_id).strip()
        if k.startswith("NOMBRE::"):
            return ""
        v = plantilla_idgrupo.get(k, "")
        if v and v.lower() != "nan":
            return v
        return cfg.empleado_a_idgrupo.get(k, "")
    # Asegurar columna IDGRUPO al inicio (sin duplicarla)
    if "IDGRUPO" in df_idgrupo.columns:
        df_idgrupo = df_idgrupo.drop(columns=["IDGRUPO"])
    df_idgrupo.insert(0, "IDGRUPO", df_idgrupo["ID"].map(_idgrupo_of))
    # En archivo _IDGRUPO solo se conserva IDGRUPO (sin columna ID)
    df_idgrupo = df_idgrupo.drop(columns=["ID"], errors="ignore")
# Guardar config sin tocar si no hubo cambios
    guardar_config(script_dir, cfg)
    base = in_path.with_suffix("")
    out1 = Path(str(base) + "_PROCESADO.xlsx")
    out2 = Path(str(base) + "_IDGRUPO.xlsx")
    # ---------------------------
    # Resúmenes + Faltas (Opción A: plantilla de empleados activos)
    # ---------------------------
    # plantilla ya cargada arriba
    # plantilla = cargar_plantilla_empleados(script_dir, plantilla_path)
    if plantilla is None:
        faltas_sem, faltas_mes, detalle_faltas = (pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    else:
        # Marcar IDs que aparecen en asistencia pero NO existen en la plantilla (útil para auditoría).
        try:
            ids_plant = set(plantilla["ID"].astype(str).str.strip().tolist())
            if "ID" in df_out.columns and "Notas" in df_out.columns:
                for _i, _r in df_out.iterrows():
                    _id = str(_r.get("ID","")).strip()
                    if _id and _id not in ids_plant:
                        oldn = str(df_out.at[_i, "Notas"]).strip()
                        tag = "ID no está en plantilla (no cuenta para faltas)"
                        if oldn and oldn.lower() != "nan":
                            if tag not in oldn:
                                df_out.at[_i, "Notas"] = oldn + " | " + tag
                        else:
                            df_out.at[_i, "Notas"] = tag
        except Exception:
            pass
        faltas_sem, faltas_mes, detalle_faltas = calcular_faltas(df_out, plantilla, cfg)
    resumen_semanal = construir_resumen_semanal(df_out, cfg, faltas_semanal=faltas_sem)
    resumen_semanal_vertical = construir_resumen_semanal_vertical(df_out, cfg, faltas_semanal=faltas_sem)
    resumen_mensual = construir_resumen_mensual(df_out, cfg, faltas_mensual=faltas_mes)

    # ---------------------------
        # Incidencias RRHH (revisión)
    # ---------------------------
    # Hoja profesional para RRHH: separa Motivo (código), Severidad, Acción sugerida y Evidencia,
    # además del Detalle (Notas) y el Registro crudo del export (solo aquí).
    #
    # Filosofía:
    # - No altera cálculos.
    # - No elimina opciones.
    # - Útil para filtrar y trabajar por lotes (RRHH/Admin).
    incidencias_rrhh = pd.DataFrame()
    try:
        if "Notas" in df_out.columns:
            reg_col = cols.get("registro") if isinstance(cols, dict) else None
            # Mapeo de patrones -> (codigo, severidad, accion)
            reglas = [
                (r"Registros extra ignorados:\s*(\d+)", "INC-CHK-006", "ALTA",
                 "Confirmar cuáles checadas son válidas; capturar Ajuste si aplica. Si el lector duplicó, documentar y dejar evidencia."),
                (r"Comida incompleta", "INC-CHK-002", "MEDIA",
                 "Validar si hubo comida; si aplica, capturar Ajuste (hora faltante) o justificar incidencia."),
                (r"Cena incompleta", "INC-CHK-002", "MEDIA",
                 "Validar si hubo cena; si aplica, capturar Ajuste (hora faltante) o justificar incidencia."),
                (r"NoLaborado fuera de jornada", "INC-PER-002", "MEDIA",
                 "Revisar permiso (fecha/hora). Si fue correcto, ajustar jornada con Ajuste; si fue error de captura, corregir el permiso."),
                (r"Solape interno NoLaborado", "INFO-PER-003", "BAJA",
                 "Sin acción: el sistema fusionó intervalos de permiso para evitar doble descuento (solo informativo)."),
                (r"NoLaborado solapado con comida/cena|NoLaborado solapado con comida/cena", "INFO-PER-004", "BAJA",
                 "Sin acción: el sistema evitó doble descuento con comida/cena (solo informativo)."),
            ]
            # Incidencia admin: sin IDGRUPO (no aparece en reportes por grupo)
            # Se detecta con base en el IDGRUPO calculado para *_IDGRUPO.xlsx
            def _idgrupo_of(emp_id: object) -> str:
                k = str(emp_id).strip()
                if k.startswith("NOMBRE::"):
                    return ""
                # prioridad: plantilla (si trae IDGRUPO) luego JSON
                v = ""
                try:
                    if plantilla is not None and "IDGRUPO" in plantilla.columns and "ID" in plantilla.columns:
                        m = dict(zip(plantilla["ID"].astype(str).str.strip(), plantilla["IDGRUPO"].astype(str).fillna("").str.strip()))
                        v = m.get(k, "")
                except Exception:
                    v = ""
                if v and str(v).lower() != "nan":
                    return v
                return cfg.empleado_a_idgrupo.get(k, "")
            ids = df_out["ID"].astype(str).fillna("").str.strip()
            idgrupo_series = ids.map(_idgrupo_of)
            mask_sin_idgrupo = idgrupo_series.astype(str).fillna("").str.strip().eq("")
            # Construcción de incidencias (pueden existir múltiples motivos por fila)
            incid_list = []
            # Registro crudo (solo para evidencias)
            reg_raw = None
            if reg_col and reg_col in df_in.columns:
                try:
                    reg_raw = df_in[reg_col].astype(str).fillna("").tolist()
                except Exception:
                    reg_raw = None
            for i, row in df_out.iterrows():
                notas = str(row.get("Notas", "") or "")
                if not notas or notas.lower() == "nan":
                    notas = ""
                motivos = []
                # reglas por texto
                for pat, cod, sev, accion in reglas:
                    m = re.search(pat, notas, flags=re.IGNORECASE)
                    if m:
                        evidencia = ""
                        # caso especial: >6 checadas, mostrar extras si hay registro crudo

                        if cod == "INC-CHK-006":
                            extras_n = None
                            try:
                                extras_n = int(m.group(1))
                            except Exception:
                                extras_n = None
                            rraw = (reg_raw[i] if reg_raw is not None and i < len(reg_raw) else "")
                            if rraw:
                                parts = [p.strip() for p in rraw.replace("|", ";").split(";") if p.strip()]
                                if len(parts) > 6:
                                    extras = parts[6:]
                                    evidencia = f"Extras={len(extras)}: " + ", ".join(extras[:6]) + (" ..." if len(extras) > 6 else "")
                                else:
                                    evidencia = f"Registro={rraw}"
                            else:
                                evidencia = "Checadas extra detectadas (ver Registro en export original)"
                            if extras_n is not None and evidencia:
                                evidencia = f"ExtraIgnorados={extras_n}. " + evidencia
                        else:
                            # evidencia estándar: última frase de notas o registro crudo
                            rraw = (reg_raw[i] if reg_raw is not None and i < len(reg_raw) else "")
                            if rraw:
                                evidencia = f"Registro={rraw}"
                            else:
                                evidencia = notas[:200]
                        motivos.append((cod, sev, accion, evidencia))
                # sin IDGRUPO (excepto casos SIN ID)
                if bool(mask_sin_idgrupo.iloc[i]) and (not str(row.get("ID","")).startswith("NOMBRE::")):
                    cod = "INC-GRP-001"
                    sev = "ALTA"
                    accion = "Asignar Grupo e IDGRUPO en Modo Admin y reprocesar para que aparezca en reportes por grupo."
                    evidencia = f"IDGRUPO vacío. GrupoActual={cfg.empleado_a_grupo.get(str(row.get('ID','')).strip(), '') or '—'}"
                    motivos.append((cod, sev, accion, evidencia))
                if not motivos:
                    continue
                # datos base
                base = {
                    "ID": row.get("ID", ""),
                    "Fecha": row.get("Fecha", ""),
                    "Nombre": row.get("Nombre", ""),
                    "Registro": (reg_raw[i] if reg_raw is not None and i < len(reg_raw) else ""),
                    "Detalle": notas,
                }
                # generar una fila por motivo (facilita filtrado)
                for cod, sev, accion, evidencia in motivos:
                    r = dict(base)
                    r["Motivo"] = cod
                    r["Severidad"] = sev
                    r["Acción sugerida"] = accion
                    r["Evidencia"] = evidencia
                    incid_list.append(r)
            if incid_list:
                incidencias_rrhh = pd.DataFrame(incid_list)
                # Orden lógico
                cols_order = ["Motivo","Severidad","Acción sugerida","Evidencia","ID","Fecha","Nombre","Registro","Detalle"]
                cols_order = [c for c in cols_order if c in incidencias_rrhh.columns] + [c for c in incidencias_rrhh.columns if c not in cols_order]
                incidencias_rrhh = incidencias_rrhh[cols_order]
    except Exception:
        incidencias_rrhh = pd.DataFrame()

    # --- IDs visibles (caso SIN ID) ---
    # Internamente usamos claves NOMBRE:: para poder calcular bien. Antes de exportar,
    # dejamos la columna ID en blanco para esos casos, pero conservamos el Nombre.
    df_out_export = apply_id_display(df_out, id_col="ID")
    df_idgrupo_export = df_idgrupo.copy()
    # Asegurar que IDGRUPO quede vacío para claves SIN ID (aunque exista mapeo)
    try:
        # no tenemos la clave interna en df_idgrupo (se eliminó 'ID'), así que tomamos la máscara desde df_out
        sin_id_keys = set(df_out[df_out["ID"].astype(str).str.startswith("NOMBRE::")]["ID"].astype(str).unique().tolist())
        if len(sin_id_keys) > 0 and "IDGRUPO" in df_idgrupo_export.columns:
            # sin ID no podemos cruzar, pero si la plantilla asignó por error, se limpia globalmente por seguridad
            pass
    except Exception:
        pass
    resumen_semanal_export = apply_id_display(resumen_semanal, id_col="ID")
    resumen_mensual_export = apply_id_display(resumen_mensual, id_col="ID")
    detalle_faltas_export = apply_id_display(detalle_faltas, id_col="ID")
    incidencias_rrhh_export = apply_id_display(incidencias_rrhh, id_col="ID")
    extra_sheets = {
        "RESUMEN_SEMANAL": resumen_semanal_export,
        "RESUMEN_MENSUAL": resumen_mensual_export,
        "DETALLE_FALTAS": detalle_faltas_export,
        "INCIDENCIAS_RRHH": incidencias_rrhh_export,
    }
    # Hoja adicional: resumen semanal de checadas (control interno)
    try:
        df_chec_sem_proc = crear_resumen_semanal_checadas(df_out_export, cfg, modo="PROCESADO")
        if df_chec_sem_proc is not None and len(df_chec_sem_proc) > 0:
            extra_sheets["RESUM_SEM_CHECADAS"] = df_chec_sem_proc
    except Exception:
        pass
    # Hoja adicional: CONTROL (resumen ejecutivo + listas)
    try:
        control_rows = []
        control_rows.append({"Sección":"RESUMEN","Campo":"Filas en Reporte","Valor":len(df_out_export)})
        control_rows.append({"Sección":"RESUMEN","Campo":"Incidencias RRHH","Valor":len(incidencias_rrhh_export)})
        # Faltas
        try:
            control_rows.append({"Sección":"RESUMEN","Campo":"Faltas (detalle)","Valor":len(detalle_faltas_export)})
        except Exception:
            pass
        # Permisos/NoLaborado aplicados (conteo por nota)
        try:
            n_nolab = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("NoLaborado", na=False).sum())
            control_rows.append({"Sección":"RRHH","Campo":"Registros con NoLaborado (menciones)","Valor":n_nolab})
        except Exception:
            pass
        # Ajustes manuales aplicados
        try:
            if "Ajuste manual" in df_out_export.columns:
                n_adj = int(df_out_export["Ajuste manual"].astype(str).str.upper().isin(["SI","SÍ","TRUE","1"]).sum())
                control_rows.append({"Sección":"RRHH","Campo":"Ajustes manuales aplicados","Valor":n_adj})
        except Exception:
            pass
        # Incidencias por motivo/severidad (top)
        try:
            if {"Motivo","Severidad"}.issubset(set(incidencias_rrhh_export.columns)):
                piv = incidencias_rrhh_export.pivot_table(index=["Motivo","Severidad"], values="ID", aggfunc="count").reset_index()
                piv = piv.rename(columns={"ID":"Cantidad"}).sort_values("Cantidad", ascending=False)
                for _, r in piv.head(15).iterrows():
                    control_rows.append({"Sección":"INCIDENCIAS","Campo":f"{r['Motivo']}|{r['Severidad']}", "Valor":int(r["Cantidad"])})
        except Exception:
            pass
        # Empleados SIN ID (internos): lista
        try:
            n_sin = len(sin_id_keys) if 'sin_id_keys' in locals() else 0
            control_rows.append({"Sección":"SIN_ID","Campo":"Total SIN ID","Valor":n_sin})
            if 'sin_id_keys' in locals() and n_sin>0:
                # listar hasta 50 para no hacer el archivo enorme
                for k in list(sorted(sin_id_keys))[:50]:
                    nombre = str(k).replace("NOMBRE::","")
                    control_rows.append({"Sección":"SIN_ID","Campo":"Nombre","Valor":nombre})

                if n_sin>50:
                    control_rows.append({"Sección":"SIN_ID","Campo":"Nota","Valor":"(Lista truncada a 50)"})
        except Exception:
            pass
        df_control = pd.DataFrame(control_rows)
    except Exception:
        df_control = pd.DataFrame()
    if df_control is not None and len(df_control)>0:
        extra_sheets["CONTROL"] = df_control
    # Preparar hojas extra por archivo (PROCESADO vs IDGRUPO)
    extra_sheets_proc = dict(extra_sheets)
    extra_sheets_idg = dict(extra_sheets)
    # En el archivo por IDGRUPO, regenerar el resumen semanal de checadas sin columna ID
    try:
        df_chec_sem_idg = crear_resumen_semanal_checadas(df_idgrupo, cfg, modo="IDGRUPO")
        if df_chec_sem_idg is not None and len(df_chec_sem_idg) > 0:
            extra_sheets_idg["RESUM_SEM_CHECADAS"] = df_chec_sem_idg
    except Exception:
        pass
    exportar_excel(df_out_export, out1, extra_sheets=extra_sheets_proc)
    exportar_excel(df_idgrupo, out2, extra_sheets=extra_sheets_idg)
    # Guardar auditoría de cambios (si hubo)
    try:
        aud_path = Path(str(base) + "_auditoria_cambios.json")
        guardar_auditoria_json(aud_path, audit_log)
    except Exception:
        pass
    # Log simple de ejecución (auditoría)
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = Path(str(base) + f"_LOG_{ts}.txt")
        n_rows = len(df_out)
        n_extra = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("Registros extra ignorados", na=False).sum())
        n_ov = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("Solape NoLaborado con comida/cena", na=False).sum())
        n_intov = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("Solape interno NoLaborado", na=False).sum())
        n_out = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("NoLaborado fuera de jornada", na=False).sum())
        n_com_inc = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("Comida incompleta", na=False).sum())
        n_cen_inc = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("Cena incompleta", na=False).sum())
        with open(log_path, "w", encoding="utf-8") as f:
            f.write("LOG DE EJECUCIÓN - PROCESADOR ASISTENCIAS\n")
            f.write(f"Entrada: {in_path}\n")
            f.write(f"Salida 1: {out1}\n")
            f.write(f"Salida 2: {out2}\n")
            f.write(f"Filas procesadas: {n_rows}\n")
            f.write(f"Con registros extra (>6): {n_extra}\n")
            f.write(f"Con solape NoLaborado vs comida/cena: {n_ov}\n")
            f.write(f"Con solape interno NoLaborado: {n_intov}\n")
            f.write(f"Con NoLaborado fuera de jornada: {n_out}\n")
            f.write(f"Comida incompleta: {n_com_inc}\n")
            f.write(f"Cena incompleta: {n_cen_inc}\n")
            if plantilla is None:
                f.write("Plantilla empleados: NO (faltas no calculadas)\n")
            else:
                f.write("Plantilla empleados: SI (faltas calculadas)\n")
            f.write("\nNotas: Este log es informativo y no altera cálculos.\n")
    except Exception:
        pass
    return out1, out2

def cargar_plantilla_empleados(script_dir: Path, ruta: str = "") -> "pd.DataFrame | None":
    """Carga plantilla de empleados activos (Opción A).
    Mínimo: columna ID. Opcionales: Nombre, Activo (SI/NO), FechaAlta, FechaBaja.
    """
    path = Path(ruta) if ruta else (script_dir / "plantilla_empleados.xlsx")
    if not path.exists():
        print(f"[AVISO] No se encontró la plantilla de empleados: {path}.")
        print("        Se omitirá el cálculo de FALTAS.")
        print("        (Solución: coloca plantilla_empleados.xlsx junto al script o usa --plantilla.)")
        return None
    try:
        df = pd.read_excel(path, dtype=str)
    except Exception:
        print(f"[AVISO] No se pudo leer la plantilla de empleados: {path}.")
        print("        Se omitirá el cálculo de FALTAS.")
        return None
    df.columns = [str(c).replace('\ufeff','').strip() for c in df.columns]
    id_col = None
    for c in df.columns:
        if str(c).strip().lower() in ("id","id_empleado","empleado","employeeid"):
            id_col = c
            break
    if id_col is None:
        return pd.DataFrame()
    df = df.rename(columns={id_col:"ID"})
    df["ID"] = df["ID"].astype(str).str.strip()
    df["ID"] = df["ID"].apply(lambda x: _coerce_id_str(x, 3))
    # Caso SIN ID: si ID es texto (nombre) lo convertimos a clave interna NOMBRE:: para que coincida con el export.
    df["ID"] = df.apply(lambda r: make_emp_key(r.get("ID",""), r.get("Nombre",""), 3)[0], axis=1)
    if "Nombre" not in df.columns:
        for c in df.columns:
            if str(c).strip().lower() in ("nombre","name"):
                df = df.rename(columns={c:"Nombre"})
                break
    if "Nombre" not in df.columns:
        df["Nombre"] = ""
    
    # Columna opcional: IDGRUPO (ej. FT-123). Si existe, se conserva para alimentar el reporte *_IDGRUPO.xlsx
    if "IDGRUPO" not in df.columns:
        for c in df.columns:
            if str(c).strip().lower() in ("idgrupo","id_grupo","grupo","grupo_id","id grupo"):
                df = df.rename(columns={c:"IDGRUPO"})
                break
    if "IDGRUPO" not in df.columns:
        df["IDGRUPO"] = ""
    else:
        df["IDGRUPO"] = df["IDGRUPO"].astype(str).fillna("").str.strip()
    if "Activo" not in df.columns:
        for c in df.columns:
            if str(c).strip().lower() in ("activo","active","estatus","status"):
                df = df.rename(columns={c:"Activo"})
                break
    if "Activo" not in df.columns:
        df["Activo"] = "SI"
    def _is_active(x):
        s = str(x).strip().lower()
        return s in ("si","sí","s","1","true","activo","active","yes","y")
    df["_activo"] = df["Activo"].apply(_is_active)
    for col in ("FechaAlta","FechaBaja"):
        if col not in df.columns:
            for c in df.columns:
                if str(c).strip().lower() == col.lower():
                    df = df.rename(columns={c:col})
                    break
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date
        else:
            df[col] = pd.NaT
    df = df[df["ID"] != ""]
    return df
def _preguntar_confirmacion_correcciones(corr_eventos, corr_nolabor, path_corr: Path) -> bool:
    """Devuelve True si se deben aplicar correcciones; False para ignorarlas."""
    try:
        import sys

        is_tty = hasattr(sys.stdin, "isatty") and sys.stdin.isatty()
    except Exception:
        is_tty = True
    n_ajustes = sum(len(v) for v in (corr_eventos or {}).values())
    n_nolabor = sum(len(v) for v in (corr_nolabor or {}).values())
    total = n_ajustes + n_nolabor
    if total <= 0:
        return False  # no hay nada que aplicar
    if not is_tty:
        # En ejecuciones no interactivas, aplicar por defecto.
        print(f"[INFO] Correcciones detectadas ({total}) en {path_corr}. Ejecución no interactiva: se aplicarán por defecto.")
        return True
    print("\n=== Confirmación de correcciones manuales ===")
    print(f"Archivo: {path_corr}")
    print(f"Ajustes de eventos (Entrada/Comida/Cena/Salida): {n_ajustes}")
    print(f"Permisos NoLaborado (tiempo no laborado): {n_nolabor}")
    print("¿Deseas APLICAR estas correcciones al procesar? (S/N)")
    while True:
        resp = input("> ").strip().lower()
        if resp in ("s", "si", "sí", "y", "yes"):
            return True
        if resp in ("n", "no"):
            print("Se procesará SIN aplicar correcciones.")
            return False
        print("Respuesta no válida. Escribe S o N.")
def main() -> int:
    parser = argparse.ArgumentParser(
        description="Procesa exportaciones de asistencia (HikCentral/iVMS) y genera 2 archivos:\n"
                    "  1) *_PROCESADO.xlsx (por empleado y día)\n"
                    "  2) *_IDGRUPO.xlsx (misma salida, ordenada por el orden de grupos configurado)",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument("archivo", nargs="?", help="Ruta al Excel/CSV exportado (con columnas tipo: ID, Fecha, Semana, Nombre, Pases, Registro).")
    parser.add_argument("--correccion", action="store_true",
                        help="Modo corrección/alta manual (asignación de grupo e ID de grupo para empleados nuevos).")
    parser.add_argument("--admin", action="store_true",
                        help="Abre el menú de administración (gestión de grupos/IDs) y termina.")
    parser.add_argument("--confirmar-correcciones", action="store_true", help="Pide confirmación antes de aplicar correcciones del archivo.")
    parser.add_argument("--correccion-interactiva", action="store_true",
                    help="Durante el procesamiento por lote, pregunta si deseas editar checadas por ID/Fecha.")
    parser.add_argument("--usuario-editor", type=str, default="RRHH",
                    help="Etiqueta del usuario que autoriza correcciones (para auditoría).")
    parser.add_argument("--modo-seguro", action="store_true",
                    help="No aplica normalización automática (p. ej. reorden por cruce medianoche) sin edición humana.")
    parser.add_argument("--correcciones", default=None,
                        help="Ruta a correcciones manuales (xlsx/csv). Si no se indica, se busca 'correcciones_asistencia.xlsx' junto al script.")
    parser.add_argument("--plantilla-correcciones", action="store_true",
                        help="Genera una plantilla 'correcciones_asistencia.xlsx' junto al script y termina.")
    parser.add_argument("--demo", action="store_true",
                        help="Genera un archivo demo (test_asistencia_demo.xlsx) con datos ficticios tipo HikCentral y termina.")
    parser.add_argument("--selftest", action="store_true",
                        help="Ejecuta una prueba rápida con el archivo demo para validar que el procesamiento corre sin errores.")
    parser.add_argument("--plantilla", type=str, default="",
        help="Ruta a plantilla_empleados.xlsx (opcional). Si se omite, se busca plantilla_empleados.xlsx junto al script.")
    args = parser.parse_args()
    script_dir = Path(__file__).resolve().parent
    if args.admin:
        modo_administracion(script_dir, args.plantilla)
        return 0
    # ---------------------------
    # Demo / Selftest

    # ---------------------------
    if args.demo or args.selftest:
        demo_path = script_dir / "test_asistencia_demo.xlsx"
        demo_df = pd.DataFrame([
            {"ID":"003","Fecha":"2026-01-19","Semana":"Lunes","Nombre":"Ana Perez","Pases":"6","Registro":"09:10; 09:29; 09:30; 09:35; 09:40; 09:42"},
            {"ID":"004","Fecha":"2026-01-19","Semana":"Lunes","Nombre":"Luis Gomez","Pases":"2","Registro":"08:00; 17:00"},
            {"ID":"005","Fecha":"2026-01-19","Semana":"Lunes","Nombre":"Marta Ruiz","Pases":"5","Registro":"07:55; 12:00; 12:45; 16:00; 16:10"},
            {"ID":"006","Fecha":"2026-01-19","Semana":"Lunes","Nombre":"Carlos Soto","Pases":"7","Registro":"06:00; 12:00; 15:00; 19:00; 21:30; 04:00; 09:00"},
        ])
        demo_df.to_excel(demo_path, index=False)
        print(f"OK. Demo generado: {demo_path}")
        if args.demo:
            return 0
        # selftest
        try:
            out1, out2 = procesar_archivo(demo_path, correccion_interactiva=False, correcciones_eventos=None, correcciones_nolabor=None, plantilla_path=args.plantilla, edicion_interactiva=False, usuario_editor="SELFTEST", modo_seguro=False)
            print(f"OK. Selftest generado:\n - {out1}\n - {out2}")
            return 0
        except Exception as e:
            print(f"ERROR en selftest: {e}")
            return 2
    if args.plantilla_correcciones:
        tplA = pd.DataFrame(columns=["ID", "Fecha", "Evento", "Hora", "Nota"])
        tplN = pd.DataFrame(columns=["ID", "Fecha", "Inicio", "Fin", "Nota"])
        outp = script_dir / "correcciones_asistencia.xlsx"
        with pd.ExcelWriter(outp, engine="openpyxl") as writer:
            tplA.to_excel(writer, index=False, sheet_name="Ajustes")
            tplN.to_excel(writer, index=False, sheet_name="NoLaborado")
        print(f"Plantilla creada: {outp}")
        return 0
    if not args.archivo:
        raise SystemExit("Debes indicar el archivo exportado. Ejemplo: python script.py export.xlsx")
    in_path = Path(args.archivo).expanduser().resolve()
    if not in_path.exists():
        raise SystemExit(f"No existe el archivo: {in_path}")
    # Correcciones manuales (opcional)
    path_corr = Path(args.correcciones).expanduser().resolve() if args.correcciones else (script_dir / "correcciones_asistencia.xlsx")
    try:
        correcciones_eventos, correcciones_nolabor = cargar_correcciones(path_corr)
        if correcciones_eventos or correcciones_nolabor:
            print(f"Correcciones cargadas: {path_corr}")
        else:
            if path_corr.exists():
                print(f"Archivo de correcciones encontrado pero sin filas aplicables: {path_corr}")
    except Exception as e:
        correcciones_eventos, correcciones_nolabor = {}, {}
        print(f"Advertencia: no se pudieron leer correcciones ({path_corr}). Se continúa sin correcciones. Detalle: {e}")
        correcciones_eventos, correcciones_nolabor = {}, {}
    
    # Confirmación opcional antes de aplicar correcciones
    if args.confirmar_correcciones and (correcciones_eventos or correcciones_nolabor):
        aplicar = _preguntar_confirmacion_correcciones(correcciones_eventos, correcciones_nolabor, path_corr)
        if not aplicar:
            correcciones_eventos, correcciones_nolabor = {}, {}
    out1, out2 = procesar_archivo(
        in_path,
        correccion_interactiva=args.correccion,
        correcciones_eventos=correcciones_eventos,
        correcciones_nolabor=correcciones_nolabor,
        plantilla_path=args.plantilla,
        edicion_interactiva=args.correccion_interactiva,
        usuario_editor=args.usuario_editor,
        modo_seguro=args.modo_seguro,
    )
    print(f"OK. Generado:\n - {out1}\n - {out2}")
    return 0
if __name__ == "__main__":
    main()
