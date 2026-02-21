"""I/O: export multi-hoja y backups (v18-compatible)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd

import re as _re


from .logger import log_exception
from .utils import chmod_restringido, normalize_id
import logging
import re
from .config import AppConfig
from .groups import sort_df_by_group


# Columnas técnicas (solo para dashboard/diagnóstico) que NO deben exportarse a Excel final.
_EXPORT_DROP_COLS = {
    "registro original",
    "registros parseados",
    "registros normalizados",
    "fuente checadas",
    "discrepancias",
}

def _drop_export_debug_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Retorna una vista del DF sin columnas técnicas.

    Mantiene el DF original intacto para dashboard/validación.
    Drop es case-insensitive y tolerante a espacios.
    """
    if df is None or df.empty:
        return df
    cols_norm = {c: " ".join(str(c).strip().lower().split()) for c in df.columns}
    to_drop = [c for c, n in cols_norm.items() if n in _EXPORT_DROP_COLS]
    if not to_drop:
        return df
    return df.drop(columns=to_drop, errors="ignore")


# Protege contra inyección de fórmulas en Excel (Excel/CSV Injection).
# Si un campo de texto comienza con =, +, -, @, Excel puede interpretarlo como fórmula.
# Mitigación: anteponer un apóstrofe (') solo en strings que inician con esos caracteres.
_EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@")

def _sanitize_excel_injection(df: pd.DataFrame) -> pd.DataFrame:
    """Retorna una copia del DF con strings sanitizadas para Excel.

    - Solo afecta columnas tipo object/string.
    - No modifica NaN/None.
    - No altera valores numéricos.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        if str(out[col].dtype) not in ("object", "string"):
            continue
        ser = out[col]
        def _fix(v):
            if v is None:
                return v
            if isinstance(v, float) and pd.isna(v):
                return v
            s = str(v)
            if s.startswith(_EXCEL_FORMULA_PREFIXES):
                return "'" + s
            return v
        out[col] = ser.map(_fix)
    return out


def _find_id_col_for_sort(df: pd.DataFrame) -> str | None:
    """Devuelve el nombre de la columna ID más probable para ordenar, o None si no existe."""
    if df is None or len(df.columns) == 0:
        return None
    candidates = [
        "ID", "Id", "id",
        "ID Empleado", "ID empleado", "id empleado",
        "ID_EMPLEADO", "id_empleado",
        "EmpleadoID", "empleado_id",
    ]
    cols = list(df.columns)
    # exact match
    for c in candidates:
        if c in cols:
            return c
    # normalized match
    def norm(x: str) -> str:
        return " ".join(str(x).strip().lower().split()).replace("_"," ")
    ncols = {c: norm(c) for c in cols}
    cand_norm = {norm(c) for c in candidates}
    for c, n in ncols.items():
        if n in cand_norm:
            return c
    return None

def _sort_for_group_order_export(df: pd.DataFrame, cfg: "AppConfig | None") -> pd.DataFrame:
    """Ordena filas por orden de grupos (cfg.grupos_orden) y luego por ID/Fecha.
    Solo aplica si existe columna ID y hay configuración de grupos.
    No modifica el df original.
    """
    try:
        if cfg is None or df is None or len(df) == 0:
            return df
        id_col = _find_id_col_for_sort(df)
        if not id_col:
            return df
        grupos_orden = getattr(cfg, "grupos_orden", []) or []
        emp_a_grupo = getattr(cfg, "empleado_a_grupo", {}) or {}
        if not grupos_orden or not emp_a_grupo:
            return df

        g2i = {g: i for i, g in enumerate(grupos_orden)}

        ids = df[id_col].astype(str).str.strip()
        grupos = ids.map(lambda x: emp_a_grupo.get(x, ""))
        idxs = grupos.map(lambda g: g2i.get(g, 9999))

        out = df.copy()
        out["_grp_idx__"] = idxs

        # Si existe columna IDGRUPO en la hoja, úsala; si no, deriva desde cfg.empleado_a_idgrupo
        idg_col = None
        for cand in ("IDGRUPO", "ID_GRUPO", "idgrupo", "id_grupo"):
            if cand in out.columns:
                idg_col = cand
                break
        if idg_col:
            idg_key = out[idg_col].astype(str).str.strip()
        else:
            emp_a_idg = getattr(cfg, "empleado_a_idgrupo", {}) or {}
            if emp_a_idg:
                idg_key = out[id_col].astype(str).str.strip().map(lambda x: emp_a_idg.get(x, ""))
            else:
                idg_key = None

        if idg_key is not None:
            out["_idgrupo_key__"] = idg_key
            out["_idgrupo_key_norm__"] = out["_idgrupo_key__"].map(_normalize_idgrupo_sort_key)
            sort_cols = ["_idgrupo_key_norm__", "_grp_idx__", id_col]
        else:
            sort_cols = ["_grp_idx__", id_col]

        if "Fecha" in out.columns:
            sort_cols.append("Fecha")

        out = out.sort_values(sort_cols, kind="mergesort").drop(columns=["_grp_idx__", "_idgrupo_key__", "_idgrupo_key_norm__"], errors="ignore")
        return out
    except Exception:
        return df



def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def backup_if_exists(path: Path) -> Optional[Path]:
    """Crea un backup timestamp en ./backups si el archivo existe."""
    if not path.exists():
        return None
    bdir = path.parent / "backups"
    bdir.mkdir(parents=True, exist_ok=True)
    try:
        chmod_restringido(bdir)
    except Exception:
        pass
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bkp = bdir / f"{path.stem}_backup_{ts}{path.suffix}"
    bkp.write_bytes(path.read_bytes())
    try:
        chmod_restringido(bkp)
    except Exception:
        pass
    return bkp

def exportar_excel(df: pd.DataFrame, out_path: Path, extra_sheets: dict = None, cfg: 'AppConfig | None' = None) -> None:
    """Exporta un DataFrame principal y hojas extra a un XLSX.

    Formato aplicado (determinístico / estable):
    - Congela encabezados (freeze panes A2)
    - Encabezados en negritas
    - AutoFiltro en el rango usado
    - Ancho de columnas determinístico (por nombre de columna; fallback por longitud de header)
    - Columna ID como texto (preserva ceros)
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    _ensure_dir(out_path.parent)
    extra_sheets = extra_sheets or {}
    cfg = cfg or AppConfig()

    def _expected_width(header: object) -> float:
        h = "" if header is None else str(header)
        # exact match
        if h in (cfg.column_widths or {}):
            return float(cfg.column_widths[h])
        # patterns
        for item in (cfg.column_width_patterns or []):
            try:
                pat = str(item.get("pattern", ""))
                w = float(item.get("width", 0) or 0)
                if pat and re.search(pat, h):
                    return float(w)
            except Exception:
                continue
        # fallback by header length
        return float(min(45, max(10, len(h) + 2)))

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        _sanitize_excel_injection(_drop_export_debug_cols(_sort_for_group_order_export(df, cfg))).to_excel(writer, index=False, sheet_name="Reporte")

        for sheet_name, sdf in extra_sheets.items():
            if sdf is None or len(sdf) == 0:
                continue
            sname = str(sheet_name)[:31]
            _sanitize_excel_injection(_drop_export_debug_cols(_sort_for_group_order_export(sdf, cfg))).to_excel(writer, index=False, sheet_name=sname)

        header_font = Font(bold=True)

        for sname, ws in writer.sheets.items():
            ws.freeze_panes = "A2"

            for c in range(1, ws.max_column + 1):
                ws.cell(row=1, column=c).font = header_font

            try:
                last_col = get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
            except Exception:
                log_exception("No se pudo aplicar auto_filter", level=logging.DEBUG)

            # Normalize ID as TEXT if present
            try:
                headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                if "ID" in headers:
                    id_col = headers.index("ID") + 1
                    ws.cell(row=1, column=id_col).number_format = "@"
                    for r in range(2, ws.max_row + 1):
                        cell = ws.cell(row=r, column=id_col)
                        if cell.value is None:
                            continue
                        cell.value = normalize_id(cell.value, 3)
                        cell.number_format = "@"
            except Exception:
                log_exception("Fallo best-effort de formateo en Excel", level=logging.DEBUG)

            # Deterministic widths based on header name
            try:
                headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                for idx, header in enumerate(headers, start=1):
                    letter = get_column_letter(idx)
                    h = "" if header is None else str(header)
                    ws.column_dimensions[letter].width = float(_expected_width(h))
            except Exception:
                log_exception("No se pudo aplicar anchos de columna", level=logging.DEBUG)

    try:
        chmod_restringido(out_path)
    except Exception:
        log_exception("No se pudo endurecer permisos del archivo de salida", level=logging.DEBUG)

def _normalize_idgrupo_sort_key(s: str) -> tuple:
    """Clave numérica para ordenar IDGRUPO (evita 000-05 antes de 000-01)."""
    if s is None:
        return (999999, 999999, "")
    st = str(s).strip()
    if not st:
        return (999999, 999999, "")
    st = st.split()[0]
    m = _re.match(r"^\s*(\d+)\s*-\s*(\d+)(?:\s*-\s*([A-Za-z0-9]+))?\s*$", st)
    if m:
        return (int(m.group(1)), int(m.group(2)), (m.group(3) or ""))
    nums = _re.findall(r"\d+", st)
    if len(nums) >= 2:
        return (int(nums[0]), int(nums[1]), "")
    if len(nums) == 1:
        return (int(nums[0]), 0, "")
    return (999999, 999999, st)
