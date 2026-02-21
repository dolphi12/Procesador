"""Pipeline de procesamiento (v18-compatible) sin depender del legacy."""

from __future__ import annotations

import copy
import re
from datetime import datetime, timedelta, time, date
import uuid
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd

def _read_csv_flexible(path: Path) -> pd.DataFrame:
    """Lee CSV/TXT de forma tolerante a encodings comunes.
    - Intenta utf-8-sig, utf-8, latin1.
    - Usa sep=None (sniff) con engine=python para tolerar delimitadores.
    """
    last_err: Exception | None = None
    for enc in ("utf-8-sig", "utf-8", "latin1"):
        try:
            return pd.read_csv(path, dtype=str, encoding=enc, sep=None, engine="python")
        except Exception as e:
            last_err = e
            continue
    # Si todos fallan, re-lanzar el último error para diagnóstico
    assert last_err is not None
    raise last_err

import openpyxl
from .logger import log_exception

from .config import AppConfig, cargar_config, guardar_config
from .utils import _coerce_id_str, _guess_column, fmt_hhmm, hhmm_to_minutes, minutes_to_hhmm, normalize_id, sha256_file
from .parsers import parse_date, parse_registro, parse_time
from .core import normalize_registro_times, map_eventos, calcular_trabajado, minutos_entre
from .io import exportar_excel, backup_if_exists
from .faltas import cargar_plantilla_empleados, calcular_faltas
from .groups import aplicar_grupos_y_idgrupo, build_idgrupo_label, transform_sheet_procesado, transform_sheet_idgrupo, make_emp_key, apply_id_display, _grupo_sort_key
from .summaries import (
    construir_resumen_semanal,
    construir_resumen_mensual,
    construir_resumen_semanal_vertical,
    crear_resumen_semanal_checadas,
)
import logging
from dataclasses import dataclass
from typing import Any
from .corrections import AuditEntry, editar_checadas_interactivo, guardar_auditoria_bundle, guardar_auditoria_json

def leer_input(path: Path, cfg: AppConfig) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """
    Lee el archivo y retorna:
      - DataFrame con columnas originales
      - dict canon->col_real: {'id','fecha','semana','nombre','pases','registro'}
    """
    if path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]:
        df = pd.read_excel(path, dtype=str)  # dtype=str para no perder ceros
    else:
        df = _read_csv_flexible(path)
    # Normalizar encabezados (quita BOM/espacios)
    df.columns = [str(c).replace('\ufeff','').strip() for c in df.columns]
    # Detectar columnas con tolerancia
    col_id = _guess_column(df, ["ID", "Id", "Empleado", "No. empleado", "No empleado"])
    col_fecha = _guess_column(df, ["Fecha", "Fecha de registro", "Fecha registro", "Date"])
    col_semana = _guess_column(df, ["Semana", "Día", "Dia", "Día de la semana", "Day"])
    col_nombre = _guess_column(df, ["Nombre", "Name", "Empleado", "Colaborador"])
    col_pases = _guess_column(df, ["Número de pases de la tarjeta", "Numero de pases de la tarjeta", "Pases", "Numero de pases", "Pases de la tarjeta"])
    col_registro = _guess_column(df, ["Registro", "Registros", "Registro de asistencia", "Registro de tiempo", "Time", "Punches"])
    missing = []
    for key, col in [("ID", col_id), ("Fecha", col_fecha), ("Nombre", col_nombre), ("Pases", col_pases), ("Registro", col_registro)]:
        if col is None:
            missing.append(key)

    if missing:
        raise KeyError(
            f"Faltan columnas requeridas: {', '.join(missing)}.\n"
            f"Columnas encontradas en el archivo: {list(df.columns)}\n"
            f"Tip: Asegúrate de exportar con columnas ID/Fecha/Nombre/Pases/Registro (o equivalentes)."
        )
    # Coerciones mínimas
    df[col_id] = df[col_id].map(lambda v: _coerce_id_str(v, cfg.id_min_width))
    # Caso SIN ID: cuando ID == Nombre (texto), se crea una clave interna NOMBRE:: y se deja ID visible en blanco.    # Guardamos el ID visible en columna auxiliar para usarla al exportar.
    df["_ID_DISPLAY"] = ""
    keys = []
    disps = []
    for _idv, _nv in zip(df[col_id].tolist(), df[col_nombre].tolist()):
        k, disp = make_emp_key(_idv, _nv, cfg.id_min_width)
        keys.append(k)
        disps.append(disp)
    df["_EMP_KEY"] = keys
    df["_ID_DISPLAY"] = disps
    df[col_id] = df["_EMP_KEY"]
    if col_pases:
        # Dejar como entero si es posible (pero conservamos dtype=str)
        df[col_pases] = df[col_pases].fillna("").map(lambda x: re.sub(r"\.0$", "", str(x).strip()))
    return df, {"id": col_id, "fecha": col_fecha, "semana": col_semana, "nombre": col_nombre, "pases": col_pases, "registro": col_registro}


@dataclass
class VerifyReport:
    ok: bool
    checks: list[str]
    errors: list[str]

    @property
    def summary(self) -> str:
        if self.ok:
            return "OK"
        return "; ".join(self.errors[:8])


def _norm_headers(row) -> list[str]:
    out = []
    for c in (row or []):
        if c is None:
            continue
        s = str(c).strip()
        if not s:
            continue
        out.append(s.upper())
    return out


def _headers_of_sheet(path: Path, sheet: str) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        return _norm_headers(row)
    finally:
        wb.close()


def _sheet_headers_ok(path: Path, sheet: str) -> bool:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        wb.close()
        if not row:
            return False
        # al menos 2 headers no vacíos
        nonempty = [c for c in row if c not in (None, "", " ")]
        return len(nonempty) >= 2
    except Exception:
        return False


def verify_outputs(*, out_procesado: Path, out_idgrupo: Path, out_dir: Path, script_dir: Path, cfg: Any) -> VerifyReport:
    """Verifica que una corrida sea 'válida' para producción.

    Checks:
    - existen archivos _PROCESADO y _IDGRUPO
    - contienen hoja CONTROL
    - ambos tienen el mismo set de hojas
    - existe auditoria/latest.json y el bundle
    - si audit_signing_enabled, la firma del bundle es válida
    """
    checks: list[str] = []
    errors: list[str] = []

    if not out_procesado.exists():
        errors.append(f"Falta PROCESADO: {out_procesado.name}")
    else:
        checks.append("OK PROCESADO")

    if not out_idgrupo.exists():
        errors.append(f"Falta IDGRUPO: {out_idgrupo.name}")
    else:
        checks.append("OK IDGRUPO")

    # Sheets
    if out_procesado.exists() and out_idgrupo.exists():
        try:
            wb1 = openpyxl.load_workbook(out_procesado, read_only=True, data_only=True)
            wb2 = openpyxl.load_workbook(out_idgrupo, read_only=True, data_only=True)
            s1 = set(wb1.sheetnames)
            s2 = set(wb2.sheetnames)
            wb1.close()
            wb2.close()

            if "CONTROL" not in s1:
                errors.append("PROCESADO sin hoja CONTROL")
            if "CONTROL" not in s2:
                errors.append("IDGRUPO sin hoja CONTROL")
            allowed_idgrupo_extra = {"IDGRUPO"}
            # Si se habilita split por grupo, también se permiten hojas GRUPO_XXX
            if bool(getattr(cfg, "excel_idgrupo_split_by_group", False)):
                allowed_idgrupo_extra |= {s for s in s2 if str(s).startswith("GRUPO_")}

            extra_in_proc = s1 - s2
            extra_in_idg = s2 - s1
            # PROCESADO no debe tener hojas "extra" respecto a IDGRUPO
            if extra_in_proc:
                errors.append("PROCESADO/IDGRUPO con hojas diferentes")
            # IDGRUPO puede tener hojas extra permitidas
            elif any(s not in allowed_idgrupo_extra for s in extra_in_idg):
                errors.append("PROCESADO/IDGRUPO con hojas diferentes")
            else:
                checks.append("OK hojas compatibles")

            # Header sanity (CONTROL + una hoja operativa cualquiera)
            if "CONTROL" in s1 and not _sheet_headers_ok(out_procesado, "CONTROL"):
                errors.append("CONTROL sin encabezados válidos (PROCESADO)")
            if "CONTROL" in s2 and not _sheet_headers_ok(out_idgrupo, "CONTROL"):
                errors.append("CONTROL sin encabezados válidos (IDGRUPO)")
        except Exception:
            errors.append("No se pudieron leer hojas (xlsx corrupto o bloqueado)")

    # Auditoría
    audit_dir = Path(out_dir) / "auditoria"
    latest = audit_dir / "latest.json"
    if not latest.exists():
        errors.append("Falta auditoria/latest.json")
    else:
        checks.append("OK latest.json")
        try:
            obj = json.loads(latest.read_text(encoding="utf-8"))
            bundle = audit_dir / str(obj.get("bundle") or "")
            if not bundle.exists():
                errors.append("Falta bundle de auditoría")
            else:
                checks.append("OK bundle")
                if bool(getattr(cfg, "audit_signing_enabled", True)):
                    from .corrections import verificar_auditoria_bundle
                    ok = verificar_auditoria_bundle(bundle, script_dir=script_dir, cfg=cfg)
                    if not ok:
                        errors.append("Firma de auditoría inválida")
                    else:
                        checks.append("OK firma auditoría")
        except Exception:
            errors.append("No se pudo validar auditoría (latest/bundle)")

    return VerifyReport(ok=(len(errors) == 0), checks=checks, errors=errors)

def aplicar_correcciones(row: Dict[str, object],
                         correcciones: Dict[Tuple[str, date], Dict[str, Tuple[Optional[time], str]]],
                         id_col: str,
                         fecha_col: str) -> Tuple[Dict[str, object], bool, str]:
    """
    Aplica correcciones sobre un dict de salida (row). Devuelve:
      - row corregido
      - flag de si se aplicó algo
      - nota concatenada
    """
    rid = str(row.get(id_col, "")).strip()
    f = row.get(fecha_col)
    if not rid or not isinstance(f, date):
        return row, False, ""
    key = (rid, f)
    if key not in correcciones:
        return row, False, ""
    aplicado = False
    notas = []
    for ev, (hh, nota) in correcciones[key].items():
        if ev not in row:
            # si la columna no existe, la crea (por seguridad)
            row[ev] = ""
        old = row.get(ev, "")
        if hh is None:
            # borrar
            if old != "":
                row[ev] = ""
                aplicado = True
        else:
            new = fmt_hhmm(hh)
            if str(old).strip() != new:
                row[ev] = new
                aplicado = True
        if nota:
            notas.append(f"{ev}: {nota}")
    return row, aplicado, " | ".join(notas)


def construir_salida(
    df: pd.DataFrame,
    cols: Dict[str, str],
    cfg: AppConfig,
    run_id: str = "",
    *,
    correcciones_nolabor: Optional[Dict[Tuple[str, date], List[Tuple[Optional[time], Optional[time], str]]]] = None,
    correcciones_eventos: Optional[Dict[Tuple[str, date], Dict[str, Tuple[Optional[time], str]]]] = None,
    edicion_interactiva: bool = False,
    interactive_anomalias: bool = False,
    usuario_editor: str = "RRHH",
    audit_log: Optional[List[AuditEntry]] = None,
    modo_seguro: bool = False,
    bulk_pending: Optional[Dict[str, Dict[str, object]]] = None,
) -> pd.DataFrame:
    """Construye el DataFrame principal.

    Robustez:
    - Tolera fechas/horas mal formateadas (parse best-effort)
    - No revienta si faltan entradas/salidas
    - Detecta y registra discrepancias registro_raw vs normalizado

    Interactividad (opcional):
    - Si edicion_interactiva=True, puede abrir editor por registro.
    - Si interactive_anomalias=True, solo pregunta cuando hay anomalías.

    Bulk:
    - bulk_pending puede contener un plan de replicación de checadas por empleado
      generado desde el editor (opción 7).
    """

    correcciones_eventos = correcciones_eventos or {}
    correcciones_nolabor = correcciones_nolabor or {}
    bulk_pending = bulk_pending or {}
    audit_log = audit_log if audit_log is not None else []

    col_id = cols.get("_EMP_KEY") or cols["id"]
    col_nombre = cols.get("nombre")
    col_fecha = cols.get("fecha")
    col_semana = cols.get("semana")
    col_pases = cols.get("pases")
    col_registro = cols.get("registro")

    def _detect_anomalias(registro_raw: str, times: List[time], times_norm: List[time], reord: bool) -> bool:
        # Segundos en texto
        has_seconds = bool(re.search(r"\b\d{1,2}:\d{2}:\d{2}\b", registro_raw or ""))
        # Duplicados en texto (comparando HH:MM)
        toks = re.findall(r"\b\d{1,2}:\d{2}(?::\d{2})?\b", registro_raw or "")
        toks_hm = [t.split(":")[0].zfill(2) + ":" + t.split(":")[1] for t in toks if ":" in t]
        has_dups = len(toks_hm) != len(set(toks_hm)) if toks_hm else False
        # Vacío o >6
        empty = len(times) == 0
        gt6 = len(times) > 6
        # Normalización cambió algo
        changed = reord or (len(times_norm) != len(times))
        return bool(has_seconds or has_dups or empty or gt6 or changed)

    out_rows: List[Dict[str, object]] = []

    # Para rendimiento (archivos grandes): itertuples es mucho más rápido que iterrows
    for r in df.itertuples(index=False, name=None):
        # Acceso por posición: mapeo columnas -> índice
        # Construimos un dict liviano solo para las columnas necesarias
        # (evita overhead de Series por fila)
        row = dict(zip(df.columns, r))

        emp_id = str(row.get(col_id, "") or "").strip()
        nombre = str(row.get(col_nombre, "") or "").strip() if col_nombre else ""
        fecha_raw = row.get(col_fecha, "") if col_fecha else ""
        fecha_d = parse_date(fecha_raw, cfg) if fecha_raw not in (None, "") else None
        semana = str(row.get(col_semana, "") or "").strip() if col_semana else ""
        pases = str(row.get(col_pases, "") or "").strip() if col_pases else ""
        registro = row.get(col_registro, "") if col_registro else ""
        registro_raw = "" if registro is None else str(registro)

        times = parse_registro(registro)

        # Aplicación de plan bulk (si existe) antes de cualquier cálculo
        if emp_id and isinstance(fecha_d, date):
            plan = bulk_pending.get(emp_id)
            if isinstance(plan, dict):
                dates = set(plan.get("dates", []) or [])
                if fecha_d.isoformat() in dates:
                    before = [fmt_hhmm(t) for t in times]
                    plan_times = plan.get("times", []) or []
                    # plan_times viene como ['HH:MM', ...]
                    new_times: List[time] = []
                    for tstr in plan_times:
                        t = parse_time(str(tstr))
                        if t:
                            new_times.append(t)
                    times = new_times
                    # auditoría
                    audit_log.append(
                        AuditEntry(
                            run_id=str(run_id),
                            emp_id=str(emp_id),
                            fecha=fecha_d.isoformat(),
                            usuario=str(plan.get("usuario") or usuario_editor or "RRHH"),
                            ts=datetime.now().isoformat(timespec="seconds"),
                            accion="BULK_APPLY",
                            campo="times",
                            antes=before,
                            despues=[fmt_hhmm(t) for t in times],
                            motivo=str(plan.get("motivo") or "aplicación por lote"),
                        )
                    )

        # Normalización (cruce medianoche) – puede desactivarse con modo_seguro
        if modo_seguro:
            times_norm, reord = times[:], False
        else:
            times_norm, reord = normalize_registro_times(times)

        # Interactividad por registro
        ajuste_manual = ""
        nota_ajuste = ""
        fuente_checadas = "CHECADOR"
        if edicion_interactiva and isinstance(fecha_d, date) and emp_id and not str(emp_id).startswith("NOMBRE::"):
            do_prompt = True
            if interactive_anomalias:
                do_prompt = _detect_anomalias(registro_raw, times, times_norm, reord)
            if do_prompt:
                resp = _safe_input(
                    f"¿Deseas revisar/editar las checadas de este registro (ID: {emp_id}, Fecha: {fecha_d.isoformat()})? (S/N): ",
                    "N",
                ).strip().lower()
                if resp in ("s", "si", "sí", "y", "yes"):
                    no_labor = correcciones_nolabor.get((emp_id, fecha_d), None) if correcciones_nolabor else None
                    try:
                        times_edit, nota_final, bulk_plan, no_labor_edit = editar_checadas_interactivo(
                            emp_id=emp_id,
                            run_id=run_id,
                            nombre=nombre,
                            fecha_d=fecha_d,
                            registro_raw=str(registro_raw or ""),
                            cfg=cfg,
                            usuario=usuario_editor,
                            audit_log=audit_log,
                script_dir=script_dir,
                processed_ids=processed_ids_for_dashboard,
                            modo_seguro=modo_seguro,
                            no_labor=no_labor,
                        )
                        if bulk_plan and isinstance(bulk_plan, dict):
                            bulk_pending[emp_id] = bulk_plan
                        if times_edit is not None:
                            times = times_edit
                            # Persistir NoLaborado solo si el usuario guardó (times_edit no None)
                            if isinstance(fecha_d, date):
                                key = (emp_id, fecha_d)
                                if no_labor_edit:
                                    correcciones_nolabor[key] = copy.deepcopy(no_labor_edit)
                                else:
                                    correcciones_nolabor.pop(key, None)
                            fuente_checadas = "EDIT"
                            if nota_final:
                                ajuste_manual = "Sí"
                                nota_ajuste = nota_final
                            # recalcular normalización luego de edición
                            if modo_seguro:
                                times_norm, reord = times[:], False
                            else:
                                times_norm, reord = normalize_registro_times(times)
                    except Exception:
                        log_exception("Fallo en editor interactivo", extra={"id": emp_id, "fecha": str(fecha_d)}, level=logging.WARNING)

        eventos = map_eventos(times_norm)

        # Aplicar correcciones de eventos (Ajustes) ANTES del cálculo
        if isinstance(fecha_d, date):
            corr_evs = correcciones_eventos.get((emp_id, fecha_d), None)
        else:
            corr_evs = None
        if corr_evs:
            ajuste_manual = "Sí"
            parts = []
            for ev, (t_corr, nota_corr) in corr_evs.items():
                if ev in eventos:
                    eventos[ev] = t_corr
                if nota_corr:
                    parts.append(f"{ev}: {nota_corr}")
                else:
                    if t_corr is None:
                        parts.append(f"{ev}: borrado")
                    else:
                        try:
                            parts.append(f"{ev}: {t_corr.strftime('%H:%M')}")
                        except Exception:
                            parts.append(f"{ev}: ajustado")
            nota_ajuste = " | ".join(parts)

        # Sanidad: evita descuentos absurdos si 'Salida a cenar' quedó antes de terminar la comida.
        # Puede ocurrir por edición accidental en dashboard o datos sucios.
        # Regla conservadora: si la cena inicia antes del fin de comida, se ignora la cena y se deja nota.
        if eventos.get("Entrada") and eventos.get("Salida") and eventos.get("Salida a cenar"):
            ent_t = eventos.get("Entrada")
            sal_t = eventos.get("Salida")
            sal_cen = eventos.get("Salida a cenar")
            meal_end = eventos.get("Regreso de comer") or (sal_t if eventos.get("Salida a comer") else None)
            if ent_t and sal_cen and meal_end:
                try:
                    if minutos_entre(ent_t, sal_cen) < minutos_entre(ent_t, meal_end):
                        eventos["_cena_ignorada_anomalia"] = 1
                        eventos["Salida a cenar"] = None
                        eventos["Regreso de cenar"] = None
                except Exception:
                    pass
        no_lab = correcciones_nolabor.get((emp_id, fecha_d), []) if isinstance(fecha_d, date) else []
        trabajado_min, extra_min, comida_ded, cena_ded, no_lab_ded, nolab_ov, nolab_intov, nolab_ign = calcular_trabajado(eventos, cfg, no_lab)

        # Auditoría / Notas
        notas_partes = []
        extra_regs = int(eventos.get("_extra_registros", 0) or 0)
        if extra_regs > 0:
            notas_partes.append(f"Registros extra ignorados: {extra_regs}")
        if reord:
            notas_partes.append("Registro fuera de orden; reordenado")
        if no_lab_ded and no_lab_ded > 0:
            notas_partes.append(f"Descuento NoLaborado: {minutes_to_hhmm(no_lab_ded)}")
            if nolab_ov and nolab_ov > 0:
                notas_partes.append(getattr(cfg, "nota_nolab_solape_cd", "NoLaborado solapado con comida/cena: fusionado"))
        if nolab_intov and nolab_intov > 0:
            notas_partes.append(f"Solape interno NoLaborado: {minutes_to_hhmm(nolab_intov)} (fusionado)")
        if nolab_ign and nolab_ign > 0:
            notas_partes.append(f"{getattr(cfg, 'nota_nolab_fuera_jornada', 'NoLaborado fuera de jornada ignorado')}: {minutes_to_hhmm(nolab_ign)}")
        if eventos.get("Salida a comer") and not eventos.get("Regreso de comer") and eventos.get("Salida"):
            notas_partes.append("Comida incompleta: fin asumido=Salida")
        if eventos.get("Salida a cenar") and not eventos.get("Regreso de cenar") and eventos.get("Salida"):
            notas_partes.append("Cena incompleta: fin asumido=Salida")
        if eventos.get("_cena_ignorada_anomalia"):
            notas_partes.append("Cena anómala: inicia antes de fin de comida (ignorada)")

        # Discrepancias raw vs normalizadas
        disc = []
        raw_list = [fmt_hhmm(t) for t in times]
        norm_list = [fmt_hhmm(t) for t in times_norm]
        if raw_list != norm_list:
            disc.append("raw!=normalizado")
        if re.search(r"\b\d{1,2}:\d{2}:\d{2}\b", registro_raw or ""):
            disc.append("segundos_recortados")
        if extra_regs > 0:
            disc.append(">6_checadas")
        discrepancias = ",".join(disc)

        notas_str = " | ".join(notas_partes)

        out_rows.append(
            {
                "ID": emp_id,
                "Fecha": fecha_d if isinstance(fecha_d, date) else fecha_raw,
                "Semana": semana,
                "Nombre": nombre,
                "Pases": pases,
                "Registro original": registro_raw,
                "Registros parseados": ", ".join(raw_list),
                "Registros normalizados": ", ".join(norm_list),
                "NoLaborado": _fmt_no_labor_list(no_lab),
                "Fuente checadas": fuente_checadas,
                "Discrepancias": discrepancias,
                "Entrada": fmt_hhmm(eventos["Entrada"]),
                "Salida a comer": fmt_hhmm(eventos["Salida a comer"]),
                "Regreso de comer": fmt_hhmm(eventos["Regreso de comer"]),
                "Salida a cenar": fmt_hhmm(eventos["Salida a cenar"]),
                "Regreso de cenar": fmt_hhmm(eventos["Regreso de cenar"]),
                "Salida": fmt_hhmm(eventos["Salida"]),
                "Horas trabajadas": minutes_to_hhmm(trabajado_min),
                "Horas extra": minutes_to_hhmm(extra_min),
                "Notas": notas_str,
                "Ajuste manual": ajuste_manual,
                "Nota ajuste": nota_ajuste,
            }
        )

    out = pd.DataFrame(out_rows)
    # Compactar Semana si viene vacía
    if "Semana" in out.columns and out["Semana"].astype(str).str.strip().eq("").all():
        out = out.drop(columns=["Semana"])
    return out


def _parse_times_list(s: str) -> List[time]:
    """Convierte "09:10, 09:29" (o con espacios) -> [time,...] tolerante."""
    if s is None:
        return []
    txt = str(s).replace(",", " ")
    toks = re.findall(r"\b\d{1,2}:\d{2}\b", txt)
    out: List[time] = []
    for t in toks:
        tt = parse_time(t)
        if tt:
            out.append(tt)
    return out


def _fmt_no_labor_list(no_lab: List[Tuple[Optional[time], Optional[time], str]]) -> str:
    """Convierte NoLaborado [(ini,fin,nota)] -> 'HH:MM-HH:MM (nota); ...'"""
    if not no_lab:
        return ""
    parts: List[str] = []
    for a, b, nota in no_lab:
        sa = fmt_hhmm(a)
        sb = fmt_hhmm(b)
        n = str(nota or "").strip()
        if n:
            parts.append(f"{sa}-{sb} ({n})")
        else:
            parts.append(f"{sa}-{sb}")
    return "; ".join(parts)


def _recalcular_out_row(
    *,
    emp_id: str,
    nombre: str,
    fecha_val: object,
    semana: str,
    pases: str,
    registro_original: str,
    times: List[time],
    cfg: AppConfig,
    run_id: str,
    usuario_editor: str,
    audit_log: List[AuditEntry],
    modo_seguro: bool,
    correcciones_eventos: Dict[Tuple[str, date], Dict[str, Tuple[Optional[time], str]]],
    correcciones_nolabor: Dict[Tuple[str, date], List[Tuple[Optional[time], Optional[time], str]]],
    ajuste_manual: str,
    nota_ajuste: str,
) -> Dict[str, object]:
    """Recalcula el dict de salida de un registro tras una edición por ID."""
    # Normaliza fecha
    fecha_d: Optional[date]
    if isinstance(fecha_val, date):
        fecha_d = fecha_val
    else:
        try:
            fecha_d = parse_date(fecha_val, cfg)
        except Exception:
            fecha_d = None

    # Normalización (cruce medianoche)
    if modo_seguro:
        times_norm, reord = times[:], False
    else:
        times_norm, reord = normalize_registro_times(times)

    eventos = map_eventos(times_norm)

    # Aplicar correcciones de eventos si existen
    if isinstance(fecha_d, date):
        corr_evs = correcciones_eventos.get((emp_id, fecha_d))
    else:
        corr_evs = None
    if corr_evs:
        ajuste_manual = "Sí"
        parts = []
        for ev, (t_corr, nota_corr) in corr_evs.items():
            if ev in eventos:
                eventos[ev] = t_corr
            if nota_corr:
                parts.append(f"{ev}: {nota_corr}")
            else:
                if t_corr is None:
                    parts.append(f"{ev}: borrado")
                else:
                    try:
                        parts.append(f"{ev}: {t_corr.strftime('%H:%M')}")
                    except Exception:
                        parts.append(f"{ev}: ajustado")
        nota_ajuste = " | ".join(parts) if parts else nota_ajuste

    # Sanidad: evita descuentos absurdos si 'Salida a cenar' quedó antes de terminar la comida.
    # Puede ocurrir por edición accidental en dashboard o datos sucios.
    # Regla conservadora: si la cena inicia antes del fin de comida, se ignora la cena y se deja nota.
    if eventos.get("Entrada") and eventos.get("Salida") and eventos.get("Salida a cenar"):
        ent_t = eventos.get("Entrada")
        sal_t = eventos.get("Salida")
        sal_cen = eventos.get("Salida a cenar")
        meal_end = eventos.get("Regreso de comer") or (sal_t if eventos.get("Salida a comer") else None)
        if ent_t and sal_cen and meal_end:
            try:
                if minutos_entre(ent_t, sal_cen) < minutos_entre(ent_t, meal_end):
                    eventos["_cena_ignorada_anomalia"] = 1
                    eventos["Salida a cenar"] = None
                    eventos["Regreso de cenar"] = None
            except Exception:
                pass
    no_lab = correcciones_nolabor.get((emp_id, fecha_d), []) if isinstance(fecha_d, date) else []
    trabajado_min, extra_min, comida_ded, cena_ded, no_lab_ded, nolab_ov, nolab_intov, nolab_ign = calcular_trabajado(eventos, cfg, no_lab)

    # Notas
    notas_partes = []
    extra_regs = int(eventos.get("_extra_registros", 0) or 0)
    if extra_regs > 0:
        notas_partes.append(f"Registros extra ignorados: {extra_regs}")
    if reord:
        notas_partes.append("Registro fuera de orden; reordenado")
    if no_lab_ded and no_lab_ded > 0:
        notas_partes.append(f"Descuento NoLaborado: {minutes_to_hhmm(no_lab_ded)}")
        if nolab_ov and nolab_ov > 0:
            notas_partes.append(getattr(cfg, "nota_nolab_solape_cd", "NoLaborado solapado con comida/cena: fusionado"))
    if nolab_intov and nolab_intov > 0:
        notas_partes.append(f"Solape interno NoLaborado: {minutes_to_hhmm(nolab_intov)} (fusionado)")
    if nolab_ign and nolab_ign > 0:
        notas_partes.append(f"{getattr(cfg, 'nota_nolab_fuera_jornada', 'NoLaborado fuera de jornada ignorado')}: {minutes_to_hhmm(nolab_ign)}")
    if eventos.get("Salida a comer") and not eventos.get("Regreso de comer") and eventos.get("Salida"):
        notas_partes.append("Comida incompleta: fin asumido=Salida")
    if eventos.get("Salida a cenar") and not eventos.get("Regreso de cenar") and eventos.get("Salida"):
        notas_partes.append("Cena incompleta: fin asumido=Salida")
    if eventos.get("_cena_ignorada_anomalia"):
        notas_partes.append("Cena anómala: inicia antes de fin de comida (ignorada)")

    raw_list = [fmt_hhmm(t) for t in times]
    norm_list = [fmt_hhmm(t) for t in times_norm]
    disc = []
    if raw_list != norm_list:
        disc.append("raw!=normalizado")
    if re.search(r"\b\d{1,2}:\d{2}:\d{2}\b", registro_original or ""):
        disc.append("segundos_recortados")
    if extra_regs > 0:
        disc.append(">6_checadas")

    return {
        "ID": emp_id,
        "Fecha": fecha_d if isinstance(fecha_d, date) else fecha_val,
        "Semana": semana,
        "Nombre": nombre,
        "Pases": pases,
        "Registro original": registro_original,
        "Registros parseados": ", ".join(raw_list),
        "Registros normalizados": ", ".join(norm_list),
        "NoLaborado": _fmt_no_labor_list(no_lab),
        "Discrepancias": ",".join(disc),
        "Entrada": fmt_hhmm(eventos["Entrada"]),
        "Salida a comer": fmt_hhmm(eventos["Salida a comer"]),
        "Regreso de comer": fmt_hhmm(eventos["Regreso de comer"]),
        "Salida a cenar": fmt_hhmm(eventos["Salida a cenar"]),
        "Regreso de cenar": fmt_hhmm(eventos["Regreso de cenar"]),
        "Salida": fmt_hhmm(eventos["Salida"]),
        "Horas trabajadas": minutes_to_hhmm(trabajado_min),
        "Horas extra": minutes_to_hhmm(extra_min),
        "Notas": " | ".join(notas_partes),
        "Ajuste manual": ajuste_manual,
        "Nota ajuste": nota_ajuste,
    }


def dashboard_revision_por_id(
    *,
    df_out: pd.DataFrame,
    cfg: AppConfig,
    run_id: str,
    usuario_editor: str,
    audit_log: List[AuditEntry],
    modo_seguro: bool,
    correcciones_eventos: Dict[Tuple[str, date], Dict[str, Tuple[Optional[time], str]]],
    correcciones_nolabor: Dict[Tuple[str, date], List[Tuple[Optional[time], Optional[time], str]]],
    script_dir: Path,
    processed_ids: Optional[List[str]] = None,
) -> pd.DataFrame:
    """Dashboard en consola para buscar por ID y editar checadas aunque no haya anomalías.

    No interrumpe el lote por cada registro: se usa al final (flag --review).
    """
    if df_out is None or len(df_out) == 0:
        print("No hay datos para revisar.")
        return df_out

    # Lista de IDs detectados en el archivo procesado (se usa en dashboard y admin de grupos).
    processed_ids = list(processed_ids or [])

    def _split_disc(s: object) -> List[str]:
        """Parsea el campo 'Discrepancias' a lista de tokens."""
        txt = str(s or "").strip()
        if not txt:
            return []
        # Admite separadores comunes: coma, punto y coma, pipe
        txt = txt.replace(";", ",").replace("|", ",")
        out = []
        for part in txt.split(","):
            p = part.strip()
            if p:
                out.append(p)
        return out

    def _incidence_stats() -> Tuple[int, int, Dict[str, int]]:
        """(total_rows, rows_with_inc, counts_by_type)"""
        total = int(len(df_out))
        with_inc = 0
        counts: Dict[str, int] = {}
        for v in df_out.get("Discrepancias", pd.Series([], dtype=str)).astype(str).tolist():
            parts = _split_disc(v)
            if not parts:
                continue
            with_inc += 1
            for t in parts:
                counts[t] = counts.get(t, 0) + 1
        return total, with_inc, counts

    def _print_incidence_summary() -> None:
        total, with_inc, counts = _incidence_stats()
        print("\n=== RESUMEN DE INCIDENCIAS ===")
        print(f"Total registros procesados: {total}")
        print(f"Registros con incidencias: {with_inc}")
        if with_inc == 0:
            print("No se detectaron incidencias.")
            return
        # Top 10 por tipo
        items = sorted(counts.items(), key=lambda x: (-x[1], x[0]))[:10]
        print("Top tipos:")
        for i, (k, n) in enumerate(items, start=1):
            print(f"  {i}) {k}: {n}")

    # Índice rápido: ID -> lista de índices
    id_to_idx: Dict[str, List[int]] = {}
    for i, v in enumerate(df_out["ID"].astype(str).tolist()):
        k = str(v).strip()
        if not k:
            continue
        id_to_idx.setdefault(k, []).append(i)

    def _row_date_iso(irow: int) -> Tuple[str, Optional[date]]:
        """Devuelve (iso, date|None) para la fila de salida."""
        try:
            fv = df_out.at[irow, "Fecha"]
        except Exception:
            fv = ""
        if isinstance(fv, date):
            return fv.isoformat(), fv
        try:
            d = parse_date(fv, cfg)
            if isinstance(d, date):
                return d.isoformat(), d
        except Exception:
            pass
        s = str(fv or "").strip()
        return s, None

    def _build_incidence_list(tipo: str = "", limit: int = 30) -> List[Tuple[int, str, str, str, str]]:
        """Devuelve lista: (irow, iso_fecha, emp_id, nombre, disc)."""
        out: List[Tuple[int, str, str, str, str]] = []
        if "Discrepancias" not in df_out.columns:
            return out
        for irow in range(len(df_out)):
            disc = str(df_out.at[irow, "Discrepancias"] or "").strip()
            if not disc:
                continue
            parts = _split_disc(disc)
            if tipo and (tipo not in parts):
                continue
            iso, _ = _row_date_iso(irow)
            emp_id = str(df_out.at[irow, "ID"] or "").strip()
            nombre = str(df_out.at[irow, "Nombre"] or "").strip() if "Nombre" in df_out.columns else ""
            out.append((irow, iso, emp_id, nombre, disc))
            if len(out) >= limit:
                break
        return out

    def _sorted_rows_for_id(idxs: List[int]) -> List[Tuple[str, int]]:
        rows: List[Tuple[str, int]] = []
        for irow in idxs:
            iso, _ = _row_date_iso(irow)
            rows.append((iso, irow))
        rows.sort(key=lambda x: x[0])
        return rows

    def _parse_selection(sel: str, rows: List[Tuple[str, int]]) -> List[Tuple[str, int]]:
        """Soporta:
        - número (1..N)
        - fecha ISO YYYY-MM-DD
        - rango YYYY-MM-DD..YYYY-MM-DD (cascada)
        """
        sel = (sel or "").strip()
        if not sel:
            return []
        if sel.isdigit():
            k = int(sel)
            if 1 <= k <= len(rows):
                return [rows[k - 1]]
            return []

        if ".." in sel:
            a, b = [x.strip() for x in sel.split("..", 1)]
            try:
                da = date.fromisoformat(a)
                db = date.fromisoformat(b)
            except Exception:
                return []
            if db < da:
                da, db = db, da
            picked: List[Tuple[str, int]] = []
            for iso, irow in rows:
                try:
                    d = date.fromisoformat(iso)
                except Exception:
                    continue
                if da <= d <= db:
                    picked.append((iso, irow))
            return picked

        # fecha ISO exacta
        try:
            date.fromisoformat(sel)
        except Exception:
            return []
        for iso, irow in rows:
            if iso == sel:
                return [(iso, irow)]
        return []

    def _print_rows_for_id(emp_id: str, idxs: List[int]) -> None:
        rows = _sorted_rows_for_id(idxs)
        print(f"\nID {emp_id} | {len(rows)} día(s)")
        for j, (iso, i) in enumerate(rows, start=1):
            r = df_out.iloc[i]
            fecha = iso or r.get("Fecha", "")
            disc = str(r.get("Discrepancias", "") or "").strip()
            ht = str(r.get("Horas trabajadas", "") or "")
            he = str(r.get("Horas extra", "") or "")
            print(f" {j:>2}) {fecha} | Trab={ht} Extra={he} | {('DISC:'+disc) if disc else ''}")

    def _editar_cascada(emp_id: str, picked: List[Tuple[str, int]]) -> None:
        """Edita en cascada una lista de (iso, irow) para un mismo empleado."""
        for pos, (iso, irow) in enumerate(picked, start=1):
            r = df_out.iloc[irow].to_dict()
            nombre = str(r.get("Nombre", "") or "")
            fecha_val = r.get("Fecha", "")
            semana = str(r.get("Semana", "") or "")
            pases = str(r.get("Pases", "") or "")
            registro_original = str(r.get("Registro original", "") or "")

            # Para el editor, usa el estado actual (normalizados) como fuente parseable
            registro_actual = str(r.get("Registros normalizados", "") or "")
            registro_parseable = registro_actual.replace(",", " ").strip() or registro_original
            registro_display = f"ORIGINAL: {registro_original}\nACTUAL: {registro_actual}".strip()

            # Fecha como date
            if isinstance(fecha_val, date):
                fecha_d = fecha_val
            else:
                fecha_d = parse_date(fecha_val, cfg)

            if not isinstance(fecha_d, date):
                print(f"No se pudo interpretar la fecha para la fila {iso}. Se omite.")
                continue

            print(f"\n--- CASCADA {pos}/{len(picked)} | ID={emp_id} | Fecha={fecha_d.isoformat()} ---")
            no_labor = correcciones_nolabor.get((emp_id, fecha_d), [])

            times_edit, nota_final, _bulk_plan, no_labor_edit = editar_checadas_interactivo(
                run_id=run_id,
                emp_id=emp_id,
                nombre=nombre,
                fecha_d=fecha_d,
                registro_raw=registro_parseable,
                registro_display=registro_display,
                cfg=cfg,
                usuario=usuario_editor,
                audit_log=audit_log,
                script_dir=script_dir,
                processed_ids=processed_ids,
                modo_seguro=modo_seguro,
                no_labor=no_labor,
            )

            if times_edit is None:
                # 0 o salida sin guardar: permite saltar el día
                cont = _safe_input("(ENTER) siguiente día | (Q) salir cascada: ", "").strip().lower()
                if cont == "q":
                    break
                continue

            # Persistir NoLaborado solo si el usuario guardó (times_edit no None)
            key = (emp_id, fecha_d)
            if no_labor_edit:
                correcciones_nolabor[key] = copy.deepcopy(no_labor_edit)
            else:
                correcciones_nolabor.pop(key, None)

            # Recalcular fila con las nuevas checadas
            nota_adj = nota_final or "Edición por dashboard (cascada)"
            new_row = _recalcular_out_row(
                emp_id=emp_id,
                nombre=nombre,
                fecha_val=fecha_val,
                semana=semana,
                pases=pases,
                registro_original=registro_original,
                times=times_edit,
                cfg=cfg,
                run_id=run_id,
                usuario_editor=usuario_editor,
                audit_log=audit_log,
                modo_seguro=modo_seguro,
                correcciones_eventos=correcciones_eventos,
                correcciones_nolabor=correcciones_nolabor,
                ajuste_manual="Sí",
                nota_ajuste=nota_adj,
            )

            # Actualizar df_out
            for c, v in new_row.items():
                if c in df_out.columns:
                    df_out.at[irow, c] = v
                else:
                    df_out[c] = ""
                    df_out.at[irow, c] = v

            print("Recalculado con cambios.")

            cont = _safe_input("(ENTER) siguiente día | (Q) salir cascada: ", "").strip().lower()
            if cont == "q":
                break

    # Mostrar resumen al entrar (antes de editar)
    _print_incidence_summary()

    def _editar_fila(irow: int) -> None:
        """Abre el editor para una fila específica y, si se guarda, actualiza df_out."""
        r = df_out.iloc[irow].to_dict()
        emp_id = normalize_id(str(r.get("ID", "") or ""), cfg.id_min_width)
        if not emp_id:
            print("Fila sin ID válido.")
            return
        nombre = str(r.get("Nombre", "") or "")
        fecha_val = r.get("Fecha", "")
        semana = str(r.get("Semana", "") or "")
        pases = str(r.get("Pases", "") or "")
        registro_original = str(r.get("Registro original", "") or "")
        registro_actual = str(r.get("Registros normalizados", "") or "")
        registro_parseable = registro_actual.replace(",", " ").strip() or registro_original
        registro_display = f"ORIGINAL: {registro_original}\nACTUAL: {registro_actual}".strip()
        if isinstance(fecha_val, date):
            fecha_d = fecha_val
        else:
            fecha_d = parse_date(fecha_val, cfg)
        if not isinstance(fecha_d, date):
            print("No se pudo interpretar la fecha para esta fila.")
            return
        print(f"\n--- EDITOR | ID={emp_id} | Fecha={fecha_d.isoformat()} ---")
        no_labor = correcciones_nolabor.get((emp_id, fecha_d), [])
        times_edit, nota_final, _bulk_plan, no_labor_edit = editar_checadas_interactivo(
            run_id=run_id,
            emp_id=emp_id,
            nombre=nombre,
            fecha_d=fecha_d,
            registro_raw=registro_parseable,
            registro_display=registro_display,
            cfg=cfg,
            usuario=usuario_editor,
            audit_log=audit_log,
            script_dir=script_dir,
            processed_ids=processed_ids,
            modo_seguro=modo_seguro,
            no_labor=no_labor,
        )
        if times_edit is None:
            return
        key = (emp_id, fecha_d)
        if no_labor_edit:
            correcciones_nolabor[key] = copy.deepcopy(no_labor_edit)
        else:
            correcciones_nolabor.pop(key, None)
        nota_adj = nota_final or "Edición por incidencias"
        new_row = _recalcular_out_row(
            emp_id=emp_id,
            nombre=nombre,
            fecha_val=fecha_val,
            semana=semana,
            pases=pases,
            registro_original=registro_original,
            times=times_edit,
            cfg=cfg,
            run_id=run_id,
            usuario_editor=usuario_editor,
            audit_log=audit_log,
            modo_seguro=modo_seguro,
            correcciones_eventos=correcciones_eventos,
            correcciones_nolabor=correcciones_nolabor,
            ajuste_manual="Sí",
            nota_ajuste=nota_adj,
        )
        for c, v in new_row.items():
            if c in df_out.columns:
                df_out.at[irow, c] = v
            else:
                df_out[c] = ""
                df_out.at[irow, c] = v
        print("Recalculado con cambios.")

    while True:
        print("\n=== DASHBOARD REVISION POR ID ===")
        print("1) Buscar por ID")
        print("2) Ver resumen de incidencias")
        print("3) Listar incidencias (top 30) y abrir editor")
        print("4) Filtrar incidencias por tipo")
        print("5) Salir y continuar exportación")
        op = _safe_input("> ", "5").strip()
        if op == "5":
            return df_out

        if op == "2":
            _print_incidence_summary()
            continue

        if op == "3":
            inc_list = _build_incidence_list("", limit=30)
            if not inc_list:
                print("No hay incidencias.")
                continue
            print("\n--- INCIDENCIAS (top 30) ---")
            for j, (_irow, iso, emp, nom, disc) in enumerate(inc_list, start=1):
                who = f"{emp} {nom}".strip()
                print(f" {j:>2}) {iso} | {who} | {disc}")
            sel = _safe_input("Abrir # (1..N) o ENTER para volver: ", "").strip()
            if not sel:
                continue
            if sel.isdigit():
                k = int(sel)
                if 1 <= k <= len(inc_list):
                    _editar_fila(inc_list[k - 1][0])
                else:
                    print("Selección inválida.")
            else:
                print("Selección inválida.")
            continue

        if op == "4":
            tipo = _safe_input("Tipo (ej. COMIDA_INCOMPLETA): ", "").strip()
            if not tipo:
                continue
            inc_list = _build_incidence_list(tipo, limit=30)
            if not inc_list:
                print("No hay incidencias para ese tipo.")
                continue
            print(f"\n--- INCIDENCIAS [{tipo}] (top 30) ---")
            for j, (_irow, iso, emp, nom, disc) in enumerate(inc_list, start=1):
                who = f"{emp} {nom}".strip()
                print(f" {j:>2}) {iso} | {who} | {disc}")
            sel = _safe_input("Abrir # (1..N) o ENTER para volver: ", "").strip()
            if not sel:
                continue
            if sel.isdigit():
                k = int(sel)
                if 1 <= k <= len(inc_list):
                    _editar_fila(inc_list[k - 1][0])
                else:
                    print("Selección inválida.")
            else:
                print("Selección inválida.")
            continue

        if op != "1":
            print("Opción inválida.")
            continue

        emp_id = _safe_input("ID empleado: ").strip()
        if not emp_id:
            continue
        emp_id = normalize_id(emp_id, cfg.id_min_width)
        idxs = id_to_idx.get(emp_id, [])
        if not idxs:
            print("No encontré ese ID en el archivo procesado.")
            continue

        _print_rows_for_id(emp_id, idxs)
        print("\nSelecciona:")
        print("- Un día:  NÚMERO (1..N) o YYYY-MM-DD")
        print("- Rango cascada:  YYYY-MM-DD..YYYY-MM-DD")
        sel = _safe_input("Selección o ENTER para volver: ").strip()
        if not sel:
            continue

        rows = _sorted_rows_for_id(idxs)
        picked = _parse_selection(sel, rows)
        if not picked:
            print("Selección inválida.")
            continue

        # --- Cascada: procesa los días seleccionados uno por uno ---
        for pos, (_iso, irow) in enumerate(picked, start=1):
            print(f"\n--- CASCADA {pos}/{len(picked)} | ID={emp_id} ---")
            _editar_fila(irow)
            cont = _safe_input("(ENTER) siguiente día | (Q) salir cascada: ", "").strip().lower()
            if cont == "q":
                break

        print("\n(Cascada terminada)\n")


def procesar_archivo(
    in_path: Path,
    *,
    correccion_interactiva: bool = False,
    correcciones_eventos: Optional[Dict[Tuple[str, date], Dict[str, Tuple[Optional[time], str]]]] = None,
    correcciones_nolabor: Optional[Dict[Tuple[str, date], List[Tuple[Optional[time], Optional[time], str]]]] = None,
    plantilla_path: str = "",
    edicion_interactiva: bool = False,
    usuario_editor: str = "RRHH",
    modo_seguro: bool = False,
    verify: bool = False,
    interactive_anomalias: bool = False,
    interactive_grupos: bool = False,
    review_por_id: bool = False,
    dry_run: bool = False,
    no_interactive: bool = False,
    script_dir_override: Optional[Path] = None,
) -> Tuple[Path, Path]:
    script_dir = Path(script_dir_override) if script_dir_override else Path(__file__).resolve().parent
    cfg = cargar_config(script_dir)
    # Flags de ejecución
    no_interactive = bool(no_interactive) or bool(getattr(cfg, 'no_interactive_default', False))
    dry_run = bool(dry_run)

    # Compat: correccion_interactiva se usa como alias de interactive_grupos
    interactive_grupos = bool(interactive_grupos) or bool(correccion_interactiva)

    # --- Contexto de corrida (run_id + hash de entrada) ---
    run_id = str(uuid.uuid4())
    try:
        input_sha256 = sha256_file(in_path)
    except Exception:
        input_sha256 = ""
    try:
        import procesador as _pkg  # type: ignore
        version = getattr(_pkg, "__version__", "")
    except Exception:
        version = ""
    started_at = datetime.now().isoformat(timespec="seconds")

    correcciones_eventos = correcciones_eventos or {}
    correcciones_nolabor = correcciones_nolabor or {}
    df_in, cols = leer_input(in_path, cfg)

    # IDs detectados en el archivo procesado (para dashboard y para plantilla de empleados)
    processed_ids_for_dashboard: List[str] = []
    try:
        col_id = cols.get('id')
        if col_id and col_id in df_in.columns:
            processed_ids_for_dashboard = (
                df_in[col_id].fillna('').astype(str).str.strip().tolist()
            )
    except Exception:
        processed_ids_for_dashboard = []

    plantilla = cargar_plantilla_empleados(
        script_dir,
        plantilla_path,
        cfg=cfg,
        empleados_detectados=processed_ids_for_dashboard,
    )

    # Completar mapeos faltantes (grupos/IDGRUPO) interactivo (solo si se solicita)
    if interactive_grupos and (not no_interactive) and (not dry_run):
        try:
            _tmp, _tmp_idg = aplicar_grupos_y_idgrupo(df_in.copy(), cols, cfg, permitir_interactivo=True)
            guardar_config(script_dir, cfg)
        except Exception:
            log_exception("Fallo en mapeo interactivo de grupos", level=logging.WARNING)
    # Construir salida “base”
    audit_log: List[AuditEntry] = []
    bulk_pending: Dict[str, Dict[str, object]] = {}
    df_out = construir_salida(
        df_in,
        cols,
        cfg,
        run_id=run_id,
        correcciones_nolabor=correcciones_nolabor,
        correcciones_eventos=correcciones_eventos,
        edicion_interactiva=bool(edicion_interactiva) and (not no_interactive) and (not dry_run),
        interactive_anomalias=bool(interactive_anomalias) and (not no_interactive) and (not dry_run),
        usuario_editor=usuario_editor,
        audit_log=audit_log,
        modo_seguro=modo_seguro,
        bulk_pending=bulk_pending,
    )

    # Dashboard de revisión por ID (post-scan). Permite editar cualquier empleado
    # aunque no tenga anomalías, sin bombardear con preguntas durante el loop.
    if review_por_id and (not no_interactive) and (not dry_run):
        try:
            dashboard_revision_por_id(
                df_out=df_out,
                cfg=cfg,
                run_id=run_id,
                usuario_editor=str(usuario_editor or "RRHH"),
                audit_log=audit_log,
                modo_seguro=bool(modo_seguro),
                correcciones_eventos=correcciones_eventos,
                correcciones_nolabor=correcciones_nolabor,
                script_dir=script_dir,
                processed_ids=processed_ids_for_dashboard,
            )
        except Exception:
            log_exception("Fallo en dashboard de revisión por ID", level=logging.WARNING)
    # Correcciones manuales (si existe correcciones_asistencia.xlsx / --correcciones)
    if False and correcciones_eventos:
        ajustado_flags: List[str] = []
        notas_flags: List[str] = []
        for idx, r in df_out.iterrows():
            row = r.to_dict()
            row2, aplicado, nota = aplicar_correcciones(row, correcciones_eventos, id_col="ID", fecha_col="Fecha")
            ajustado_flags.append("Sí" if aplicado else "")
            notas_flags.append(nota if aplicado else "")
            # Actualiza columnas de eventos (si existen en el DF)
            for ev in CORR_EVENTOS:
                if ev in df_out.columns:
                    df_out.at[idx, ev] = row2.get(ev, "")
        # Para auditoría. Si no quieres estas columnas, se pueden desactivar.
        df_out["Ajuste manual"] = ajustado_flags
        df_out["Nota ajuste"] = notas_flags
    # Aplicar orden por grupos usando el ID original (columna "ID" en df_out)
    # Creamos DF temporal para ordenar con cfg.empleado_a_grupo.
    df_out["_grp_idx"] = df_out["ID"].map(lambda x: _grupo_sort_key(str(x), cfg)[0])
    df_out = df_out.sort_values(by=["_grp_idx", "ID"], kind="stable").drop(columns=["_grp_idx"]).reset_index(drop=True)
    # IDGRUPO: agregar columna al inicio según mapping manual
    
    df_idgrupo = df_out.copy()
    # IDGRUPO: por prioridad se toma de la plantilla (si existe y está capturado); si no, usa el mapeo del JSON
    plantilla_idgrupo: Dict[str, str] = {}
    try:
        if plantilla is not None and len(plantilla) > 0 and "IDGRUPO" in plantilla.columns:
            plantilla_idgrupo = dict(
                zip(
                    plantilla["ID"].astype(str).str.strip(),
                    plantilla["IDGRUPO"].astype(str).fillna("").astype(str).str.strip(),
                )
            )
    except Exception:
        plantilla_idgrupo = {}

    def _idgrupo_of(emp_id: object) -> str:
        k = str(emp_id).strip()
        if k.startswith("NOMBRE::"):
            return ""
        # 1) plantilla explícita (si existe)
        v = plantilla_idgrupo.get(k, "")
        if v and str(v).lower() != "nan":
            return str(v)
        # 2) mapping explícito del JSON
        try:
            v2 = (cfg.empleado_a_idgrupo or {}).get(k)
            if v2:
                return str(v2)
        except Exception:
            pass
        # 3) derivado por grupo
        try:
            grp = (cfg.empleado_a_grupo or {}).get(k, "")
            if grp:
                return str(cfg.prefijo_de_grupo(grp))
        except Exception:
            pass
        return ""
    # Asegurar columna IDGRUPO al inicio (sin duplicarla)
    if "IDGRUPO" in df_idgrupo.columns:
        df_idgrupo = df_idgrupo.drop(columns=["IDGRUPO"])
    df_idgrupo.insert(0, "IDGRUPO", df_idgrupo["ID"].map(lambda x: build_idgrupo_label(_idgrupo_of(x), x, cfg)))
    # En _IDGRUPO NO se conserva la columna ID original (control por archivo)
    df_idgrupo = df_idgrupo.drop(columns=["ID"], errors="ignore")

    # Guardar config sin tocar si no hubo cambios
    guardar_config(script_dir, cfg)
    base = in_path.with_suffix("")
    # En _PROCESADO NO se incluye la columna IDGRUPO (control por archivo)
    df_out = df_out.drop(columns=["IDGRUPO"], errors="ignore")
    out1 = Path(str(base) + "_PROCESADO.xlsx")
    out2 = Path(str(base) + "_IDGRUPO.xlsx")
    # ---------------------------
    # Resúmenes + Faltas (Opción A: plantilla de empleados activos)
    # ---------------------------
    # plantilla ya cargada arriba
    # plantilla = cargar_plantilla_empleados(script_dir, plantilla_path, cfg=cfg, empleados_detectados=processed_ids_for_dashboard)
    if plantilla is None:
        faltas_sem, faltas_mes, detalle_faltas = (pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
    else:
        # Marcar IDs que aparecen en asistencia pero NO existen en la plantilla (útil para auditoría).
        try:
            ids_plant = set(plantilla["ID"].astype(str).str.strip().tolist())
            if "ID" in df_out.columns and "Notas" in df_out.columns:
                for _i, _r in df_out.iterrows():
                    _id = str(_r.get("ID","")).strip()
                    if _id and _id not in ids_plant:
                        oldn = str(df_out.at[_i, "Notas"]).strip()
                        tag = "ID no está en plantilla (no cuenta para faltas)"
                        if oldn and oldn.lower() != "nan":
                            if tag not in oldn:
                                df_out.at[_i, "Notas"] = oldn + " | " + tag
                        else:
                            df_out.at[_i, "Notas"] = tag
        except Exception:
            pass
        faltas_sem, faltas_mes, detalle_faltas = calcular_faltas(df_out, plantilla, cfg)
    resumen_semanal = construir_resumen_semanal(df_out, cfg, faltas_semanal=faltas_sem)
    resumen_semanal_vertical = construir_resumen_semanal_vertical(df_out, cfg, faltas_semanal=faltas_sem)
    resumen_mensual = construir_resumen_mensual(df_out, cfg, faltas_mensual=faltas_mes)

    # ---------------------------
        # Incidencias RRHH (revisión)
    # ---------------------------
    # Hoja profesional para RRHH: separa Motivo (código), Severidad, Acción sugerida y Evidencia,
    # además del Detalle (Notas) y el Registro crudo del export (solo aquí).
    #
    # Filosofía:
    # - No altera cálculos.
    # - No elimina opciones.
    # - Útil para filtrar y trabajar por lotes (RRHH/Admin).
    incidencias_rrhh = pd.DataFrame()
    try:
        if "Notas" in df_out.columns:
            reg_col = cols.get("registro") if isinstance(cols, dict) else None

            # Mapeo de patrones -> (codigo, severidad, accion)
            reglas = [
                (r"Registros extra ignorados:\s*(\d+)", "INC-CHK-006", "ALTA",
                 "Confirmar cuáles checadas son válidas; capturar Ajuste si aplica. Si el lector duplicó, documentar y dejar evidencia."),
                (r"Comida incompleta", "INC-CHK-002", "MEDIA",
                 "Validar si hubo comida; si aplica, capturar Ajuste (hora faltante) o justificar incidencia."),
                (r"Cena incompleta", "INC-CHK-002", "MEDIA",
                 "Validar si hubo cena; si aplica, capturar Ajuste (hora faltante) o justificar incidencia."),
                (r"NoLaborado fuera de jornada", "INC-PER-002", "MEDIA",
                 "Revisar permiso (fecha/hora). Si fue correcto, ajustar jornada con Ajuste; si fue error de captura, corregir el permiso."),
                (r"Solape interno NoLaborado", "INFO-PER-003", "BAJA",
                 "Sin acción: el sistema fusionó intervalos de permiso para evitar doble descuento (solo informativo)."),
                (r"NoLaborado solapado con comida/cena|NoLaborado solapado con comida/cena", "INFO-PER-004", "BAJA",
                 "Sin acción: el sistema evitó doble descuento con comida/cena (solo informativo)."),
            ]

            # Serie de Notas normalizada
            notas_s = df_out.get("Notas", pd.Series(dtype=str)).fillna("").astype(str)
            notas_s = notas_s.where(~notas_s.str.lower().eq("nan"), "")

            # Registro crudo (solo para evidencias)
            if reg_col and reg_col in df_in.columns:
                try:
                    reg_raw_s = df_in[reg_col].fillna("").astype(str)
                except Exception:
                    reg_raw_s = pd.Series([""] * len(df_out))
            else:
                reg_raw_s = pd.Series([""] * len(df_out))

            # Incidencia admin: sin IDGRUPO (no aparece en reportes por grupo)
            # Se detecta con base en el IDGRUPO calculado para *_IDGRUPO.xlsx
            ids = df_out["ID"].astype(str).fillna("").str.strip()

            plantilla_idgrupo_map = {}
            try:
                if plantilla is not None and "IDGRUPO" in plantilla.columns and "ID" in plantilla.columns:
                    plantilla_idgrupo_map = dict(
                        zip(
                            plantilla["ID"].astype(str).str.strip(),
                            plantilla["IDGRUPO"].astype(str).fillna("").astype(str).str.strip(),
                        )
                    )
            except Exception:
                plantilla_idgrupo_map = {}

            def _idgrupo_fast(emp_id: str) -> str:
                if not emp_id or emp_id.startswith("NOMBRE::"):
                    return ""
                v = plantilla_idgrupo_map.get(emp_id, "")
                if v and str(v).lower() != "nan":
                    return str(v)
                try:
                    v2 = (cfg.empleado_a_idgrupo or {}).get(emp_id, "")
                    return str(v2 or "")
                except Exception:
                    return ""

            idgrupo_series = ids.map(_idgrupo_fast)
            mask_sin_idgrupo = idgrupo_series.fillna("").astype(str).str.strip().eq("") & ~ids.str.startswith("NOMBRE::")

            frames: list[pd.DataFrame] = []

            # Generar incidencias por reglas (vectorizado)
            for pat, cod, sev, accion in reglas:
                try:
                    contains_pat = (r"Registros extra ignorados:\\s*\\d+" if cod == "INC-CHK-006" else pat)
                    mask = notas_s.str.contains(contains_pat, case=False, regex=True, na=False)
                except Exception:
                    # por seguridad, fallback a máscara vacía si el regex falla
                    mask = pd.Series([False] * len(notas_s))
                if not bool(mask.any()):
                    continue
                tmp = df_out.loc[mask, ["ID", "Fecha", "Nombre"]].copy()
                tmp["Registro"] = reg_raw_s.loc[mask].values
                tmp["Detalle"] = notas_s.loc[mask].values
                tmp["Motivo"] = cod
                tmp["Severidad"] = sev
                tmp["Acción sugerida"] = accion
                # Evidencia: prioriza Registro crudo, si no existe, recorta notas
                if cod == "INC-CHK-006":
                    # ExtraIgnorados (si se puede extraer)
                    try:
                        extras_n = notas_s.loc[mask].str.extract(pat, flags=re.IGNORECASE)[0].fillna("")
                        tmp["Evidencia"] = ("ExtraIgnorados=" + extras_n.astype(str) + ". Registro=" + tmp["Registro"].astype(str)).str.strip()
                    except Exception:
                        tmp["Evidencia"] = ("Registro=" + tmp["Registro"].astype(str)).str.strip()
                else:
                    evid = np.where(
                        tmp["Registro"].astype(str).str.strip().ne(""),
                        "Registro=" + tmp["Registro"].astype(str),
                        tmp["Detalle"].astype(str).str.slice(0, 200),
                    )
                    tmp["Evidencia"] = evid
                frames.append(tmp)

            # Incidencia por falta de IDGRUPO (vectorizado)
            if bool(mask_sin_idgrupo.any()):
                tmp = df_out.loc[mask_sin_idgrupo, ["ID", "Fecha", "Nombre"]].copy()
                tmp["Registro"] = reg_raw_s.loc[mask_sin_idgrupo].values
                tmp["Detalle"] = notas_s.loc[mask_sin_idgrupo].values
                tmp["Motivo"] = "INC-GRP-001"
                tmp["Severidad"] = "ALTA"
                tmp["Acción sugerida"] = "Asignar Grupo e IDGRUPO en Modo Admin y reprocesar para que aparezca en reportes por grupo."
                tmp["Evidencia"] = tmp["ID"].astype(str).map(lambda k: f"IDGRUPO vacío. GrupoActual={cfg.empleado_a_grupo.get(str(k).strip(), '') or '—'}")
                frames.append(tmp)

            if frames:
                incidencias_rrhh = pd.concat(frames, ignore_index=True)
                cols_order = ["Motivo", "Severidad", "Acción sugerida", "Evidencia", "ID", "Fecha", "Nombre", "Registro", "Detalle"]
                cols_order = [c for c in cols_order if c in incidencias_rrhh.columns] + [c for c in incidencias_rrhh.columns if c not in cols_order]
                incidencias_rrhh = incidencias_rrhh[cols_order]
    except Exception:
        incidencias_rrhh = pd.DataFrame()

    # --- IDs visibles (caso SIN ID) ---
    # Internamente usamos claves NOMBRE:: para poder calcular bien. Antes de exportar,
    # dejamos la columna ID en blanco para esos casos, pero conservamos el Nombre.
    df_out_export = apply_id_display(df_out, id_col="ID")
    df_idgrupo_export = df_idgrupo.copy()
    # Asegurar que IDGRUPO quede vacío para claves SIN ID (aunque exista mapeo)
    try:
        # no tenemos la clave interna en df_idgrupo (se eliminó 'ID'), así que tomamos la máscara desde df_out
        sin_id_keys = set(df_out[df_out["ID"].astype(str).str.startswith("NOMBRE::")]["ID"].astype(str).unique().tolist())
        if len(sin_id_keys) > 0 and "IDGRUPO" in df_idgrupo_export.columns:
            # sin ID no podemos cruzar, pero si la plantilla asignó por error, se limpia globalmente por seguridad
            pass
    except Exception:
        pass
    resumen_semanal_export = apply_id_display(resumen_semanal, id_col="ID")
    resumen_mensual_export = apply_id_display(resumen_mensual, id_col="ID")
    detalle_faltas_export = apply_id_display(detalle_faltas, id_col="ID")
    incidencias_rrhh_export = apply_id_display(incidencias_rrhh, id_col="ID")

    # Incidencias (discrepancias técnicas)
    try:
        inc_cols = ["ID", "Fecha", "Nombre", "Registro original", "Registros normalizados", "Discrepancias", "Notas"]
        inc_cols = [c for c in inc_cols if c in df_out_export.columns]
        df_incidencias = df_out_export.copy()
        mask = df_incidencias.get("Discrepancias", pd.Series(dtype=str)).astype(str).str.strip().ne("") | df_incidencias.get("Notas", pd.Series(dtype=str)).astype(str).str.strip().ne("")
        df_incidencias = df_incidencias.loc[mask, inc_cols] if inc_cols else df_incidencias.loc[mask]
        df_incidencias = df_incidencias.reset_index(drop=True)
    except Exception:
        df_incidencias = pd.DataFrame()

    extra_sheets = {
        "RESUMEN_SEMANAL": resumen_semanal_export,
        "RESUMEN_MENSUAL": resumen_mensual_export,
        "DETALLE_FALTAS": detalle_faltas_export,
        "INCIDENCIAS_RRHH": incidencias_rrhh_export,
        "INCIDENCIAS": df_incidencias,
    }
    # Hoja adicional: resumen semanal de checadas (control interno)
    try:
        df_chec_sem_proc = crear_resumen_semanal_checadas(df_out_export, cfg, modo="PROCESADO")
        if df_chec_sem_proc is not None and len(df_chec_sem_proc) > 0:
            extra_sheets["RESUM_SEM_CHECADAS"] = df_chec_sem_proc
    except Exception:
        pass
    # Hoja adicional: CONTROL (resumen ejecutivo + listas)
    try:
        control_rows = []
        # Metadata de ejecución (run)
        try:
            control_rows.append({"Sección":"EJECUCION","Campo":"run_id","Valor":run_id})
            control_rows.append({"Sección":"EJECUCION","Campo":"started_at","Valor":started_at})
            control_rows.append({"Sección":"EJECUCION","Campo":"version","Valor":version})
            control_rows.append({"Sección":"EJECUCION","Campo":"input_sha256","Valor":input_sha256})
        except Exception:
            pass
        # Auditoría (conteos)
        try:
            n_aud = len(audit_log) if audit_log is not None else 0
            control_rows.append({"Sección":"AUDITORIA","Campo":"Ediciones registradas","Valor":n_aud})
            if n_aud > 0:
                emps = sorted(set([str(a.emp_id) for a in audit_log]))
                fechas = sorted(set([str(a.fecha) for a in audit_log]))
                control_rows.append({"Sección":"AUDITORIA","Campo":"Empleados afectados","Valor":len(emps)})
                control_rows.append({"Sección":"AUDITORIA","Campo":"Fechas afectadas","Valor":len(fechas)})
        except Exception:
            pass
        control_rows.append({"Sección":"RESUMEN","Campo":"Filas en Reporte","Valor":len(df_out_export)})
        control_rows.append({"Sección":"RESUMEN","Campo":"Incidencias RRHH","Valor":len(incidencias_rrhh_export)})
        # Faltas
        try:
            control_rows.append({"Sección":"RESUMEN","Campo":"Faltas (detalle)","Valor":len(detalle_faltas_export)})
        except Exception:
            pass
        # Permisos/NoLaborado aplicados (conteo por nota)
        try:
            n_nolab = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("NoLaborado", na=False).sum())
            control_rows.append({"Sección":"RRHH","Campo":"Registros con NoLaborado (menciones)","Valor":n_nolab})
        except Exception:
            pass
        # Ajustes manuales aplicados
        try:
            if "Ajuste manual" in df_out_export.columns:
                n_adj = int(df_out_export["Ajuste manual"].astype(str).str.upper().isin(["SI","SÍ","TRUE","1"]).sum())
                control_rows.append({"Sección":"RRHH","Campo":"Ajustes manuales aplicados","Valor":n_adj})
        except Exception:
            pass
        # Incidencias por motivo/severidad (top)
        try:
            if {"Motivo","Severidad"}.issubset(set(incidencias_rrhh_export.columns)):
                piv = incidencias_rrhh_export.pivot_table(index=["Motivo","Severidad"], values="ID", aggfunc="count").reset_index()
                piv = piv.rename(columns={"ID":"Cantidad"}).sort_values("Cantidad", ascending=False)
                for _, r in piv.head(15).iterrows():
                    control_rows.append({"Sección":"INCIDENCIAS","Campo":f"{r['Motivo']}|{r['Severidad']}", "Valor":int(r["Cantidad"])})
        except Exception:
            pass
        # Empleados SIN ID (internos): lista
        try:
            n_sin = len(sin_id_keys) if 'sin_id_keys' in locals() else 0
            control_rows.append({"Sección":"SIN_ID","Campo":"Total SIN ID","Valor":n_sin})
            if 'sin_id_keys' in locals() and n_sin>0:
                # listar hasta 50 para no hacer el archivo enorme
                for k in list(sorted(sin_id_keys))[:50]:
                    nombre = str(k).replace("NOMBRE::","")
                    control_rows.append({"Sección":"SIN_ID","Campo":"Nombre","Valor":nombre})

                if n_sin>50:
                    control_rows.append({"Sección":"SIN_ID","Campo":"Nota","Valor":"(Lista truncada a 50)"})
        except Exception:
            pass
        df_control = pd.DataFrame(control_rows)
    except Exception:
        df_control = pd.DataFrame()
    if df_control is not None and len(df_control)>0:
        extra_sheets["CONTROL"] = df_control
    # Preparar hojas extra por archivo (PROCESADO vs IDGRUPO)
    extra_sheets_proc = {k: (transform_sheet_procesado(v) if hasattr(v, "columns") else v) for k, v in extra_sheets.items()}
    extra_sheets_idg = {k: (transform_sheet_idgrupo(v, cfg, _idgrupo_of) if hasattr(v, "columns") else v) for k, v in extra_sheets.items()}

    # Opcional: crear hojas por grupo (solo cuando exista mapeo empleado->grupo)
    try:
        if bool(getattr(cfg, "excel_idgrupo_split_by_group", False)):
            mapa = (cfg.empleado_a_grupo or {})
            if mapa:
                for g in (cfg.grupos_orden or []):
                    if "ID" not in df_out_export.columns:
                        break
                    mask = df_out_export["ID"].astype(str).map(lambda x: mapa.get(str(x).strip(), "") == g)
                    sub = df_out_export.loc[mask].copy()
                    if sub is None or len(sub) == 0:
                        continue
                    sheet_name = f"GRUPO_{g}"
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31]
                    extra_sheets_idg[sheet_name] = transform_sheet_idgrupo(sub, cfg, _idgrupo_of)
    except Exception:
        pass
    
    # Hoja IDGRUPO (lista visual): catálogo para nómina / prefijos.
    # Esta hoja es SOLO para visualización / listas desplegables en Excel; no afecta cálculos.
    # Formato (una sola columna):
    #   000-03
    #   F-110
    #   FT-124
    #   000-04
    #   ...
    # NOTA: F/FT NO se calculan; son prefijos/códigos de nómina definidos en mapa_grupos.json.
    try:
        grupos = list(cfg.grupos_orden or [])
        # Si no hay grupos definidos, inferir desde el propio mapeo empleado_a_grupo/idgrupo
        if not grupos:
            try:
                grupos = sorted({str(g).strip() for g in (cfg.empleado_a_grupo or {}).values() if str(g).strip()})
            except Exception:
                grupos = []
        items = []
        # Prefijos globales opcionales (top-level): data['nomina_prefijos'] (lista)
        prefijos_globales = []
        try:
            prefijos_globales = getattr(cfg, 'nomina_prefijos', None) or []
        except Exception:
            prefijos_globales = []
        # Construir lista por grupo, preservando orden.
        for g in grupos:
            g = str(g).strip()
            if not g:
                continue
            items.append(g)
            meta = (cfg.grupos_meta.get(g, {}) or {})
            pref = meta.get('nomina_prefijos') or meta.get('prefijos_nomina') or meta.get('nomina') or []
            if isinstance(pref, str):
                pref = [pref]
            # Si el grupo no define prefijos, usar los globales si existen.
            if (not pref) and prefijos_globales:
                pref = prefijos_globales
            # Normalizar y anexar
            for x in pref:
                sx = str(x).strip()
                if sx:
                    items.append(sx)
        # De-duplicar sin perder orden
        seen = set()
        ordered = []
        for it in items:
            if it not in seen:
                seen.add(it)
                ordered.append(it)
        if ordered:
            extra_sheets_idg['IDGRUPO'] = pd.DataFrame({'IDGRUPO': ordered})
    except Exception:
        pass

    # En el archivo por IDGRUPO, regenerar el resumen semanal de checadas sin columna ID
    try:
        df_chec_sem_idg = crear_resumen_semanal_checadas(df_idgrupo, cfg, modo="IDGRUPO")
        if df_chec_sem_idg is not None and len(df_chec_sem_idg) > 0:
            extra_sheets_idg["RESUM_SEM_CHECADAS"] = df_chec_sem_idg
    except Exception:
        pass
    if dry_run:
        logging.info("[DRY-RUN] Se omite escritura de Excel/auditoría. Salidas propuestas: %s | %s", out1.name, out2.name)
    else:
        backup_if_exists(out1)
        backup_if_exists(out2)
        exportar_excel(df_out_export, out1, extra_sheets=extra_sheets_proc, cfg=cfg)
        exportar_excel(df_idgrupo, out2, extra_sheets=extra_sheets_idg, cfg=cfg)

    # Auditoría avanzada (bundle + índice + firma + rotación)
    try:
        run_meta = {
            "run_id": run_id,
            "started_at": started_at,
            "version": version,
            "entrada": str(in_path),
            "input_sha256": input_sha256,
            "usuario": usuario_editor or "RRHH",
            "modo_seguro": bool(modo_seguro),
            "edicion_interactiva": bool(edicion_interactiva),
            "plantilla_path": str(plantilla_path) if plantilla_path else "",
        }
        if not dry_run:
            guardar_auditoria_bundle(out_dir=Path(base).parent, script_dir=script_dir, audit_log=audit_log, run_meta=run_meta, cfg=cfg)
        else:
            logging.info('[DRY-RUN] Se omite guardar auditoría bundle (run_id=%s)', run_id)
    except Exception:
        log_exception("Fallo al guardar auditoría avanzada (bundle)", extra={"run_id": run_id, "out_dir": str(Path(base).parent)}, level=logging.ERROR)

    # Log simple de ejecución (auditoría)
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = Path(str(base) + f"_LOG_{ts}.txt")
        n_rows = len(df_out)
        n_extra = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("Registros extra ignorados", na=False).sum())
        n_ov = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("Solape NoLaborado con comida/cena", na=False).sum())
        n_intov = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("Solape interno NoLaborado", na=False).sum())
        n_out = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("NoLaborado fuera de jornada", na=False).sum())
        n_com_inc = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("Comida incompleta", na=False).sum())
        n_cen_inc = int(df_out.get("Notas", pd.Series(dtype=str)).astype(str).str.contains("Cena incompleta", na=False).sum())
        with open(log_path, "w", encoding="utf-8") as f:
            f.write("LOG DE EJECUCIÓN - PROCESADOR ASISTENCIAS\n")
            f.write(f"Entrada: {in_path}\n")
            f.write(f"Salida 1: {out1}\n")
            f.write(f"Salida 2: {out2}\n")
            f.write(f"Filas procesadas: {n_rows}\n")
            f.write(f"Con registros extra (>6): {n_extra}\n")
            f.write(f"Con solape NoLaborado vs comida/cena: {n_ov}\n")
            f.write(f"Con solape interno NoLaborado: {n_intov}\n")
            f.write(f"Con NoLaborado fuera de jornada: {n_out}\n")
            f.write(f"Comida incompleta: {n_com_inc}\n")
            f.write(f"Cena incompleta: {n_cen_inc}\n")
            if plantilla is None:
                f.write("Plantilla empleados: NO (faltas no calculadas)\n")
            else:
                f.write("Plantilla empleados: SI (faltas calculadas)\n")
            f.write("\nNotas: Este log es informativo y no altera cálculos.\n")
    except Exception:
        log_exception("Fallo al escribir log de ejecución", extra={"run_id": run_id}, level=logging.WARNING)
    # Verificación final (producción) - --verify
    if verify:
        rep = verify_outputs(out_procesado=out1, out_idgrupo=out2, out_dir=Path(base).parent, script_dir=script_dir, cfg=cfg)
        if not rep.ok:
            log_exception("VERIFY FAIL: salida inválida", extra={"run_id": run_id, "errors": rep.errors}, level=logging.ERROR)
            raise RuntimeError("VERIFY FAIL: " + rep.summary)
        logging.getLogger("procesador").info("VERIFY OK | run_id=%s", run_id)

    return out1, out2


def _safe_input(prompt: str, default: str = "") -> str:
    """EOF-safe input for optional interactive flows."""
    try:
        v = input(prompt)
    except EOFError:
        return default
    v = (v or "").strip()
    return v if v else default

def _is_yes(s: str) -> bool:
    v = (s or "").strip().lower()
    return v in ("s", "si", "sí", "y", "yes", "1", "true")