import os
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from procesador.pipeline import procesar_archivo


SNAP_ALL = Path(__file__).parent / "snapshots" / "golden_all_sheets.json"


def _make_input_excel(path: Path) -> None:
    df = pd.DataFrame(
        {
            "ID": ["3","3","115","115","124","124","999"],
            "Fecha": [
                "2026-01-28","2026-01-29",
                "2026-01-28","2026-01-29",
                "2026-01-28","2026-01-29",
                "2026-01-28",
            ],
            "Nombre": [
                "Empleado A","Empleado A",
                "Empleado B","Empleado B",
                "Empleado C","Empleado C",
                "Empleado X",
            ],
            "Número de pases de la tarjeta": ["1","1","1","1","1","1","1"],
            "Registro": [
                "09:00 13:00 13:30 18:00",
                "09:05 13:05 13:35 18:10",
                "22:00 02:00",
                "22:10 02:05",
                "09:00:30 09:00:30 18:00:15",
                "09:00 13:00",
                "09:00",
            ],
        }
    )
    df.to_excel(path, index=False)


def _make_plantilla(path: Path) -> None:
    df = pd.DataFrame(
        {
            "ID": ["03", "115", "124", "999"],
            "Nombre": ["Empleado A", "Empleado B", "Empleado C", "Empleado X"],
            "Activo": ["SI", "SI", "SI", "SI"],
            "IDGRUPO": ["000", "F", "FT", "000"],
        }
    )
    df.to_excel(path, index=False)


def _sheetnames(xlsx: Path) -> list[str]:
    wb = load_workbook(xlsx, read_only=True)
    return wb.sheetnames


def _df_sheet(xlsx: Path, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(xlsx, sheet_name=sheet, dtype=str).fillna("")
    # normalización suave
    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce").dt.date.astype(str).replace("NaT", "")
    return df



def _sanitize_control(df: pd.DataFrame) -> pd.DataFrame:
    """Hace determinística la hoja CONTROL para tests.

    La hoja CONTROL incluye campos volátiles por corrida (run_id, started_at, hash de entrada).
    En tests de 'golden', los anulamos para no romper determinismo.
    """
    if df is None or len(df) == 0:
        return df
    if not {"Sección", "Campo", "Valor"}.issubset(set(df.columns)):
        return df
    df2 = df.copy()
    mask = (df2["Sección"].astype(str) == "EJECUCION") & (
        df2["Campo"].astype(str).isin(["run_id", "started_at", "input_sha256"])
    )
    df2.loc[mask, "Valor"] = ""
    return df2

def _canonicalize(df: pd.DataFrame) -> dict:
    # Orden de columnas preservado
    cols = list(df.columns)
    # Orden de filas determinístico: por ID/IDGRUPO/Fecha si existen
    sort_cols = []
    for c in ["ID", "IDGRUPO", "Nombre", "Fecha", "Semana", "Mes"]:
        if c in df.columns:
            sort_cols.append(c)
    df2 = df.copy()
    if sort_cols:
        df2 = df2.sort_values(by=sort_cols, kind="stable").reset_index(drop=True)
    # Convertir a lista de filas (dict) preservando orden de columnas
    rows = [{c: row[c] for c in cols} for _, row in df2.iterrows()]
    return {"columns": cols, "row_count": int(len(df2)), "rows": rows}


def _build_snapshot(out_proc: Path, out_idg: Path) -> dict:
    snap = {"PROCESADO": {}, "IDGRUPO": {}}
    for xlsx, key in [(out_proc, "PROCESADO"), (out_idg, "IDGRUPO")]:
        for sh in _sheetnames(xlsx):
            df = _df_sheet(xlsx, sh)
            if sh == "CONTROL":
                df = _sanitize_control(df)
            snap[key][sh] = _canonicalize(df)
    return snap


def test_golden_all_sheets_snapshot(tmp_path: Path):
    inp = tmp_path / "mini_asistencia.xlsx"
    plantilla = tmp_path / "plantilla_empleados.xlsx"
    _make_input_excel(inp)
    _make_plantilla(plantilla)

    out_proc, out_idg = procesar_archivo(
        inp,
        correccion_interactiva=False,
        plantilla_path=str(plantilla),
        edicion_interactiva=False,
        usuario_editor="TEST",
        modo_seguro=True,
    )

    current = _build_snapshot(out_proc, out_idg)

    if os.getenv("UPDATE_SNAPSHOTS") == "1" or not SNAP_ALL.exists():
        SNAP_ALL.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")
        assert True
        return

    expected = json.loads(SNAP_ALL.read_text(encoding="utf-8"))

    # 1) mismas hojas
    assert set(current["PROCESADO"].keys()) == set(expected["PROCESADO"].keys())
    assert set(current["IDGRUPO"].keys()).issuperset(set(expected["IDGRUPO"].keys()))

    # 2) headers + row_count + rows exactas
    # Permitir hojas extra en IDGRUPO; comparar solo lo esperado
    for book, exp_book in expected.items():
        assert book in current
        for sh, exp_payload in exp_book.items():
            assert sh in current[book]
            assert current[book][sh] == exp_payload


def test_contract_rowcounts_and_headers(tmp_path: Path):
    """Contract duro: headers exactos + row count por hoja.
    Útil si no quieres comparar filas completas en cada cambio.
    """
    inp = tmp_path / "mini_asistencia.xlsx"
    plantilla = tmp_path / "plantilla_empleados.xlsx"
    _make_input_excel(inp)
    _make_plantilla(plantilla)

    out_proc, out_idg = procesar_archivo(
        inp,
        correccion_interactiva=False,
        plantilla_path=str(plantilla),
        edicion_interactiva=False,
        usuario_editor="TEST",
        modo_seguro=True,
    )

    current = _build_snapshot(out_proc, out_idg)
    if not SNAP_ALL.exists():
        # si no existe snapshot todavía, se considera aprobado (el test snapshot principal lo creará en UPDATE)
        assert True
        return

    expected = json.loads(SNAP_ALL.read_text(encoding="utf-8"))

    for key in ["PROCESADO", "IDGRUPO"]:
        for sh, meta in expected[key].items():
            assert sh in current[key]
            assert current[key][sh]["columns"] == meta["columns"], f"{key}:{sh} headers difieren"
            if sh != "CONTROL":
                assert current[key][sh]["row_count"] == meta["row_count"], f"{key}:{sh} row_count difiere"