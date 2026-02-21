import os
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from procesador.pipeline import procesar_archivo


SNAP_REAL = Path(__file__).parent / "snapshots" / "real_export_headers.json"


def _make_input_excel(path: Path) -> None:
    df = pd.DataFrame(
        {
            "ID": ["3", "3", "115", "115"],
            "Fecha": ["2026-01-28", "2026-01-29", "2026-01-28", "2026-01-29"],
            "Nombre": ["Empleado A", "Empleado A", "Empleado B", "Empleado B"],
            "Número de pases de la tarjeta": ["1", "1", "1", "1"],
            # string con horas: el parser extrae HH:MM
            "Registro": [
                "09:00 13:00 13:30 18:00",
                "09:05 13:05 13:35 18:10",
                "22:00 02:00",  # cruce medianoche (prueba de robustez)
                "22:10 02:05",
            ],
        }
    )
    df.to_excel(path, index=False)


def _make_plantilla(path: Path) -> None:
    # Incluye IDGRUPO para generar etiquetas 000-03 y F-115
    df = pd.DataFrame(
        {
            "ID": ["03", "115"],
            "Nombre": ["Empleado A", "Empleado B"],
            "Activo": ["SI", "SI"],
            "IDGRUPO": ["000", "F"],
        }
    )
    df.to_excel(path, index=False)


def _read_headers(xlsx: Path) -> dict:
    wb = load_workbook(xlsx)
    out = {}
    for sh in wb.sheetnames:
        ws = wb[sh]
        first_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # normalizar: eliminar None al final
        while first_row and first_row[-1] is None:
            first_row.pop()
        out[sh] = [str(x) for x in first_row if x is not None]
    return out


def test_real_export_headers_snapshot(tmp_path: Path):
    inp = tmp_path / "mini_asistencia.xlsx"
    plantilla = tmp_path / "plantilla_empleados.xlsx"
    _make_input_excel(inp)
    _make_plantilla(plantilla)

    out_proc, out_idg = procesar_archivo(
        inp,
        correccion_interactiva=False,
        plantilla_path=str(plantilla),
        edicion_interactiva=False,
        usuario_editor="TEST",
        modo_seguro=True,
    )

    headers = {"PROCESADO": _read_headers(out_proc), "IDGRUPO": _read_headers(out_idg)}

    if os.getenv("UPDATE_SNAPSHOTS") == "1" or not SNAP_REAL.exists():
        SNAP_REAL.write_text(json.dumps(headers, ensure_ascii=False, indent=2), encoding="utf-8")
        assert True
        return

    expected = json.loads(SNAP_REAL.read_text(encoding="utf-8"))
    # Permitir hojas extra (p.ej. hoja 'IDGRUPO' en el archivo _IDGRUPO)
    for book, exp_book in expected.items():
        assert book in headers
        for sh, exp_cols in exp_book.items():
            assert sh in headers[book]
            assert headers[book][sh] == exp_cols


def _id_from_idgrupo(v: str) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if "-" not in s:
        return s
    return s.split("-")[-1].lstrip("0") or "0"


def test_real_export_only_identifier_diff(tmp_path: Path):
    inp = tmp_path / "mini_asistencia.xlsx"
    plantilla = tmp_path / "plantilla_empleados.xlsx"
    _make_input_excel(inp)
    _make_plantilla(plantilla)

    out_proc, out_idg = procesar_archivo(
        inp,
        correccion_interactiva=False,
        plantilla_path=str(plantilla),
        edicion_interactiva=False,
        usuario_editor="TEST",
        modo_seguro=True,
    )

    # leer todas las hojas comunes
    wb_proc = load_workbook(out_proc)
    wb_idg = load_workbook(out_idg)

    common = [s for s in wb_proc.sheetnames if s in wb_idg.sheetnames]
    assert "Reporte" in common

    for sh in common:
        dfp = pd.read_excel(out_proc, sheet_name=sh, dtype=str).fillna("")
        dfi = pd.read_excel(out_idg, sheet_name=sh, dtype=str).fillna("")

        # Contrato: procesado usa ID, idgrupo usa IDGRUPO
        if "ID" in dfp.columns:
            dfp = dfp.drop(columns=["IDGRUPO"], errors="ignore")  # por seguridad
        if "IDGRUPO" in dfi.columns:
            dfi = dfi.drop(columns=["ID"], errors="ignore")

        # Si ambos tienen identificador, removemos el identificador y comparamos resto
        if "ID" in dfp.columns and "IDGRUPO" in dfi.columns:
            dfp_cmp = dfp.drop(columns=["ID"], errors="ignore")
            dfi_cmp = dfi.drop(columns=["IDGRUPO"], errors="ignore")
        else:
            dfp_cmp = dfp
            dfi_cmp = dfi

        # Comparar por contenido (sin depender de orden de filas)
        # Ordenamos columnas y filas por representación textual
        dfp_cmp = dfp_cmp.reindex(sorted(dfp_cmp.columns), axis=1)
        dfi_cmp = dfi_cmp.reindex(sorted(dfi_cmp.columns), axis=1)

        dfp_cmp = dfp_cmp.sort_values(by=list(dfp_cmp.columns)[:1] if len(dfp_cmp.columns)>0 else None).reset_index(drop=True)
        dfi_cmp = dfi_cmp.sort_values(by=list(dfi_cmp.columns)[:1] if len(dfi_cmp.columns)>0 else None).reset_index(drop=True)

        assert dfp_cmp.equals(dfi_cmp), f"Hoja {sh} difiere más allá del identificador"