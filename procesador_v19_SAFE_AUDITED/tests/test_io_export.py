from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from procesador.io import exportar_excel


def test_exportar_excel_multi_sheets(tmp_path: Path):
    df = pd.DataFrame({"ID": ["001", "002"], "X": [1, 2]})
    extra = {
        "RESUMEN_SEMANAL": pd.DataFrame({"A": [1]}),
        "RESUMEN_MENSUAL": pd.DataFrame({"B": [2]}),
    }
    out = tmp_path / "salida.xlsx"
    exportar_excel(df, out, extra_sheets=extra)

    wb = load_workbook(out)
    assert "Reporte" in wb.sheetnames
    assert "RESUMEN_SEMANAL" in wb.sheetnames
    assert "RESUMEN_MENSUAL" in wb.sheetnames


def test_exportar_excel_sheetname_truncation(tmp_path: Path):
    df = pd.DataFrame({"ID": ["001"], "X": [1]})
    long_name = "A" * 40
    out = tmp_path / "salida.xlsx"
    exportar_excel(df, out, extra_sheets={long_name: pd.DataFrame({"Z": [9]})})

    wb = load_workbook(out)
    # Excel limita a 31 caracteres
    assert ("A" * 31) in wb.sheetnames
