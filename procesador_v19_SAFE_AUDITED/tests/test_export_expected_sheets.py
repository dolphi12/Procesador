from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from procesador.io import exportar_excel


SHEETS = ['Reporte', 'CONTROL', 'DETALLE_FALTAS', 'INCIDENCIAS_RRHH', 'RESUMEN_MENSUAL', 'RESUMEN_SEMANAL', 'RESUM_SEM_CHECADAS']


def test_export_expected_sheetnames(tmp_path: Path):
    df = pd.DataFrame({"ID": ["001"], "X": [1]})
    extras = {k: pd.DataFrame({"Y": [1]}) for k in SHEETS if k != "Reporte"}
    out = tmp_path / "out.xlsx"
    exportar_excel(df, out, extra_sheets=extras)

    wb = load_workbook(out)
    for name in SHEETS:
        assert name in wb.sheetnames
