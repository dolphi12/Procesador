import os
import json
from pathlib import Path

import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from procesador.pipeline import procesar_archivo

SNAP_STYLE = Path(__file__).parent / "snapshots" / "excel_style_snapshot.json"


def _make_input_excel(path: Path) -> None:
    df = pd.DataFrame(
        {
            "ID": ["3","3","115","115","124","124","999"],
            "Fecha": [
                "2026-01-28","2026-01-29",
                "2026-01-28","2026-01-29",
                "2026-01-28","2026-01-29",
                "2026-01-28",
            ],
            "Nombre": [
                "Empleado A","Empleado A",
                "Empleado B","Empleado B",
                "Empleado C","Empleado C",
                "Empleado X",
            ],
            "Número de pases de la tarjeta": ["1","1","1","1","1","1","1"],
            "Registro": [
                "09:00 13:00 13:30 18:00",
                "09:05 13:05 13:35 18:10",
                "22:00 02:00",
                "22:10 02:05",
                "09:00:30 09:00:30 18:00:15",
                "09:00 13:00",
                "09:00",
            ],
        }
    )
    df.to_excel(path, index=False)


def _make_plantilla(path: Path) -> None:
    df = pd.DataFrame(
        {
            "ID": ["03", "115", "124", "999"],
            "Nombre": ["Empleado A", "Empleado B", "Empleado C", "Empleado X"],
            "Activo": ["SI", "SI", "SI", "SI"],
            "IDGRUPO": ["000", "F", "FT", "000"],
        }
    )
    df.to_excel(path, index=False)


def _expected_width(header: str) -> float:
    # Leer la configuración real usada por el pipeline (mapa_grupos.json en el directorio del módulo)
    import procesador
    from procesador.config import cargar_config
    script_dir = Path(procesador.__file__).resolve().parent
    cfg = cargar_config(script_dir)

    h = "" if header is None else str(header)
    if h in (cfg.column_widths or {}):
        return float(cfg.column_widths[h])
    for item in (cfg.column_width_patterns or []):
        try:
            pat = str(item.get("pattern", ""))
            w = float(item.get("width", 0) or 0)
            if pat and re.search(pat, h):
                return float(w)
        except Exception:
            continue
    return float(min(45, max(10, len(h) + 2)))


def _style_snapshot(xlsx: Path) -> dict:
    wb = load_workbook(xlsx)
    out: dict = {}
    for sh in wb.sheetnames:
        ws = wb[sh]
        freeze = str(ws.freeze_panes) if ws.freeze_panes else ""
        autof = str(ws.auto_filter.ref) if ws.auto_filter and ws.auto_filter.ref else ""

        header_bold = True
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            if not (cell.font and cell.font.bold):
                header_bold = False
                break

        id_fmt = ""
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if "ID" in headers:
            id_col = headers.index("ID") + 1
            id_fmt = str(ws.cell(row=1, column=id_col).number_format or "")

        # Contract duro de widths: por header -> width esperado
        for idx, h in enumerate(headers, start=1):
            letter = get_column_letter(idx)
            actual = ws.column_dimensions[letter].width
            exp = _expected_width(h)
            assert actual is not None
            assert round(float(actual), 2) == round(float(exp), 2), f"{xlsx.name}:{sh}:{h} width {actual} != {exp}"

        out[sh] = {
            "freeze_panes": freeze,
            "autofilter": autof,
            "header_bold": header_bold,
            "id_number_format": id_fmt,
        }
    return out


def test_excel_style_snapshot(tmp_path: Path):
    inp = tmp_path / "mini_asistencia.xlsx"
    plantilla = tmp_path / "plantilla_empleados.xlsx"
    _make_input_excel(inp)
    _make_plantilla(plantilla)

    out_proc, out_idg = procesar_archivo(
        inp,
        correccion_interactiva=False,
        plantilla_path=str(plantilla),
        edicion_interactiva=False,
        usuario_editor="TEST",
        modo_seguro=True,
    )

    current = {"PROCESADO": _style_snapshot(out_proc), "IDGRUPO": _style_snapshot(out_idg)}

    if os.getenv("UPDATE_SNAPSHOTS") == "1" or not SNAP_STYLE.exists():
        SNAP_STYLE.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")
        assert True
        return

    expected = json.loads(SNAP_STYLE.read_text(encoding="utf-8"))
    # Permitir hojas extra (p.ej. hoja 'IDGRUPO' en el archivo _IDGRUPO)
    for book, exp_book in expected.items():
        assert book in current
        cur_book = current[book]
        for sh, exp_meta in exp_book.items():
            assert sh in cur_book
            cur_meta = cur_book[sh]
            for k, v in exp_meta.items():
                assert cur_meta.get(k) == v

    for book in ["PROCESADO", "IDGRUPO"]:
        for sh, meta in current[book].items():
            assert meta["freeze_panes"] == "A2"
            assert meta["header_bold"] is True
            assert meta["autofilter"].startswith("A1:")