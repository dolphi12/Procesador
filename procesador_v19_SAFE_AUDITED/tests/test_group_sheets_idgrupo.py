import json
from pathlib import Path
import pandas as pd
import openpyxl
import subprocess

def test_idgrupo_no_group_sheets_by_default(tmp_path: Path):
    # copy project files needed: run from repo root, but we can just validate writer behavior by running cli with config
    # This test assumes default excel.idgrupo_split_by_group is False.
    import shutil
    repo = Path(__file__).resolve().parents[1]
    work = tmp_path/"repo"
    shutil.copytree(repo, work, dirs_exist_ok=True)
    # Create tiny input
    df = pd.DataFrame([{"ID":1,"Fecha":"2026-01-01","Nombre":"A","Número de pases de la tarjeta":1,"Registro":"08:00 12:00 13:00 17:00"}])
    inp = work/"inp.xlsx"
    df.to_excel(inp, index=False)
    # Ensure group mapping exists and grupos_orden exists
    mg = json.loads((work/"procesador"/"mapa_grupos.json").read_text(encoding="utf-8"))
    mg["grupos_orden"]=["010"]
    mg["empleado_a_grupo"]={"1":"010"}
    mg.setdefault("excel",{})["idgrupo_split_by_group"]=False
    (work/"procesador"/"mapa_grupos.json").write_text(json.dumps(mg, ensure_ascii=False, indent=2), encoding="utf-8")
    subprocess.run(["python","-m","procesador.cli","process","--input",str(inp),"--no-interactive"], cwd=str(work), check=True)
    out = work/"inp_IDGRUPO.xlsx"
    wb = openpyxl.load_workbook(out, data_only=True)
    assert not any(s.startswith("GRUPO_") for s in wb.sheetnames)
    assert "IDGRUPO" in wb.sheetnames

